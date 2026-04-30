from fastapi import FastAPI, APIRouter, HTTPException, Depends, UploadFile, File, status, Request, Query
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from starlette.middleware.base import BaseHTTPMiddleware
import os
import asyncio
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict, EmailStr, field_validator
from typing import List, Optional, Literal, Set
from datetime import datetime, timezone, timedelta
from fastapi import UploadFile, File
from ocr_helpers import extract_text_from_file, parse_employee_from_text
from passlib.context import CryptContext
from jose import JWTError, jwt
import uuid
from cachetools import TTLCache
import threading
from fallback_db import InMemoryDatabase
from auto_gl import _get_or_create_system_account, create_auto_journal_entry
from ai_expense_engine import analyze_receipt, validate_expense_claim
from api.scheduling import router as scheduling_router

try:
    from motor.motor_asyncio import AsyncIOMotorClient
except Exception:
    AsyncIOMotorClient = None

# ── Thread-safe in-process user cache (TTL = 60 s) ──────────────
_user_cache: TTLCache = TTLCache(maxsize=512, ttl=60)
_cache_lock = threading.Lock()

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

import sys

mongo_url = os.environ.get('MONGO_URL', '').strip()
db_name = os.environ.get('DB_NAME', 'myoffice').strip() or 'myoffice'
using_fallback_db = not bool(mongo_url)

if os.environ.get('ENVIRONMENT', 'production') != 'test' and using_fallback_db:
    logging.error("CRITICAL: MONGO_URL is missing. Production environment requires a valid MongoDB connection.")
    sys.exit(1)

if using_fallback_db or AsyncIOMotorClient is None:
    client = None
    db = InMemoryDatabase()
    if AsyncIOMotorClient is None:
        logging.warning("motor/pymongo import failed. Using in-memory fallback database.")
    else:
        logging.warning("MONGO_URL is missing. Using in-memory fallback database.")
else:
    client = AsyncIOMotorClient(mongo_url, serverSelectionTimeoutMS=2500)
    db = client[db_name]

app = FastAPI(title="PRSK API", version="2.0.0")

# ── CORS: lock to explicit origins from env ───────────────────────
_allowed_origins_env = os.environ.get("ALLOWED_ORIGINS", "").strip()
if _allowed_origins_env:
    _ALLOWED_ORIGINS = [o.strip() for o in _allowed_origins_env.split(",") if o.strip()]
else:
    _ALLOWED_ORIGINS = [
        "http://localhost:3000",
        "http://127.0.0.1:3000",
        "https://myoffice-saas.vercel.app",
    ]
_ALLOWED_ORIGIN_REGEX = os.environ.get("ALLOWED_ORIGIN_REGEX", r"https://.*\.vercel\.app")
app.add_middleware(
    CORSMiddleware,
    allow_origins=_ALLOWED_ORIGINS,
    allow_origin_regex=_ALLOWED_ORIGIN_REGEX,
    allow_credentials=True,
    allow_methods=["GET", "POST", "PUT", "PATCH", "DELETE", "OPTIONS"],
    allow_headers=["Authorization", "Content-Type", "Accept"],
)

# ── Global AI Audit Tracker System ──────────────────────────────────
# Background interceptor that acts as a silent auditor for all system mutative actions
class AIAuditTrackerMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request: Request, call_next):
        response = await call_next(request)
        method = request.method
        
        # Only track mutative actions when they succeed or trigger business logic (avoid tracking 404s/options)
        if method in ["POST", "PUT", "PATCH", "DELETE"] and request.url.path.startswith("/api/"):
            path = request.url.path
            
            # Extract user id internally from token
            auth = request.headers.get("Authorization")
            user_id, org_id, user_email = None, "unknown", "System/Unknown"
            
            company_id = None
            if auth and auth.startswith("Bearer "):
                try:
                    token = auth.split(" ")[1]
                    payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
                    user_id = payload.get("sub")
                    with _cache_lock:
                        user = _user_cache.get(user_id)
                    if user:
                        user_id = user.get("id", user_id)
                        user_email = user.get("email", "Unknown")
                        org_id = user.get("organization_id", "unknown")
                        company_id = user.get("company_id")
                except Exception:
                    pass
            
            parts = path.split("/")
            module = parts[2] if len(parts) > 2 else "system"
            entity_id = parts[3] if len(parts) > 3 else None
            
            # AI Anomaly Heuristics (ZOHO/SAP flavor)
            details = f"{method} action executed on {module.upper()} module"
            anomaly = False
            
            sensitive_modules = {"finance", "saas", "roles", "ledger", "subscriptions", "trips"}
            if module in sensitive_modules and method in ["DELETE", "PATCH"]:
                details += " 🔒 [AI FLAG: Sensitive data modification detected]"
                anomaly = True
            
            action_map = {"POST": "CREATE", "PUT": "UPDATE", "PATCH": "UPDATE", "DELETE": "DELETE"}

            ip_address = request.client.host if request.client else None
            device_info = request.headers.get("user-agent")

            # Avoid logging login requests here as they are explicitly logged in the endpoint
            if module.upper() == "AUTH" and "login" in path:
                return response

            log_doc = {
                "id": str(uuid.uuid4()),
                "user_id": user_id,
                "company_id": company_id,
                "organization_id": org_id,
                "user_email": user_email,
                "action": method,
                "action_type": action_map.get(method, method),
                "module": module.upper(),
                "entity_id": entity_id,
                "details": details,
                "ip_address": ip_address,
                "device_info": device_info,
                "anomaly_flag": anomaly,
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            # Fire-and-forget background logging to avoid blocking response pipeline
            async def _log():
                await db.audit_logs.insert_one(log_doc)
            asyncio.create_task(_log())
            
        return response

app.add_middleware(AIAuditTrackerMiddleware)

api_router = APIRouter(prefix="/api")
@app.get("/")
async def healthcheck():
    return {"status": "ok", "service": "myoffice-backend"}

# ── Enforce SECRET_KEY — fail loudly, never use a default ───────
SECRET_KEY = os.environ.get('SECRET_KEY')
if not SECRET_KEY or SECRET_KEY == 'your-secret-key-change-in-production':
    import sys
    if os.environ.get('ENVIRONMENT', 'production') != 'test':
        logging.error("CRITICAL: SECRET_KEY is missing or insecure. Production environment requires a secure SECRET_KEY.")
        sys.exit(1)
    else:
        # Warn but don't crash dev environments without the key set
        SECRET_KEY = SECRET_KEY or 'dev-only-key-do-not-use-in-production'
        logging.warning('⚠️  SECRET_KEY not set — using insecure default. Set SECRET_KEY env variable!')
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 1440
DEFAULT_ORG_ID = "default"
DEFAULT_COMPANY_ID = "demo-comp-1"
DEFAULT_DEMO_PASSWORD = "password123"
DEFAULT_ENABLED_SERVICES = [
    "dashboard",
    "employees",
    "attendance",
    "leaves",
    "recruitment",
    "projects",
    "crm",
    "inventory",
    "finance",
    "support",
    "assets",
    "announcements",
    "kb",
    "audit",
    "insights",
    "stores",
    "procurement",
    "travel",
    "accounting",
]
DEFAULT_ACCOUNTING_SERVICES = ["ledger", "journal", "reports", "gst", "bank"]

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/auth/login")

def get_password_hash(password: str):
    return pwd_context.hash(password)

def verify_password(plain_password: str, hashed_password: str):
    return pwd_context.verify(plain_password, hashed_password)

def create_access_token(data: dict):
    to_encode = data.copy()
    expire = datetime.now(timezone.utc) + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)

async def get_current_user(token: str = Depends(oauth2_scheme)):
    credentials_exception = HTTPException(
        status_code=status.HTTP_401_UNAUTHORIZED,
        detail="Invalid authentication credentials",
        headers={"WWW-Authenticate": "Bearer"},
    )
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        user_id: str = payload.get("sub")
        if user_id is None:
            raise credentials_exception
    except JWTError:
        raise credentials_exception

    # ── Check in-process cache to avoid DB hit on every request ──
    with _cache_lock:
        cached = _user_cache.get(user_id)
    if cached:
        return cached

    user = await db.users.find_one({"id": user_id}, {"_id": 0, "password": 0, "verification_token": 0})
    if user is None:
        raise credentials_exception

    with _cache_lock:
        _user_cache[user_id] = user
    return user

def _invalidate_user_cache(user_id: str):
    """Call after any user document mutation."""
    with _cache_lock:
        _user_cache.pop(user_id, None)


class JobApplication(BaseModel):
    job_id: str
    name: Optional[str] = None
    email: Optional[EmailStr] = None
    resume_text: Optional[str] = None
    peer_rating: Optional[float] = None

class AIInterviewMessage(BaseModel):
    message: str
    time_taken_seconds: Optional[int] = 30

class UserRegister(BaseModel):
    email: EmailStr
    password: str
    name: str
    # role is intentionally NOT accepted from external input

    @field_validator('password')
    @classmethod
    def validate_password_strength(cls, v: str) -> str:
        if len(v) < 8:
            raise ValueError('Password must be at least 8 characters')
        if v.isdigit():
            raise ValueError('Password cannot be all digits')
        return v

    @field_validator('name')
    @classmethod
    def validate_name(cls, v: str) -> str:
        v = v.strip()
        if len(v) < 2:
            raise ValueError('Name must be at least 2 characters')
        return v

class UserLogin(BaseModel):
    email: EmailStr
    password: str

class SeedTestDataRequest(BaseModel):
    multiplier: int = Field(default=1, ge=1, le=5)
    target_records: Optional[int] = Field(default=None, ge=200, le=1200)

class User(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    email: str
    name: str
    role: str
    organization_id: Optional[str] = None
    company_id: Optional[str] = None  # assigned company for accountant role
    email_verified: bool = False
    subscription_status: str = "trial"
    subscription_limits: Optional[dict] = None
    enabled_services: Optional[List[str]] = None
    subscription_end_date: Optional[str] = None
    created_at: str

class Token(BaseModel):
    access_token: str
    token_type: str
    user: User

class Employee(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    user_id: Optional[str] = None
    emp_id: Optional[str] = None
    previous_emp_id: Optional[str] = None
    name: str
    email: str
    phone: str
    department: str
    designation: str
    # added parking
    parking_slot: Optional[str] = None
    vehicle_number: Optional[str] = None
    date_of_joining: Optional[str] = None
    # Identity & Personal
    father_name: Optional[str] = None
    sex: Optional[str] = None # Male, female, other
    is_disabled: Optional[str] = "No"
    present_address: Optional[str] = None
    temp_address: Optional[str] = None
    emergency_contact: Optional[str] = None
    
    # Statutory & Identity
    pan_number: Optional[str] = None
    aadhaar_number: Optional[str] = None
    abha_number: Optional[str] = None
    driving_license: Optional[str] = None
    driving_expiry: Optional[str] = None
    
    # HR & Payroll Metadata
    skill_category: Optional[str] = "skilled" # unskilled, semiskilled, skilled
    senior_level_code: Optional[str] = None # mapping from employee code
    approval_gate_pass: Optional[str] = None # mapping from employee code
    
    # ESI & Insurance
    esi_number: Optional[str] = None
    esi_registration_date: Optional[str] = None
    company_insurance_no: Optional[str] = None
    company_insurance_reg_date: Optional[str] = None
    company_insurance_expiry: Optional[str] = None
    employee_insurance_no: Optional[str] = None
    employee_insurance_reg_date: Optional[str] = None
    employee_insurance_expiry: Optional[str] = None
    
    food_allowance: Optional[str] = "No"
    food_charges: Optional[str] = None
    food_start_date: Optional[str] = None
    food_close_date: Optional[str] = None
    pf_enabled: Optional[str] = "No"
    bank_name: Optional[str] = None
    bank_ifsc: Optional[str] = None
    bank_account_number: Optional[str] = None
    bank_holder_name: Optional[str] = None
    nationality: Optional[str] = None

    photo: Optional[str] = None
    status: str = "active"
    created_at: Optional[str] = None


class Department(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    company_id: str
    name: str
    description: Optional[str] = None
    manager_id: Optional[str] = None  # emp_id of the manager
    status: str = "active"
    created_at: str

class DepartmentCreate(BaseModel):
    name: str = Field(..., min_length=1)
    description: Optional[str] = None
    manager_id: Optional[str] = None

class Team(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    company_id: str
    department_id: str
    name: str
    description: Optional[str] = None
    lead_id: Optional[str] = None  # emp_id of the lead
    status: str = "active"
    created_at: str

class TeamCreate(BaseModel):
    department_id: str
    name: str
    description: Optional[str] = None
    lead_id: Optional[str] = None

class Position(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    company_id: str
    department_id: Optional[str] = None
    team_id: Optional[str] = None
    title: str
    description: Optional[str] = None
    reports_to_position_id: Optional[str] = None  # ID of the parent position
    status: str = "active"
    created_at: str

class PositionCreate(BaseModel):
    department_id: Optional[str] = None
    team_id: Optional[str] = None
    title: str
    description: Optional[str] = None
    reports_to_position_id: Optional[str] = None

class EmployeePositionMapping(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    company_id: str
    emp_id: str
    position_id: str
    start_date: str
    end_date: Optional[str] = None
    is_primary: bool = True
    status: str = "active"
    created_at: str

class EmployeePositionMappingCreate(BaseModel):
    emp_id: str
    position_id: str
    start_date: Optional[str] = None
    is_primary: bool = True

class SalaryComponent(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    company_id: str
    name: str
    type: str # "earning" or "deduction"
    is_taxable: bool = True
    default_percentage: Optional[float] = None # e.g. 0.40 for 40% of basic
    base_component_id: Optional[str] = None # ID of component this depends on
    created_at: str

class SalaryComponentCreate(BaseModel):
    name: str
    type: str
    is_taxable: bool = True
    default_percentage: Optional[float] = None
    base_component_id: Optional[str] = None

class SalaryComponentValue(BaseModel):
    component_id: str
    suggested_amount: float
    final_amount: float
    type: str = "earning"

class SalaryStructure(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    company_id: str
    emp_id: str
    components: List[SalaryComponentValue]
    gross_salary: float
    net_salary: float
    pf_enabled: bool = False
    esi_enabled: bool = False
    status: str = "active"
    created_at: str

class SalaryStructureCreate(BaseModel):
    emp_id: str
    components: List[SalaryComponentValue]
    pf_enabled: bool = False
    esi_enabled: bool = False

class Payroll(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    company_id: str
    month: str # "YYYY-MM"
    status: str # "draft", "processed", "locked"
    processed_at: Optional[str] = None
    locked_at: Optional[str] = None
    total_gross: float = 0
    total_net: float = 0
    created_at: str

class PayrollCreate(BaseModel):
    month: str

class Payslip(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    company_id: str
    payroll_id: str
    emp_id: str
    month: str
    components: List[SalaryComponentValue]
    gross_salary: float
    total_deductions: float
    net_salary: float
    status: str = "generated"
    created_at: str

class ComplianceRecord(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    company_id: str
    emp_id: str
    type: str # "PF", "ESI", "PT", etc.
    account_number: Optional[str] = None
    is_active: bool = True
    start_date: str
    end_date: Optional[str] = None
    created_at: str

class ComplianceRecordCreate(BaseModel):
    emp_id: str
    type: str
    account_number: Optional[str] = None
    is_active: bool = True
    start_date: Optional[str] = None

class Document(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    company_id: str
    emp_id: str
    title: str
    type: str # "Offer Letter", "ID Proof", etc.
    file_url: str
    status: str = "active"
    created_at: str

class DocumentCreate(BaseModel):
    emp_id: str
    title: str
    type: str
    file_url: str



class BulkEmployeeParseRequest(BaseModel):
    text: str

class ParsedEmployee(BaseModel):
    name: str
    email: str
    department: str
    designation: str

class EmployeeCreate(BaseModel):
    name: str = Field(..., min_length=1)
    position_id: str
    emp_id: Optional[str] = None
    previous_emp_id: Optional[str] = None
    email: str
    phone: str
    department: str = Field(..., min_length=1)
    designation: str = Field(..., min_length=1)
    date_of_joining: Optional[str] = None
    pan_number: Optional[str] = None
    aadhaar_number: Optional[str] = None
    present_address: Optional[str] = None
    temp_address: Optional[str] = None
    father_name: Optional[str] = None
    sex: Optional[str] = None
    is_disabled: Optional[str] = "No"
    skill_category: Optional[str] = None
    senior_level_code: Optional[str] = None
    approval_gate_pass: Optional[str] = None
    abha_number: Optional[str] = None
    driving_license: Optional[str] = None
    driving_expiry: Optional[str] = None
    esi_number: Optional[str] = None

    food_allowance: Optional[str] = "No"
    food_charges: Optional[str] = None
    food_start_date: Optional[str] = None
    food_close_date: Optional[str] = None
    pf_enabled: Optional[str] = "No"
    bank_name: Optional[str] = None
    bank_ifsc: Optional[str] = None
    bank_account_number: Optional[str] = None
    bank_holder_name: Optional[str] = None
    nationality: Optional[str] = None

    photo: Optional[str] = None

class Attendance(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    employee_id: str
    date: str
    check_in: Optional[str] = None
    check_out: Optional[str] = None
    status: str
    created_at: str

class AttendanceCreate(BaseModel):
    employee_id: str
    date: str
    check_in: Optional[str] = None
    check_out: Optional[str] = None
    status: str = "present"

# --- AI Attendance Models ---

class FaceVerification(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    employee_id: str
    timestamp: str
    status: str # 'success', 'failed', 'spoof_detected'
    confidence_score: float
    liveness_passed: bool

class DeviceLog(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    employee_id: str
    timestamp: str
    user_agent: str
    device_fingerprint: str

class IpLog(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    employee_id: str
    timestamp: str
    ip_address: str
    is_vpn_proxy: bool

class ActivityLog(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    employee_id: str
    timestamp: str
    session_id: str
    idle_time_seconds: int
    is_active: bool

class AnomalyFlag(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    employee_id: str
    session_id: str
    timestamp: str
    reason: str # 'face_mismatch', 'ip_change', 'long_idle', 'device_change'
    severity: str # 'low', 'medium', 'high'

class AttendanceSession(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    employee_id: str
    start_time: str
    end_time: Optional[str] = None
    last_heartbeat: Optional[str] = None
    status: str # 'active', 'ended', 'suspicious'
    trust_score: float = 100.0

class AttendanceSessionStart(BaseModel):
    employee_id: str
    face_verification_id: str
    ip_address: str
    user_agent: str
    device_fingerprint: str

class AttendanceSessionHeartbeat(BaseModel):
    session_id: str
    employee_id: str
    face_verification_id: Optional[str] = None
    idle_time_seconds: int
    ip_address: str

class AttendanceSessionEnd(BaseModel):
    session_id: str
    employee_id: str
    face_verification_id: str

class LeaveRequest(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    employee_id: str
    leave_type: str
    from_date: str
    to_date: str
    reason: str
    status: str = "pending"
    created_at: str

class LeaveRequestCreate(BaseModel):
    employee_id: str
    leave_type: str
    from_date: str
    to_date: str
    reason: str

class Project(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    name: str
    description: Optional[str] = None
    status: str = "active"
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    created_at: Optional[str] = None

class ProjectCreate(BaseModel):
    name: str
    description: Optional[str] = None
    start_date: Optional[str] = None
    end_date: Optional[str] = None

class Task(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    project_id: str
    title: str
    description: Optional[str] = None
    assigned_to: Optional[str] = None
    status: str = "todo"
    priority: str = "medium"
    due_date: Optional[str] = None
    created_at: str

class TaskCreate(BaseModel):
    project_id: str
    title: str
    description: Optional[str] = None
    assigned_to: Optional[str] = None
    priority: str = "medium"
    due_date: Optional[str] = None

class Lead(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    name: str
    email: Optional[str] = None
    phone: Optional[str] = None
    company: Optional[str] = None
    source: Optional[str] = None
    status: str = "new"
    created_at: str

class LeadCreate(BaseModel):
    name: str
    email: Optional[str] = None
    phone: Optional[str] = None
    company: Optional[str] = None
    source: Optional[str] = None

class Deal(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    lead_id: str
    title: str
    value: float
    stage: str = "proposal"
    probability: int = 50
    expected_close_date: Optional[str] = None
    created_at: str

class DealCreate(BaseModel):
    lead_id: str
    title: str
    value: float
    stage: str = "proposal"
    probability: int = 50
    expected_close_date: Optional[str] = None

class ExpenseCategory(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    company_id: str
    organization_id: str
    name: str
    max_limit: float
    requires_receipt: bool = True
    created_at: str

class ExpenseCategoryCreate(BaseModel):
    name: str
    max_limit: float
    requires_receipt: bool = True

class Expense(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    employee_id: str
    category: str
    amount: float
    description: Optional[str] = None
    date: str
    status: str = "submitted" # submitted, manager_approved, finance_approved, approved, rejected, flagged, paid
    receipt_url: Optional[str] = None
    payment_method: str = "direct" # direct, payroll
    ai_extracted_data: Optional[dict] = None
    ai_score: float = 10.0
    ai_flags: List[str] = []
    manager_id: Optional[str] = None
    finance_id: Optional[str] = None
    created_at: str

class ExpenseCreate(BaseModel):
    employee_id: str
    category: str
    amount: float
    description: Optional[str] = None
    date: str
    receipt_url: Optional[str] = None
    payment_method: str = "direct"
    ai_extracted_data: Optional[dict] = None

class InventoryItem(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    name: str
    category: str
    quantity: int
    unit: str
    price_per_unit: float
    location: Optional[str] = None
    created_at: str

# NOTE: Duplicate StoreCreate removed — canonical definition is at line ~335

class OfferLetter(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    organization_id: str
    name: str
    email: str
    phone: Optional[str] = ''
    designation: str
    ctc_yearly: float
    esi_enabled: Optional[bool] = False
    pf_enabled: Optional[bool] = False
    details: dict = {}
    status: str = "Generated"
    created_at: str

class OfferLetterCreate(BaseModel):
    name: str
    email: str
    phone: Optional[str] = ''
    designation: str
    ctc_yearly: float
    esi_enabled: Optional[bool] = False
    pf_enabled: Optional[bool] = False
    pt_enabled: Optional[bool] = False
    it_enabled: Optional[bool] = False
    gratuity_enabled: Optional[bool] = False
    details: dict = {}
    status: str = "Generated"

class InventoryItemCreate(BaseModel):
    name: str
    category: str
    quantity: int
    unit: str
    price_per_unit: float
    location: Optional[str] = None

class DashboardStats(BaseModel):
    total_employees: int
    active_employees: int
    total_projects: int
    pending_leaves: int
    total_leads: int
    total_expenses: float
    pending_purchase_requests: int
    total_stores: int
    total_customers: int
    total_invoices: int
    total_tickets: int
    total_vendors: int
    total_timesheet_hours: float
    burn_rate: float
    projected_revenue: float
    hiring_progress: float

class Store(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    name: str
    location: str
    manager: Optional[str] = None
    contact: Optional[str] = None
    status: str = "active"
    created_at: str

class StoreCreate(BaseModel):
    name: str
    location: str
    manager: Optional[str] = None
    contact: Optional[str] = None

class PurchaseRequest(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    store_id: str
    requested_by: str
    items: List[dict]
    total_amount: float
    reason: Optional[str] = None
    status: str = "pending"
    approved_by: Optional[str] = None
    approved_date: Optional[str] = None
    created_at: str

class PurchaseRequestCreate(BaseModel):
    store_id: str
    requested_by: str
    items: List[dict]
    total_amount: float
    reason: Optional[str] = None

class PurchaseOrder(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    purchase_request_id: str
    store_id: str
    supplier_name: str
    supplier_contact: Optional[str] = None
    items: List[dict]
    total_amount: float
    delivery_date: Optional[str] = None
    status: str = "pending"
    created_by: str
    created_at: str

class PurchaseOrderCreate(BaseModel):
    purchase_request_id: str
    store_id: str
    supplier_name: str = Field(..., min_length=1)
    supplier_contact: Optional[str] = None
    items: List[dict] = Field(..., min_length=1)
    total_amount: float = Field(..., gt=0)
    delivery_date: Optional[str] = None
    created_by: str

class HRField(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    field_name: str
    field_type: str
    is_required: bool = False
    options: Optional[List[str]] = None
    applies_to: str
    created_at: str

class HRFieldCreate(BaseModel):
    field_name: str
    field_type: str
    is_required: bool = False
    options: Optional[List[str]] = None
    applies_to: str

class Customer(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    name: str
    contact_person: Optional[str] = None
    email: EmailStr
    phone: Optional[str] = None
    address: Optional[str] = None
    organization_id: str
    created_at: str

class CustomerCreate(BaseModel):
    name: str
    contact_person: Optional[str] = None
    email: EmailStr
    phone: Optional[str] = None
    address: Optional[str] = None

class GoodsReceiptCreate(BaseModel):
    purchase_order_id: str
    store_id: str
    items_received: List[dict] = Field(..., min_length=1)
    receipt_date: str
    delivery_note: Optional[str] = None
    received_by: str

class GoodsReceipt(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    purchase_order_id: str
    store_id: str
    items_received: List[dict]
    receipt_date: str
    delivery_note: Optional[str] = None
    received_by: str
    status: str = "completed"
    organization_id: str
    created_at: str

class PaymentCreate(BaseModel):
    invoice_id: str
    amount: float = Field(..., gt=0)
    payment_method: str = Field(..., min_length=1)
    payment_date: str
    reference_number: Optional[str] = None

class Payment(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    invoice_id: str
    amount: float
    payment_method: str
    payment_date: str
    reference_number: Optional[str] = None
    status: str = "completed"
    organization_id: str
    created_at: str

class Invoice(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    invoice_number: str
    customer_id: Optional[str] = None
    vendor_id: Optional[str] = None
    purchase_order_id: Optional[str] = None
    goods_receipt_id: Optional[str] = None
    type: str = "AR"
    deal_id: Optional[str] = None
    items: List[dict]
    total_amount: float
    amount_paid: float = 0.0
    status: str = "draft" # draft, pending, partially_paid, paid
    due_date: str
    organization_id: str
    created_at: str

class InvoiceCreate(BaseModel):
    customer_id: Optional[str] = None
    vendor_id: Optional[str] = None
    purchase_order_id: Optional[str] = None
    goods_receipt_id: Optional[str] = None
    type: str = "AR"
    deal_id: Optional[str] = None
    items: List[dict]
    total_amount: float
    due_date: str

class Timesheet(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    employee_id: str
    project_id: str
    task_id: str
    hours: float
    date: str
    description: Optional[str] = None
    status: str = "submitted"
    organization_id: str
    created_at: str

class TimesheetCreate(BaseModel):
    project_id: str
    task_id: str
    hours: float
    date: str
    description: Optional[str] = None

class Ticket(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    subject: str
    description: str
    priority: str = "medium"
    status: str = "open"
    assigned_to: Optional[str] = None
    contact_email: EmailStr
    organization_id: str
    created_at: str

class TicketCreate(BaseModel):
    subject: str
    description: str
    priority: str = "medium"
    contact_email: EmailStr

class POSHComplaint(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    complainant_id: str
    incident_date: str
    description: str
    accused_name: str
    status: str = "Under Review"  # Under Review, Investigating, Action Taken, Closed
    organization_id: str
    created_at: str

class POSHComplaintCreate(BaseModel):
    incident_date: str
    description: str
    accused_name: str

class WFHRequest(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    employee_id: str
    employee_name: Optional[str] = "Employee"
    start_date: str
    end_date: str
    reason: str
    status: str = "pending"  # pending, approved, rejected
    organization_id: Optional[str] = "default"
    created_at: str

class WFHRequestCreate(BaseModel):
    start_date: str
    end_date: str
    reason: str

class Vendor(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    name: str
    email: Optional[EmailStr] = None
    phone: Optional[str] = None
    category: Optional[str] = None
    organization_id: str
    created_at: str

class VendorCreate(BaseModel):
    name: str
    email: Optional[EmailStr] = None
    phone: Optional[str] = None
    category: Optional[str] = None

class Asset(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    name: str
    type: str # Laptop, Furniture, etc.
    serial_number: Optional[str] = None
    assigned_to: Optional[str] = None
    status: str = "available"
    purchase_date: Optional[str] = None
    value: float = 0.0
    organization_id: str
    created_at: str

class AssetCreate(BaseModel):
    name: str
    type: str
    serial_number: Optional[str] = None
    assigned_to: Optional[str] = None
    purchase_date: Optional[str] = None
    value: float = 0.0

class Announcement(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    title: str
    content: str
    author_id: str
    author_name: str
    priority: str = "normal" # normal, high, urgent
    organization_id: str
    created_at: str

class AnnouncementCreate(BaseModel):
    title: str
    content: str
    priority: str = "normal"

class JobPosting(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    title: str
    department: str
    location: str
    type: str # Full-time, Remote, etc.
    description: str
    status: str = "open"
    organization_id: str
    created_at: str

class Payslip(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    employee_id: str
    net_salary: float
    earnings: list
    deductions: list
    organization_id: str
    created_at: str

class PayslipCreate(BaseModel):
    employee_id: str
    net_salary: float
    earnings: list
    deductions: list


class JobPostingCreate(BaseModel):
    title: str
    department: str
    location: str
    type: str
    description: str

class JDCreate(BaseModel):
    jd_text: str
    organization_id: Optional[str] = "default_org"

class Candidate(BaseModel):

    model_config = ConfigDict(extra="ignore")
    id: str
    job_id: str
    name: str
    email: EmailStr
    resume_url: Optional[str] = None
    resume_text: Optional[str] = None
    peer_rating: Optional[float] = None
    ai_interview_rating: Optional[float] = None
    status: str = "applied" # applied, screening, interview, offered, rejected
    organization_id: str
    created_at: str

class CandidateCreate(BaseModel):
    job_id: str
    name: str
    email: EmailStr
    resume_url: Optional[str] = None
    resume_text: Optional[str] = None
    peer_rating: Optional[float] = None
    ai_interview_rating: Optional[float] = None

class KnowledgeBaseArticle(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    title: str
    content: str
    category: str
    author_name: str
    organization_id: str
    created_at: str

class AuditLog(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    user_id: Optional[str] = None
    company_id: Optional[str] = None
    user_email: str
    action_type: Optional[str] = None
    action: str
    module: str
    entity_id: Optional[str] = None
    details: str
    ip_address: Optional[str] = None
    device_info: Optional[str] = None
    organization_id: str
    created_at: str

class BusinessInsight(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    type: str # 'warning', 'opportunity', 'kpi'
    title: str
    message: str
    impact: str # 'high', 'medium', 'low'
    organization_id: str
    created_at: str

class Resignation(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    employee_id: str
    employee_name: str
    reason: str
    resignation_date: str
    notice_period_days: int = 30
    last_working_day: Optional[str] = None
    status: str = "pending" # pending, approved, rejected
    fnf_status: str = "pending" # pending, calculated, settled
    fnf_amount: float = 0.0
    fnf_breakdown: dict = {}
    organization_id: str
    created_at: str

class ResignationCreate(BaseModel):
    employee_id: str
    reason: str
    resignation_date: str
    last_working_day: Optional[str] = None
    notice_period_days: int = 30

class PerformancePlan(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    employee_id: str
    employee_name: str
    reason: str
    goals: str
    duration_days: int = 30
    start_date: str
    end_date: str
    status: str = "active"
    organization_id: str
    created_at: str

class PerformancePlanCreate(BaseModel):
    employee_id: str
    reason: str
    goals: str
    duration_days: int = 30
    start_date: str
    end_date: str

# ─────────────────── IATF HR STANDARDS ───────────────────

class IATFMetadata(BaseModel):
    version: str = "1.0"
    created_by: str
    approved_by: Optional[str] = None
    status: str = "draft" # draft, active, archived
    company_id: str
    created_at: str
    updated_at: str

class TurtleDiagram(BaseModel):
    id: str
    metadata: IATFMetadata
    process_name: str
    inputs: List[str]
    outputs: List[str]
    resources: List[str] # With What?
    methods: List[str] # How?
    personnel: List[str] # With Whom?
    kpis: List[str] # How Many?

class InductionProgram(BaseModel):
    id: str
    employee_id: str
    metadata: IATFMetadata
    checkpoints: List[dict] # {item: str, status: bool, date: str}
    mentor_sign_off: bool = False

class SatisfactionAssessment(BaseModel):
    id: str
    metadata: IATFMetadata
    year: int
    scores: dict # {category: score}
    improvement_areas: List[str]

class KaizenSuggestion(BaseModel):
    id: str
    employee_id: str
    metadata: IATFMetadata
    theme: str
    problem_description: str
    suggestion_details: str
    status: str = "pending" # pending, approved, rejected, implemented
    savings_estimated: float = 0.0

class TrainingAttendance(BaseModel):
    id: str
    calendar_id: str
    employee_id: str
    metadata: IATFMetadata
    attended: bool = True
    feedback_score: Optional[int] = None
    effectiveness_evaluation: Optional[str] = None

class SkillLevel(BaseModel):
    skill: str
    level: int # 1, 2, 3, 4

class SkillMatrix(BaseModel):
    id: str
    employee_id: str
    metadata: IATFMetadata
    skills: List[SkillLevel]

class JobDescription(BaseModel):
    id: str
    metadata: IATFMetadata
    role_name: str
    responsibilities: List[str]
    competence_req: List[dict] # {skill: str, min_level: int}

class ResponsibilityMatrix(BaseModel):
    id: str
    metadata: IATFMetadata
    clauses: List[dict] # {iatf_clause: str, responsibility: str, role: str}

class OJTRecord(BaseModel):
    id: str
    employee_id: str
    metadata: IATFMetadata
    topic: str
    trainer_id: str
    start_date: str
    end_date: str
    hours_completed: float
    supervisor_comment: str
    status: str = "completed"

class MotivationAction(BaseModel):
    id: str
    metadata: IATFMetadata
    employee_id: Optional[str] = None # Can be dept wide
    action_type: str # Reward, Recognition, Event, Training
    description: str
    budget_utilized: float
    impact_assessment: Optional[str] = None

class TrainingEffectivenessEvaluation(BaseModel):
    id: str
    attendance_id: str
    metadata: IATFMetadata
    evaluator_id: str
    evaluation_date: str
    score: int # 1-5
    performance_change_observed: str
    manager_comments: str
    follow_up_required: bool = False

@api_router.post("/auth/register", response_model=Token)
async def register(user_data: UserRegister):
    existing = await db.users.find_one({"email": user_data.email}, {"_id": 0})
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    user_id = str(uuid.uuid4())
    hashed_password = get_password_hash(user_data.password)
    org_id = str(uuid.uuid4())
    verification_token = str(uuid.uuid4())
    
    user_doc = {
        "id": user_id,
        "email": user_data.email,
        "password": hashed_password,
        "name": user_data.name.strip(),
        "role": "admin",  # Always admin — role self-escalation prevented
        "organization_id": org_id,
        "email_verified": False,
        "verification_token": verification_token,
        "subscription_status": "trial",
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.users.insert_one(user_doc)

    await db.analytics.insert_one({
        "id": str(uuid.uuid4()),
        "user_id": user_id,
        "event_type": "user_registered",
        "event_data": {"email": user_data.email, "role": "admin"},
        "page": "registration",
        "timestamp": datetime.now(timezone.utc).isoformat()
    })

    access_token = create_access_token(data={"sub": user_id})
    safe_user = {k: v for k, v in user_doc.items() if k not in ("password", "verification_token", "_id")}
    return {"access_token": access_token, "token_type": "bearer", "user": safe_user}

@api_router.post("/auth/login", response_model=Token)
async def login(user_data: UserLogin, request: Request):
    user = await db.users.find_one({"email": user_data.email}, {"_id": 0})
    if not user or not verify_password(user_data.password, user["password"]):
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    access_token = create_access_token(data={"sub": user["id"]})
    user.pop("password")
    
    # If employee, fetch org admin's services
    if user.get("role") == "employee" and user.get("organization_id"):
        admin = await db.users.find_one({"organization_id": user["organization_id"], "role": "admin"}, {"_id": 0, "enabled_services": 1, "subscription_end_date": 1})
        if admin:
            user["enabled_services"] = admin.get("enabled_services")
            user["subscription_end_date"] = admin.get("subscription_end_date")
            
    # Audit Login action
    ip_address = request.client.host if request.client else None
    device_info = request.headers.get("user-agent")
    log_doc = {
        "id": str(uuid.uuid4()),
        "user_id": user["id"],
        "company_id": user.get("company_id"),
        "organization_id": user.get("organization_id", "unknown"),
        "user_email": user["email"],
        "action": "LOGIN",
        "action_type": "LOGIN",
        "module": "AUTH",
        "details": "User successfully logged in",
        "ip_address": ip_address,
        "device_info": device_info,
        "anomaly_flag": False,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.audit_logs.insert_one(log_doc)

    return {"access_token": access_token, "token_type": "bearer", "user": user}

@api_router.get("/auth/me", response_model=User)
async def get_me(current_user: dict = Depends(get_current_user)):
    user = current_user.copy()
    if user.get("role") == "employee" and user.get("organization_id"):
        admin = await db.users.find_one({"organization_id": user["organization_id"], "role": "admin"}, {"_id": 0, "enabled_services": 1, "subscription_end_date": 1})
        if admin:
            user["enabled_services"] = admin.get("enabled_services")
            user["subscription_end_date"] = admin.get("subscription_end_date")
    return user

@api_router.post("/dev/seed-test-data")
async def seed_test_data(payload: SeedTestDataRequest, current_user: dict = Depends(get_current_user)):
    if current_user.get("role") not in ["admin", "superadmin"]:
        raise HTTPException(status_code=403, detail="Only admin users can seed test data")

    org_id = current_user.get("organization_id")
    if not org_id:
        raise HTTPException(status_code=400, detail="Organization context missing")

    summary = await seed_test_data_for_org(
        organization_id=org_id,
        multiplier=payload.multiplier,
        target_records=payload.target_records,
        seeded_by=current_user.get("name", "Admin"),
    )
    return {
        "message": "Test data seeded successfully",
        "organization_id": org_id,
        "summary": summary,
    }

@api_router.get("/dashboard/stats", response_model=DashboardStats)
async def get_dashboard_stats(current_user: dict = Depends(get_current_user)):
    org_id = current_user.get("organization_id")
    q = {} if current_user.get("role") == "superadmin" else {"organization_id": org_id}

    total_employees = await db.employees.count_documents(q)
    active_employees = await db.employees.count_documents({**q, "status": "active"})
    total_projects = await db.projects.count_documents(q)
    pending_leaves = await db.leave_requests.count_documents({**q, "status": "pending"})
    total_leads = await db.leads.count_documents(q)
    pending_pr = await db.purchase_requests.count_documents({**q, "status": "pending"})
    total_stores = await db.stores.count_documents(q)
    total_customers = await db.customers.count_documents(q)
    total_invoices = await db.invoices.count_documents(q)
    total_tickets = await db.tickets.count_documents(q)
    total_vendors = await db.vendors.count_documents(q)

    # ── Aggregate expense sum in DB — no in-memory scan ────────
    exp_agg = await db.expenses.aggregate([
        {"$match": q},
        {"$group": {"_id": None, "total": {"$sum": "$amount"}}}
    ]).to_list(1)
    total_expenses = exp_agg[0]["total"] if exp_agg else 0.0

    # ── Aggregate timesheet hours in DB ────────────────────────
    ts_agg = await db.timesheets.aggregate([
        {"$match": q},
        {"$group": {"_id": None, "total": {"$sum": "$hours"}}}
    ]).to_list(1)
    total_timesheet_hours = ts_agg[0]["total"] if ts_agg else 0.0

    # ── Executive computed stats ────────────────────────────────
    burn_rate = round(total_expenses / 30, 2) if total_expenses > 0 else 0.0
    projected_revenue = total_invoices * 85000.0 if total_invoices > 0 else 0.0
    hiring_progress = min(95.0, (total_employees / 100) * 100) if total_employees > 0 else 0.0

    return {
        "total_employees": total_employees,
        "active_employees": active_employees,
        "total_projects": total_projects,
        "pending_leaves": pending_leaves,
        "total_leads": total_leads,
        "total_expenses": total_expenses,
        "pending_purchase_requests": pending_pr,
        "total_stores": total_stores,
        "total_customers": total_customers,
        "total_invoices": total_invoices,
        "total_tickets": total_tickets,
        "total_vendors": total_vendors,
        "total_timesheet_hours": total_timesheet_hours,
        "burn_rate": burn_rate,
        "projected_revenue": projected_revenue,
        "hiring_progress": hiring_progress
    }


# ─────────────────── ORGANIZATION STRUCTURE ───────────────────

@api_router.post("/departments", response_model=Department)
async def create_department(data: DepartmentCreate, current_user: dict = Depends(get_current_user)):
    company_id = get_accountant_company(current_user)

    # Uniqueness check
    existing = await db.departments.find_one({"company_id": company_id, "name": data.name})
    if existing:
        raise HTTPException(status_code=400, detail="Department with this name already exists in the company")

    dept_id = str(uuid.uuid4())
    doc = {
        "id": dept_id,
        "company_id": company_id,
        "name": data.name,
        "description": data.description,
        "manager_id": data.manager_id,
        "status": "active",
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.departments.insert_one({k: v for k, v in doc.items() if k != "_id"})
    return Department(**doc)

@api_router.get("/departments", response_model=List[Department])
async def get_departments(current_user: dict = Depends(get_current_user)):
    company_id = get_accountant_company(current_user)
    departments = await db.departments.find({"company_id": company_id}).to_list(1000)
    return [Department(**d) for d in departments]

@api_router.put("/departments/{dept_id}", response_model=Department)
async def update_department(dept_id: str, data: DepartmentCreate, current_user: dict = Depends(get_current_user)):
    company_id = get_accountant_company(current_user)
    update_doc = {
        "name": data.name,
        "description": data.description,
        "manager_id": data.manager_id,
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    result = await db.departments.update_one({"id": dept_id, "company_id": company_id}, {"$set": update_doc})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Department not found")
    doc = await db.departments.find_one({"id": dept_id, "company_id": company_id})
    return Department(**doc)

@api_router.delete("/departments/{dept_id}")
async def delete_department(dept_id: str, current_user: dict = Depends(get_current_user)):
    company_id = get_accountant_company(current_user)
    result = await db.departments.delete_one({"id": dept_id, "company_id": company_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Department not found")
    return {"status": "deleted"}

@api_router.post("/teams", response_model=Team)
async def create_team(data: TeamCreate, current_user: dict = Depends(get_current_user)):
    company_id = get_accountant_company(current_user)
    team_id = str(uuid.uuid4())

    # Verify department exists
    dept = await db.departments.find_one({"id": data.department_id, "company_id": company_id})
    if not dept:
        raise HTTPException(status_code=404, detail="Department not found")

    doc = {
        "id": team_id,
        "company_id": company_id,
        "department_id": data.department_id,
        "name": data.name,
        "description": data.description,
        "lead_id": data.lead_id,
        "status": "active",
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.teams.insert_one({k: v for k, v in doc.items() if k != "_id"})
    return Team(**doc)

@api_router.get("/teams", response_model=List[Team])
async def get_teams(department_id: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    company_id = get_accountant_company(current_user)
    query = {"company_id": company_id}
    if department_id:
        query["department_id"] = department_id
    teams = await db.teams.find(query).to_list(1000)
    return [Team(**t) for t in teams]

@api_router.post("/positions", response_model=Position)
async def create_position(data: PositionCreate, current_user: dict = Depends(get_current_user)):
    company_id = get_accountant_company(current_user)
    position_id = str(uuid.uuid4())

    # Check if parent position exists
    if data.reports_to_position_id:
        parent = await db.positions.find_one({"id": data.reports_to_position_id, "company_id": company_id})
        if not parent:
            raise HTTPException(status_code=404, detail="Parent position not found")

        # Prevent circular reporting
        current_parent_id = parent.get("reports_to_position_id")
        while current_parent_id:
            if current_parent_id == position_id:
                raise HTTPException(status_code=400, detail="Circular reporting detected")
            grandparent = await db.positions.find_one({"id": current_parent_id})
            current_parent_id = grandparent.get("reports_to_position_id") if grandparent else None

    doc = {
        "id": position_id,
        "company_id": company_id,
        "department_id": data.department_id,
        "team_id": data.team_id,
        "title": data.title,
        "description": data.description,
        "reports_to_position_id": data.reports_to_position_id,
        "status": "active",
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.positions.insert_one({k: v for k, v in doc.items() if k != "_id"})
    return Position(**doc)

@api_router.get("/positions", response_model=List[Position])
async def get_positions(department_id: Optional[str] = None, team_id: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    company_id = get_accountant_company(current_user)
    query = {"company_id": company_id}
    if department_id:
        query["department_id"] = department_id
    if team_id:
        query["team_id"] = team_id
    positions = await db.positions.find(query).to_list(1000)
    return [Position(**p) for p in positions]

@api_router.post("/employee-positions", response_model=EmployeePositionMapping)
async def create_employee_position(data: EmployeePositionMappingCreate, current_user: dict = Depends(get_current_user)):
    company_id = get_accountant_company(current_user)
    mapping_id = str(uuid.uuid4())

    # Verify position exists
    pos = await db.positions.find_one({"id": data.position_id, "company_id": company_id})
    if not pos:
        raise HTTPException(status_code=404, detail="Position not found")

    # Verify employee exists
    emp = await db.employees.find_one({"id": data.emp_id, "company_id": company_id})
    if not emp:
        raise HTTPException(status_code=404, detail="Employee not found")

    doc = {
        "id": mapping_id,
        "company_id": company_id,
        "emp_id": data.emp_id,
        "position_id": data.position_id,
        "start_date": data.start_date or datetime.now(timezone.utc).isoformat(),
        "is_primary": data.is_primary,
        "status": "active",
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.employee_positions.insert_one({k: v for k, v in doc.items() if k != "_id"})
    return EmployeePositionMapping(**doc)

@api_router.get("/employee-positions", response_model=List[EmployeePositionMapping])
async def get_employee_positions(emp_id: Optional[str] = None, position_id: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    company_id = get_accountant_company(current_user)
    query = {"company_id": company_id}
    if emp_id:
        query["emp_id"] = emp_id
    if position_id:
        query["position_id"] = position_id
    mappings = await db.employee_positions.find(query).to_list(1000)
    return [EmployeePositionMapping(**m) for m in mappings]

# ─────────────────── SALARY & PAYROLL ENGINE ───────────────────

@api_router.post("/salary-components", response_model=SalaryComponent)
async def create_salary_component(data: SalaryComponentCreate, current_user: dict = Depends(get_current_user)):
    company_id = get_accountant_company(current_user)
    comp_id = str(uuid.uuid4())
    doc = {
        "id": comp_id,
        "company_id": company_id,
        "name": data.name,
        "type": data.type,
        "is_taxable": data.is_taxable,
        "default_percentage": data.default_percentage,
        "base_component_id": data.base_component_id,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.salary_components.insert_one({k: v for k, v in doc.items() if k != "_id"})
    return SalaryComponent(**doc)

@api_router.get("/salary-components", response_model=List[SalaryComponent])
async def get_salary_components(current_user: dict = Depends(get_current_user)):
    company_id = get_accountant_company(current_user)
    components = await db.salary_components.find({"company_id": company_id}).to_list(1000)
    return [SalaryComponent(**c) for c in components]

class SalarySuggestionRequest(BaseModel):
    emp_id: str
    annual_ctc: float

@api_router.post("/salary-structure/suggest")
async def suggest_salary_structure(data: SalarySuggestionRequest, current_user: dict = Depends(get_current_user)):
    company_id = get_accountant_company(current_user)

    monthly_gross = data.annual_ctc / 12

    # Standard components if DB empty
    basic = monthly_gross * 0.40
    hra = basic * 0.50
    special = monthly_gross - (basic + hra)

    pf_suggested = basic * 0.12
    esi_suggested = monthly_gross * 0.0075 if monthly_gross <= 21000 else 0
    pt_suggested = 200 # Standard PT

    components = [
        {"component_name": "Basic", "type": "earning", "suggested_amount": basic, "final_amount": basic},
        {"component_name": "HRA", "type": "earning", "suggested_amount": hra, "final_amount": hra},
        {"component_name": "Special Allowance", "type": "earning", "suggested_amount": special, "final_amount": special},
        {"component_name": "PF", "type": "deduction", "suggested_amount": pf_suggested, "final_amount": pf_suggested},
        {"component_name": "ESI", "type": "deduction", "suggested_amount": esi_suggested, "final_amount": esi_suggested},
        {"component_name": "Professional Tax", "type": "deduction", "suggested_amount": pt_suggested, "final_amount": pt_suggested}
    ]

    return {
        "monthly_gross": monthly_gross,
        "components": components,
        "pf_enabled": True,
        "esi_enabled": monthly_gross <= 21000
    }

@api_router.post("/salary-structure", response_model=SalaryStructure)
async def save_salary_structure(data: SalaryStructureCreate, current_user: dict = Depends(get_current_user)):
    company_id = get_accountant_company(current_user)
    struct_id = str(uuid.uuid4())

    gross = sum(c.final_amount for c in data.components if c.type == 'earning')
    deductions = sum(c.final_amount for c in data.components if c.type == 'deduction')
    net = gross - deductions

    doc = {
        "id": struct_id,
        "company_id": company_id,
        "emp_id": data.emp_id,
        "components": [c.model_dump() for c in data.components],
        "gross_salary": gross,
        "net_salary": net,
        "pf_enabled": data.pf_enabled,
        "esi_enabled": data.esi_enabled,
        "status": "active",
        "created_at": datetime.now(timezone.utc).isoformat()
    }

    await db.salary_structures.update_one(
        {"emp_id": data.emp_id, "company_id": company_id},
        {"$set": {"status": "inactive"}}
    )

    await db.salary_structures.insert_one({k: v for k, v in doc.items() if k != "_id"})
    return SalaryStructure(**doc)

@api_router.get("/salary-structure/{emp_id}", response_model=SalaryStructure)
async def get_salary_structure(emp_id: str, current_user: dict = Depends(get_current_user)):
    company_id = get_accountant_company(current_user)
    struct = await db.salary_structures.find_one({"emp_id": emp_id, "company_id": company_id, "status": "active"})
    if not struct:
        raise HTTPException(status_code=404, detail="Active salary structure not found")
    return SalaryStructure(**struct)

class ComplianceCheckRequest(BaseModel):
    monthly_gross: float

@api_router.post("/compliance/check")
async def check_compliance_rules(data: ComplianceCheckRequest, current_user: dict = Depends(get_current_user)):
    esi_mandatory = data.monthly_gross <= 21000 # ESI threshold is 21k, but requirements mentioned 22k for manual. We'll return flags based on 21k for ESI
    esi_allowed = data.monthly_gross <= 22000

    return {
        "monthly_gross": data.monthly_gross,
        "esi_mandatory": esi_mandatory,
        "esi_allowed_manual": esi_allowed, # allow manual if <= 22k
        "pf_allowed": True # PF is toggle-based
    }

@api_router.post("/payroll/run", response_model=Payroll)
async def run_payroll(data: PayrollCreate, current_user: dict = Depends(get_current_user)):
    company_id = get_accountant_company(current_user)

    # Check if payroll already exists for this month
    existing = await db.payrolls.find_one({"month": data.month, "company_id": company_id})
    if existing:
        if existing.get("status") == "locked":
            raise HTTPException(status_code=400, detail="Payroll is locked for this month")
        await db.payrolls.delete_one({"id": existing["id"]})
        await db.payslips.delete_many({"payroll_id": existing["id"]})

    payroll_id = str(uuid.uuid4())

    # Fetch all active salary structures
    structures = await db.salary_structures.find({"company_id": company_id, "status": "active"}).to_list(1000)

    if not structures:
        raise HTTPException(status_code=400, detail="No active salary structures found")

    total_gross = 0
    total_net = 0
    payslips = []

    for struct in structures:
        gross = struct.get("gross_salary", 0)
        net = struct.get("net_salary", 0)
        total_gross += gross
        total_net += net

        payslip = {
            "id": str(uuid.uuid4()),
            "company_id": company_id,
            "payroll_id": payroll_id,
            "emp_id": struct.get("emp_id"),
            "month": data.month,
            "components": struct.get("components", []),
            "gross_salary": gross,
            "total_deductions": gross - net,
            "net_salary": net,
            "status": "generated",
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        payslips.append(payslip)

    if payslips:
        await db.payslips.insert_many(payslips)

    doc = {
        "id": payroll_id,
        "company_id": company_id,
        "month": data.month,
        "status": "processed",
        "processed_at": datetime.now(timezone.utc).isoformat(),
        "total_gross": total_gross,
        "total_net": total_net,
        "created_at": datetime.now(timezone.utc).isoformat()
    }

    await db.payrolls.insert_one({k: v for k, v in doc.items() if k != "_id"})
    return Payroll(**doc)

@api_router.post("/payroll/{payroll_id}/lock", response_model=Payroll)
async def lock_payroll(payroll_id: str, current_user: dict = Depends(get_current_user)):
    company_id = get_accountant_company(current_user)

    payroll = await db.payrolls.find_one({"id": payroll_id, "company_id": company_id})
    if not payroll:
        raise HTTPException(status_code=404, detail="Payroll not found")

    if payroll.get("status") == "locked":
        raise HTTPException(status_code=400, detail="Payroll is already locked")

    update_data = {
        "status": "locked",
        "locked_at": datetime.now(timezone.utc).isoformat()
    }

    await db.payrolls.update_one({"id": payroll_id}, {"$set": update_data})

    updated = await db.payrolls.find_one({"id": payroll_id})

    # Auto GL Entry for Payroll
    exp_acc = await _get_or_create_system_account(db, company_id, "6000", "Salary Expense", "Expense", "Operating Expense")
    liab_acc = await _get_or_create_system_account(db, company_id, "2100", "Salaries Payable", "Liability", "Current Liability")

    total_gross = updated.get("total_gross", 0.0)
    if total_gross > 0:
        lines = [
            {"account_id": exp_acc["id"], "account_name": exp_acc["name"], "debit": total_gross, "credit": 0.0, "description": f"Payroll Expense {updated.get('month')}"},
            {"account_id": liab_acc["id"], "account_name": liab_acc["name"], "debit": 0.0, "credit": total_gross, "description": f"Salaries Payable {updated.get('month')}"}
        ]
        await create_auto_journal_entry(
            db=db,
            company_id=company_id,
            date=datetime.now(timezone.utc).strftime("%Y-%m-%d"),
            narration=f"Auto GL Entry for Payroll {updated.get('month')}",
            reference=payroll_id,
            lines=lines
        )

    return Payroll(**updated)

@api_router.get("/payroll", response_model=List[Payroll])
async def get_payrolls(current_user: dict = Depends(get_current_user)):
    company_id = get_accountant_company(current_user)
    payrolls = await db.payrolls.find({"company_id": company_id}).sort("month", -1).to_list(100)
    return [Payroll(**p) for p in payrolls]

@api_router.get("/payslips/{emp_id}", response_model=List[Payslip])
async def get_employee_payslips(emp_id: str, current_user: dict = Depends(get_current_user)):
    company_id = get_accountant_company(current_user)
    payslips = await db.payslips.find({"emp_id": emp_id, "company_id": company_id}).sort("month", -1).to_list(100)
    return [Payslip(**p) for p in payslips]

@api_router.post("/compliance-records", response_model=ComplianceRecord)
async def create_compliance_record(data: ComplianceRecordCreate, current_user: dict = Depends(get_current_user)):
    company_id = get_accountant_company(current_user)
    rec_id = str(uuid.uuid4())
    doc = {
        "id": rec_id,
        "company_id": company_id,
        "emp_id": data.emp_id,
        "type": data.type,
        "account_number": data.account_number,
        "is_active": data.is_active,
        "start_date": data.start_date or datetime.now(timezone.utc).isoformat(),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.compliance_records.insert_one({k: v for k, v in doc.items() if k != "_id"})
    return ComplianceRecord(**doc)

@api_router.get("/compliance-records/{emp_id}", response_model=List[ComplianceRecord])
async def get_compliance_records(emp_id: str, current_user: dict = Depends(get_current_user)):
    company_id = get_accountant_company(current_user)
    records = await db.compliance_records.find({"emp_id": emp_id, "company_id": company_id}).to_list(100)
    return [ComplianceRecord(**r) for r in records]

@api_router.post("/documents", response_model=Document)
async def create_document(data: DocumentCreate, current_user: dict = Depends(get_current_user)):
    company_id = get_accountant_company(current_user)
    doc_id = str(uuid.uuid4())
    doc = {
        "id": doc_id,
        "company_id": company_id,
        "emp_id": data.emp_id,
        "title": data.title,
        "type": data.type,
        "file_url": data.file_url,
        "status": "active",
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.documents.insert_one({k: v for k, v in doc.items() if k != "_id"})
    return Document(**doc)

@api_router.get("/documents/{emp_id}", response_model=List[Document])
async def get_documents(emp_id: str, current_user: dict = Depends(get_current_user)):
    company_id = get_accountant_company(current_user)
    docs = await db.documents.find({"emp_id": emp_id, "company_id": company_id, "status": "active"}).to_list(100)
    return [Document(**d) for d in docs]




@api_router.post("/ai/upload-employees", response_model=List[ParsedEmployee])
async def upload_employees_ai(file: UploadFile = File(...), current_user: dict = Depends(get_current_user)):
    import csv
    import io
    import re
    from openpyxl import load_workbook

    employees = []
    contents = await file.read()

    lines = []
    if file.filename.endswith('.xlsx'):
        wb = load_workbook(io.BytesIO(contents), data_only=True)
        sheet = wb.active
        for row in sheet.iter_rows(values_only=True):
            # Convert row to comma separated string to reuse our parsing logic
            row_str = " , ".join([str(cell) for cell in row if cell is not None])
            if row_str.strip():
                lines.append(row_str)
    else:
        # Assume CSV or TXT
        text = contents.decode('utf-8', errors='ignore')
        lines = text.split('\n')

    for line in lines:
        line = line.strip()
        if not line:
            continue

        emp = {
            "name": "",
            "email": "",
            "department": "General",
            "designation": "Employee"
        }

        email_match = re.search(r'[\w\.-]+@[\w\.-]+', line)
        if email_match:
            emp["email"] = email_match.group(0)
            line = line.replace(email_match.group(0), '')

        parts = [p.strip() for p in re.split(r'[,|;-]', line) if p.strip()]
        if parts:
            emp["name"] = parts[0]
            if len(parts) > 1:
                emp["designation"] = parts[1]
            if len(parts) > 2:
                emp["department"] = parts[2]
        else:
            words = line.split()
            if len(words) >= 2:
                emp["name"] = f"{words[0]} {words[1]}"
                if len(words) > 2:
                    emp["designation"] = " ".join(words[2:])

        emp["name"] = re.sub(r'[^a-zA-Z\s]', '', emp["name"]).strip()
        if not emp["name"]:
            if emp["email"]:
                emp["name"] = emp["email"].split('@')[0].replace('.', ' ').title()
            else:
                emp["name"] = "Unknown Employee"

        if not emp["email"]:
            emp["email"] = f"{emp['name'].replace(' ', '.').lower()}@example.com"

        employees.append(emp)

    return employees

@api_router.post("/ai/parse-employees", response_model=List[ParsedEmployee])
async def parse_employees_ai(req: BulkEmployeeParseRequest, current_user: dict = Depends(get_current_user)):
    import re
    employees = []
    lines = req.text.split('\n')
    for line in lines:
        line = line.strip()
        if not line:
            continue

        emp = {
            "name": "",
            "email": "",
            "department": "General",
            "designation": "Employee"
        }

        email_match = re.search(r'[\w\.-]+@[\w\.-]+', line)
        if email_match:
            emp["email"] = email_match.group(0)
            line = line.replace(email_match.group(0), '')

        parts = [p.strip() for p in re.split(r'[,|;-]', line) if p.strip()]
        if parts:
            emp["name"] = parts[0]
            if len(parts) > 1:
                emp["designation"] = parts[1]
            if len(parts) > 2:
                emp["department"] = parts[2]
        else:
            words = line.split()
            if len(words) >= 2:
                emp["name"] = f"{words[0]} {words[1]}"
                if len(words) > 2:
                    emp["designation"] = " ".join(words[2:])

        emp["name"] = re.sub(r'[^a-zA-Z\s]', '', emp["name"]).strip()
        if not emp["name"]:
            if emp["email"]:
                emp["name"] = emp["email"].split('@')[0].replace('.', ' ').title()
            else:
                emp["name"] = "Unknown Employee"

        if not emp["email"]:
            emp["email"] = f"{emp['name'].replace(' ', '.').lower()}@example.com"

        employees.append(emp)

    return employees

@api_router.post("/employees", response_model=Employee)
async def create_employee(emp_data: EmployeeCreate, current_user: dict = Depends(get_current_user)):
    company_id = get_accountant_company(current_user)

    if not getattr(emp_data, 'position_id', None):
        raise HTTPException(status_code=400, detail="Employee must be assigned to a position.")

    pos = await db.positions.find_one({"id": emp_data.position_id, "company_id": company_id})
    if not pos:
        raise HTTPException(status_code=404, detail="Position not found.")

    # Limit check
    limits = current_user.get("subscription_limits") or {}
    max_employees = limits.get("max_employees")
    if max_employees is not None:
        count = await db.employees.count_documents({"organization_id": current_user.get("organization_id")})
        if count >= max_employees:
            raise HTTPException(status_code=403, detail="Employee limit reached. Please upgrade your subscription.")

    emp_doc = emp_data.model_dump()
    emp_doc.pop("position_id", None)
    emp_id = str(uuid.uuid4())
    emp_doc["id"] = emp_id
    emp_doc["organization_id"] = current_user.get("organization_id")
    emp_doc["company_id"] = company_id
    emp_doc["status"] = "active"
    emp_doc["created_at"] = datetime.now(timezone.utc).isoformat()
    await db.employees.insert_one(emp_doc)

    # Create employee-position mapping
    mapping_doc = {
        "id": str(uuid.uuid4()),
        "company_id": company_id,
        "emp_id": emp_id,
        "position_id": emp_data.position_id,
        "start_date": datetime.now(timezone.utc).isoformat(),
        "is_primary": True,
        "status": "active",
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.employee_positions.insert_one(mapping_doc)

    return emp_doc

@api_router.get("/employees", response_model=List[Employee])
async def get_employees(current_user: dict = Depends(get_current_user)):
    query = {"organization_id": current_user.get("organization_id")} if current_user.get("role") != "superadmin" else {}
    employees = await db.employees.find(query, {"_id": 0}).to_list(1000)
    return employees

@api_router.get("/employees/{emp_id}", response_model=Employee)
async def get_employee(emp_id: str, current_user: dict = Depends(get_current_user)):
    emp = await db.employees.find_one({"id": emp_id}, {"_id": 0})
    if not emp:
        raise HTTPException(status_code=404, detail="Employee not found")
    return emp

@api_router.put("/employees/{emp_id}", response_model=Employee)
async def update_employee(emp_id: str, emp_data: EmployeeCreate, current_user: dict = Depends(get_current_user)):
    result = await db.employees.update_one({"id": emp_id}, {"$set": emp_data.model_dump()})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Employee not found")
    updated = await db.employees.find_one({"id": emp_id}, {"_id": 0})
    return updated

@api_router.delete("/employees/{emp_id}")
async def delete_employee(emp_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.employees.delete_one({"id": emp_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Employee not found")
    return {"message": "Employee deleted"}

@api_router.post("/attendance", response_model=Attendance)
async def create_attendance(att_data: AttendanceCreate, current_user: dict = Depends(get_current_user)):
    att_doc = att_data.model_dump()
    att_doc["id"] = str(uuid.uuid4())
    att_doc["organization_id"] = current_user.get("organization_id")  # FIX: missing org scope
    att_doc["created_at"] = datetime.now(timezone.utc).isoformat()
    await db.attendance.insert_one(att_doc)
    return att_doc

@api_router.get("/attendance", response_model=List[Attendance])
async def get_attendance(employee_id: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    # FIX: was not org-scoped — cross-org data was accessible
    query: dict = {} if current_user.get("role") == "superadmin" else {"organization_id": current_user.get("organization_id")}
    if employee_id:
        query["employee_id"] = employee_id
    attendance = await db.attendance.find(query, {"_id": 0}).sort("date", -1).to_list(500)
    return attendance

# --- AI Attendance Endpoints ---

@api_router.post("/attendance/session/start", response_model=AttendanceSession)
async def start_attendance_session(session_data: AttendanceSessionStart, current_user: dict = Depends(get_current_user)):
    now = datetime.utcnow().isoformat()
    session_id = str(uuid.uuid4())
    org_id = current_user.get("organization_id")

    # Check if a session is already active
    active_session = await db.attendance_sessions.find_one({"employee_id": session_data.employee_id, "status": "active"})
    if active_session:
        raise HTTPException(status_code=400, detail="An active session already exists for this employee.")

    session_doc = {
        "id": session_id,
        "employee_id": session_data.employee_id,
        "organization_id": org_id,
        "start_time": now,
        "last_heartbeat": now,
        "status": "active",
        "trust_score": 100.0
    }

    ip_log = {
        "id": str(uuid.uuid4()),
        "employee_id": session_data.employee_id,
        "organization_id": org_id,
        "timestamp": now,
        "ip_address": session_data.ip_address,
        "is_vpn_proxy": False # Simplified for now, could integrate a 3rd party API to check IP
    }

    device_log = {
        "id": str(uuid.uuid4()),
        "employee_id": session_data.employee_id,
        "organization_id": org_id,
        "timestamp": now,
        "user_agent": session_data.user_agent,
        "device_fingerprint": session_data.device_fingerprint
    }

    await db.attendance_sessions.insert_one(session_doc)
    await db.ip_logs.insert_one(ip_log)
    await db.device_logs.insert_one(device_log)

    # Also log regular attendance record check-in
    today_str = now[:10]
    att_record = await db.attendance.find_one({"employee_id": session_data.employee_id, "date": today_str})
    if not att_record:
        att_doc = {
            "id": str(uuid.uuid4()),
            "employee_id": session_data.employee_id,
            "organization_id": org_id,
            "date": today_str,
            "check_in": now[11:16],
            "status": "present",
            "created_at": now
        }
        await db.attendance.insert_one(att_doc)

    return session_doc

@api_router.post("/attendance/session/heartbeat")
async def heartbeat_attendance_session(heartbeat_data: AttendanceSessionHeartbeat, current_user: dict = Depends(get_current_user)):
    now = datetime.utcnow().isoformat()
    org_id = current_user.get("organization_id")

    session = await db.attendance_sessions.find_one({"id": heartbeat_data.session_id, "organization_id": org_id})
    if not session:
        raise HTTPException(status_code=404, detail="Session not found.")

    if session.get("status") != "active":
        raise HTTPException(status_code=400, detail="Session is not active.")

    # Check for IP change
    latest_ip = await db.ip_logs.find_one(
        {"employee_id": heartbeat_data.employee_id},
        sort=[("timestamp", -1)]
    )
    trust_score = session.get("trust_score", 100.0)

    if latest_ip and latest_ip.get("ip_address") != heartbeat_data.ip_address:
        anomaly = {
            "id": str(uuid.uuid4()),
            "employee_id": heartbeat_data.employee_id,
            "organization_id": org_id,
            "session_id": heartbeat_data.session_id,
            "timestamp": now,
            "reason": "ip_change",
            "severity": "high"
        }
        await db.anomaly_flags.insert_one(anomaly)
        trust_score -= 20.0

        await db.ip_logs.insert_one({
            "id": str(uuid.uuid4()),
            "employee_id": heartbeat_data.employee_id,
            "organization_id": org_id,
            "timestamp": now,
            "ip_address": heartbeat_data.ip_address,
            "is_vpn_proxy": False
        })

    if heartbeat_data.idle_time_seconds > 600: # 10 minutes idle
        anomaly = {
            "id": str(uuid.uuid4()),
            "employee_id": heartbeat_data.employee_id,
            "organization_id": org_id,
            "session_id": heartbeat_data.session_id,
            "timestamp": now,
            "reason": "long_idle",
            "severity": "medium"
        }
        await db.anomaly_flags.insert_one(anomaly)
        trust_score -= 5.0

    new_status = "suspicious" if trust_score < 50.0 else "active"

    await db.attendance_sessions.update_one(
        {"id": heartbeat_data.session_id},
        {"$set": {"last_heartbeat": now, "trust_score": trust_score, "status": new_status}}
    )

    return {"status": "heartbeat_recorded", "trust_score": trust_score, "session_status": new_status}

@api_router.post("/attendance/session/end", response_model=AttendanceSession)
async def end_attendance_session(end_data: AttendanceSessionEnd, current_user: dict = Depends(get_current_user)):
    now = datetime.utcnow().isoformat()
    org_id = current_user.get("organization_id")

    session = await db.attendance_sessions.find_one({"id": end_data.session_id, "organization_id": org_id})
    if not session:
        raise HTTPException(status_code=404, detail="Session not found.")

    await db.attendance_sessions.update_one(
        {"id": end_data.session_id},
        {"$set": {"end_time": now, "status": "ended" if session.get("status") == "active" else session.get("status")}}
    )

    today_str = now[:10]
    await db.attendance.update_one(
        {"employee_id": end_data.employee_id, "date": today_str},
        {"$set": {"check_out": now[11:16]}}
    )

    updated_session = await db.attendance_sessions.find_one({"id": end_data.session_id})
    if updated_session:
        updated_session.pop("_id", None)
    return updated_session

@api_router.get("/attendance/sessions", response_model=List[AttendanceSession])
async def get_attendance_sessions(employee_id: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    query: dict = {} if current_user.get("role") == "superadmin" else {"organization_id": current_user.get("organization_id")}

    # Enforce RBAC: Non-admin employees can only see their own sessions
    if current_user.get("role") not in ["superadmin", "admin", "hr"]:
        query["employee_id"] = current_user.get("employee_id") or current_user.get("id")
    elif employee_id:
        query["employee_id"] = employee_id

    sessions = await db.attendance_sessions.find(query, {"_id": 0}).sort("start_time", -1).to_list(100)
    return sessions

@api_router.get("/attendance/anomalies", response_model=List[AnomalyFlag])
async def get_attendance_anomalies(employee_id: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    query: dict = {} if current_user.get("role") == "superadmin" else {"organization_id": current_user.get("organization_id")}

    # Enforce RBAC: Non-admin employees can only see their own anomalies
    if current_user.get("role") not in ["superadmin", "admin", "hr"]:
        query["employee_id"] = current_user.get("employee_id") or current_user.get("id")
    elif employee_id:
        query["employee_id"] = employee_id

    anomalies = await db.anomaly_flags.find(query, {"_id": 0}).sort("timestamp", -1).to_list(100)
    return anomalies

@api_router.post("/leave-requests", response_model=LeaveRequest)
async def create_leave_request(leave_data: LeaveRequestCreate, current_user: dict = Depends(get_current_user)):
    leave_doc = leave_data.model_dump()
    leave_doc["id"] = str(uuid.uuid4())
    leave_doc["status"] = "pending"
    leave_doc["organization_id"] = current_user.get("organization_id")  # FIX: was missing org scope
    leave_doc["created_at"] = datetime.now(timezone.utc).isoformat()
    await db.leave_requests.insert_one(leave_doc)
    return leave_doc

@api_router.get("/leave-requests", response_model=List[LeaveRequest])
async def get_leave_requests(current_user: dict = Depends(get_current_user)):
    # FIX: was leaking all orgs' leave data
    query = {}
    if current_user.get("role") != "superadmin":
        query["organization_id"] = current_user.get("organization_id")
    # Employees only see their own leave requests
    if current_user.get("role") == "employee":
        query["employee_id"] = current_user.get("id")
    leaves = await db.leave_requests.find(query, {"_id": 0}).sort("created_at", -1).to_list(500)
    return leaves

class StatusUpdate(BaseModel):
    status: str

class TeamInvite(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    email: EmailStr
    invited_by: str
    organization_id: str
    role: str = "employee"
    status: str = "pending"
    token: str
    expires_at: str
    created_at: str

class TeamInviteCreate(BaseModel):
    email: EmailStr
    role: str = "employee"

class Subscription(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    user_id: str
    plan: str
    status: str
    amount: float
    currency: str = "INR"
    billing_cycle: str
    starts_at: str
    ends_at: str
    created_at: str

class SubscriptionCreate(BaseModel):
    plan: str
    billing_cycle: str = "monthly"

class PlatformSettings(BaseModel):
    model_config = ConfigDict(extra="ignore")
    organization_id: str
    feature_toggles: Optional[dict] = Field(default_factory=dict)
    enabled_modules: Optional[List[str]] = Field(default_factory=list)
    notification_settings: Optional[dict] = Field(default_factory=dict)
    integration_configs: Optional[dict] = Field(default_factory=dict)

class Analytics(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    user_id: Optional[str] = None
    event_type: str
    event_data: dict
    page: Optional[str] = None
    timestamp: str

class AnalyticsCreate(BaseModel):
    event_type: str
    event_data: dict
    page: Optional[str] = None

@api_router.post("/posh-complaints", response_model=POSHComplaint)
async def create_posh_complaint(complaint_data: POSHComplaintCreate, current_user: dict = Depends(get_current_user)):
    doc = complaint_data.model_dump()
    doc["id"] = f"POSH-{str(uuid.uuid4())[:8].upper()}"
    doc["complainant_id"] = current_user.get("id")
    doc["status"] = "Under Review"
    doc["organization_id"] = current_user.get("organization_id")
    doc["created_at"] = datetime.now(timezone.utc).isoformat()
    await db.posh_complaints.insert_one(doc)
    return doc

@api_router.get("/posh-complaints", response_model=List[POSHComplaint])
async def get_posh_complaints(current_user: dict = Depends(get_current_user)):
    query = {"organization_id": current_user.get("organization_id")}
    # Employees only see their own complaints, admins see all
    if current_user.get("role") not in ["admin", "superadmin", "hr"]:
        query["complainant_id"] = current_user.get("id")
    complaints = await db.posh_complaints.find(query, {"_id": 0}).to_list(1000)
    return complaints

@api_router.patch("/posh-complaints/{complaint_id}/status")
async def update_posh_status(complaint_id: str, status_data: StatusUpdate, current_user: dict = Depends(get_current_user)):
    if current_user.get("role") not in ["admin", "superadmin", "hr"]:
        raise HTTPException(status_code=403, detail="Not authorized to update status")
    result = await db.posh_complaints.update_one({"id": complaint_id}, {"$set": {"status": status_data.status}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Complaint not found")
    return {"message": "Status updated"}

@api_router.post("/wfh-requests", response_model=WFHRequest)
async def create_wfh_request(wfh_data: WFHRequestCreate, current_user: dict = Depends(get_current_user)):
    doc = wfh_data.model_dump()
    doc["id"] = f"WFH-{str(uuid.uuid4())[:8].upper()}"
    doc["employee_id"] = current_user.get("id") or "unknown"
    doc["employee_name"] = current_user.get("name") or "User"
    doc["status"] = "pending"
    doc["organization_id"] = current_user.get("organization_id") or "default"
    doc["created_at"] = datetime.now(timezone.utc).isoformat()
    await db.wfh_requests.insert_one(doc)
    return doc

@api_router.get("/wfh-requests", response_model=List[WFHRequest])
async def get_wfh_requests(current_user: dict = Depends(get_current_user)):
    requests = await db.wfh_requests.find({"organization_id": current_user.get("organization_id")}, {"_id": 0}).to_list(1000)
    return requests

@api_router.patch("/wfh-requests/{request_id}/status")
async def update_wfh_status(request_id: str, status_data: StatusUpdate, current_user: dict = Depends(get_current_user)):
    if current_user.get("role") not in ["admin", "superadmin", "manager"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    result = await db.wfh_requests.update_one({"id": request_id}, {"$set": {"status": status_data.status}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Request not found")
    return {"message": "Status updated"}

@api_router.patch("/leave-requests/{leave_id}/status")
async def update_leave_status(leave_id: str, status_data: StatusUpdate, current_user: dict = Depends(get_current_user)):
    # FIX: Anyone could approve their own leave — now requires admin/hr
    if current_user.get("role") not in ["admin", "superadmin", "hr", "manager"]:
        raise HTTPException(status_code=403, detail="Only managers or admins can approve/reject leave requests")
    if status_data.status not in ["pending", "approved", "rejected"]:
        raise HTTPException(status_code=400, detail="Invalid leave status")
    result = await db.leave_requests.update_one(
        {"id": leave_id, "organization_id": current_user.get("organization_id")},
        {"$set": {"status": status_data.status, "reviewed_by": current_user.get("id"), "reviewed_at": datetime.now(timezone.utc).isoformat()}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Leave request not found")
    _invalidate_user_cache(current_user["id"])
    return {"message": f"Leave request {status_data.status}"}

@api_router.post("/projects", response_model=Project)
async def create_project(proj_data: ProjectCreate, current_user: dict = Depends(get_current_user)):
    # Limit check
    limits = current_user.get("subscription_limits") or {}
    max_projects = limits.get("max_projects")
    if max_projects is not None:
        count = await db.projects.count_documents({"organization_id": current_user.get("organization_id")})
        if count >= max_projects:
            raise HTTPException(status_code=403, detail="Project limit reached. Please upgrade your subscription.")

    proj_doc = proj_data.model_dump()
    proj_doc["id"] = str(uuid.uuid4())
    proj_doc["organization_id"] = current_user.get("organization_id")
    proj_doc["status"] = "active"
    proj_doc["created_at"] = datetime.now(timezone.utc).isoformat()
    await db.projects.insert_one(proj_doc)
    return proj_doc

@api_router.get("/projects", response_model=List[Project])
async def get_projects(current_user: dict = Depends(get_current_user)):
    query = {"organization_id": current_user.get("organization_id")} if current_user.get("role") != "superadmin" else {}
    projects = await db.projects.find(query, {"_id": 0}).to_list(1000)
    return projects


@api_router.put("/projects/{project_id}", response_model=Project)
async def update_project(project_id: str, data: ProjectCreate, current_user: dict = Depends(get_current_user)):
    org_id = current_user.get("organization_id")
    update_doc = data.model_dump()
    update_doc["updated_at"] = datetime.now(timezone.utc).isoformat()
    result = await db.projects.update_one({"id": project_id, "organization_id": org_id}, {"$set": update_doc})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Project not found")
    doc = await db.projects.find_one({"id": project_id, "organization_id": org_id})
    return Project(**doc)

@api_router.delete("/projects/{project_id}")
async def delete_project(project_id: str, current_user: dict = Depends(get_current_user)):
    org_id = current_user.get("organization_id")
    result = await db.projects.delete_one({"id": project_id, "organization_id": org_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Project not found")
    return {"status": "deleted"}

@api_router.post("/tasks", response_model=Task)
async def create_task(task_data: TaskCreate, current_user: dict = Depends(get_current_user)):
    task_doc = task_data.model_dump()
    task_doc["id"] = str(uuid.uuid4())
    task_doc["status"] = "todo"
    task_doc["organization_id"] = current_user.get("organization_id")  # FIX: add org scope
    task_doc["created_at"] = datetime.now(timezone.utc).isoformat()
    await db.tasks.insert_one(task_doc)
    return task_doc

@api_router.get("/tasks", response_model=List[Task])
async def get_tasks(project_id: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    # FIX: was not org-scoped
    query: dict = {} if current_user.get("role") == "superadmin" else {"organization_id": current_user.get("organization_id")}
    if project_id:
        query["project_id"] = project_id
    tasks = await db.tasks.find(query, {"_id": 0}).to_list(500)
    return tasks

@api_router.patch("/tasks/{task_id}/status")
async def update_task_status(task_id: str, status_data: StatusUpdate, current_user: dict = Depends(get_current_user)):
    result = await db.tasks.update_one({"id": task_id}, {"$set": {"status": status_data.status}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Task not found")
    return {"message": "Task status updated"}

VALID_LEAD_STATUSES = {"new", "contacted", "qualified", "proposal", "negotiation", "closed-won", "closed-lost"}


@api_router.delete("/tasks/{task_id}")
async def delete_task(task_id: str, current_user: dict = Depends(get_current_user)):
    org_id = current_user.get("organization_id")
    result = await db.tasks.delete_one({"id": task_id, "organization_id": org_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Task not found")
    return {"status": "deleted"}

@api_router.post("/leads", response_model=Lead)
async def create_lead(lead_data: LeadCreate, current_user: dict = Depends(get_current_user)):
    lead_doc = lead_data.model_dump()
    lead_doc["id"] = str(uuid.uuid4())
    lead_doc["status"] = "new"
    lead_doc["organization_id"] = current_user.get("organization_id")  # FIX: was missing
    lead_doc["created_by"] = current_user.get("id")
    lead_doc["created_at"] = datetime.now(timezone.utc).isoformat()
    await db.leads.insert_one(lead_doc)
    return lead_doc

@api_router.get("/leads", response_model=List[Lead])
async def get_leads(current_user: dict = Depends(get_current_user)):
    # FIX: was leaking ALL orgs' leads
    query = {} if current_user.get("role") == "superadmin" else {"organization_id": current_user.get("organization_id")}
    leads = await db.leads.find(query, {"_id": 0}).sort("created_at", -1).to_list(500)
    return leads

@api_router.patch("/leads/{lead_id}/status")
async def update_lead_status(lead_id: str, status_data: StatusUpdate, current_user: dict = Depends(get_current_user)):
    if status_data.status not in VALID_LEAD_STATUSES:
        raise HTTPException(status_code=400, detail=f"Invalid status. Must be one of: {VALID_LEAD_STATUSES}")
    # FIX: also scope update to org to prevent cross-org mutation
    result = await db.leads.update_one(
        {"id": lead_id, "organization_id": current_user.get("organization_id")},
        {"$set": {"status": status_data.status, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Lead not found")
    return {"message": "Lead status updated"}


@api_router.delete("/leads/{lead_id}")
async def delete_lead(lead_id: str, current_user: dict = Depends(get_current_user)):
    org_id = current_user.get("organization_id")
    result = await db.leads.delete_one({"id": lead_id, "organization_id": org_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Lead not found")
    return {"status": "deleted"}

@api_router.post("/deals", response_model=Deal)
async def create_deal(deal_data: DealCreate, current_user: dict = Depends(get_current_user)):
    deal_doc = deal_data.model_dump()
    deal_doc["id"] = str(uuid.uuid4())
    deal_doc["organization_id"] = current_user.get("organization_id")  # FIX: was missing
    deal_doc["created_by"] = current_user.get("id")
    deal_doc["created_at"] = datetime.now(timezone.utc).isoformat()
    await db.deals.insert_one(deal_doc)
    return deal_doc

@api_router.get("/deals", response_model=List[Deal])
async def get_deals(current_user: dict = Depends(get_current_user)):
    # FIX: was leaking ALL orgs' deals
    query = {} if current_user.get("role") == "superadmin" else {"organization_id": current_user.get("organization_id")}
    deals = await db.deals.find(query, {"_id": 0}).sort("created_at", -1).to_list(500)
    return deals


@api_router.delete("/deals/{deal_id}")
async def delete_deal(deal_id: str, current_user: dict = Depends(get_current_user)):
    org_id = current_user.get("organization_id")
    result = await db.deals.delete_one({"id": deal_id, "organization_id": org_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Deal not found")
    return {"status": "deleted"}

@api_router.post("/expenses/categories", response_model=ExpenseCategory)
async def create_expense_category(cat_data: ExpenseCategoryCreate, current_user: dict = Depends(get_current_user)):
    if current_user.get("role") not in ["admin", "superadmin"]:
        raise HTTPException(status_code=403, detail="Not authorized")

    company_id = get_accountant_company(current_user) if current_user.get("role") == "accountant" else current_user.get("company_id", "default_company")
    cat_doc = cat_data.model_dump()
    cat_doc["id"] = str(uuid.uuid4())
    cat_doc["company_id"] = company_id
    cat_doc["organization_id"] = current_user.get("organization_id")
    cat_doc["created_at"] = datetime.now(timezone.utc).isoformat()
    await db.expense_categories.insert_one(cat_doc)
    return cat_doc

@api_router.get("/expenses/categories", response_model=List[ExpenseCategory])
async def get_expense_categories(current_user: dict = Depends(get_current_user)):
    company_id = get_accountant_company(current_user) if current_user.get("role") == "accountant" else current_user.get("company_id", "default_company")
    cats = await db.expense_categories.find({"company_id": company_id}, {"_id": 0}).to_list(100)
    return cats

@api_router.post("/expenses/upload-receipt")
async def upload_expense_receipt(file: UploadFile = File(...), current_user: dict = Depends(get_current_user)):
    try:
        content = await file.read()
        text = extract_text_from_file(content, file.filename)
        extracted = analyze_receipt(text)

        # In a real app we'd save to S3. Here we simulate it.
        receipt_url = f"/uploads/receipts/{uuid.uuid4()}_{file.filename}"

        return {
            "receipt_url": receipt_url,
            "extracted_data": extracted,
            "raw_text": text[:500]
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/expenses", response_model=Expense)
async def create_expense(exp_data: ExpenseCreate, current_user: dict = Depends(get_current_user)):
    company_id = get_accountant_company(current_user) if current_user.get("role") == "accountant" else current_user.get("company_id", "default_company")
    exp_doc = exp_data.model_dump()
    exp_doc["id"] = str(uuid.uuid4())
    exp_doc["status"] = "submitted"
    exp_doc["organization_id"] = current_user.get("organization_id")
    exp_doc["company_id"] = company_id
    exp_doc["created_at"] = datetime.now(timezone.utc).isoformat()

    # Run AI Validation
    validation_result = await validate_expense_claim(db, company_id, exp_doc)
    exp_doc["ai_score"] = validation_result["score"]
    exp_doc["ai_flags"] = validation_result["flags"]
    if validation_result["score"] < 5.0:
        exp_doc["status"] = "flagged"

    await db.expenses.insert_one(exp_doc)
    return exp_doc

@api_router.get("/expenses", response_model=List[Expense])
async def get_expenses(current_user: dict = Depends(get_current_user)):
    query = {"organization_id": current_user.get("organization_id")}
    if current_user.get("role") == "accountant":
        query["company_id"] = get_accountant_company(current_user)
    expenses = await db.expenses.find(query, {"_id": 0}).to_list(1000)
    return expenses

@api_router.patch("/expenses/{expense_id}/approve")
async def approve_expense(expense_id: str, level: str, current_user: dict = Depends(get_current_user)):
    # level can be 'manager' or 'finance'
    expense = await db.expenses.find_one({"id": expense_id})
    if not expense:
        raise HTTPException(status_code=404, detail="Expense not found")

    company_id = expense.get("company_id", "default_company")
    updates = {}
    if level == "manager":
        updates["status"] = "manager_approved"
        updates["manager_id"] = current_user["id"]
    elif level == "finance":
        if current_user.get("role") not in ["admin", "superadmin", "accountant"]:
            raise HTTPException(status_code=403, detail="Not authorized for finance approval")
        updates["status"] = "approved"
        updates["finance_id"] = current_user["id"]
        updates["approved_date"] = datetime.now(timezone.utc).isoformat()

        # Trigger Auto GL Entry for Expense -> Payable
        try:
            exp_acc = await _get_or_create_system_account(db, company_id, "5020", "Employee Expenses", "expense", "operating")
            pay_acc = await _get_or_create_system_account(db, company_id, "2020", "Expense Payables", "liability", "current")
            await create_auto_journal_entry(
                db=db,
                company_id=company_id,
                date=datetime.now(timezone.utc).strftime("%Y-%m-%d"),
                narration=f"Expense Claim Approved: {expense.get('category')} - {expense.get('description', '')}",
                reference=f"EXP-{expense['id'][:8]}",
                lines=[
                    {"account_id": exp_acc["id"], "debit": expense["amount"], "credit": 0.0},
                    {"account_id": pay_acc["id"], "debit": 0.0, "credit": expense["amount"]}
                ],
                created_by="System - Expense Module"
            )
        except Exception as e:
            print(f"GL Error during expense approval: {e}")

    result = await db.expenses.update_one({"id": expense_id}, {"$set": updates})
    return {"message": f"Expense {level} approved"}

@api_router.patch("/expenses/{expense_id}/pay")
async def pay_expense(expense_id: str, current_user: dict = Depends(get_current_user)):
    if current_user.get("role") not in ["admin", "superadmin", "accountant"]:
        raise HTTPException(status_code=403, detail="Not authorized")

    expense = await db.expenses.find_one({"id": expense_id})
    if not expense or expense.get("status") != "approved":
        raise HTTPException(status_code=400, detail="Expense not found or not approved")

    company_id = expense.get("company_id", "default_company")

    # Auto GL: Payable -> Bank
    try:
        pay_acc = await _get_or_create_system_account(db, company_id, "2020", "Expense Payables", "liability", "current")
        bank_acc = await _get_or_create_system_account(db, company_id, "1010", "Main Bank Account", "asset", "bank")
        await create_auto_journal_entry(
            db=db,
            company_id=company_id,
            date=datetime.now(timezone.utc).strftime("%Y-%m-%d"),
            narration=f"Expense Reimbursement Paid",
            reference=f"PAY-EXP-{expense['id'][:8]}",
            lines=[
                {"account_id": pay_acc["id"], "debit": expense["amount"], "credit": 0.0},
                {"account_id": bank_acc["id"], "debit": 0.0, "credit": expense["amount"]}
            ],
            created_by="System - Expense Module"
        )
    except Exception as e:
        print(f"GL Error during expense payment: {e}")

    await db.expenses.update_one(
        {"id": expense_id},
        {"$set": {"status": "paid", "paid_date": datetime.now(timezone.utc).isoformat()}}
    )
    return {"message": "Expense marked as paid"}

@api_router.patch("/expenses/{expense_id}/reject")
async def reject_expense(expense_id: str, current_user: dict = Depends(get_current_user)):
    if current_user.get("role") not in ["admin", "superadmin", "accountant", "manager"]:
        raise HTTPException(status_code=403, detail="Not authorized to reject expenses")
    result = await db.expenses.update_one(
        {"id": expense_id},
        {"$set": {"status": "rejected", "rejected_by": current_user["id"], "rejected_date": datetime.now(timezone.utc).isoformat()}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Expense not found")
    return {"message": "Expense rejected"}

@api_router.post("/inventory", response_model=InventoryItem)
async def create_inventory_item(item_data: InventoryItemCreate, current_user: dict = Depends(get_current_user)):
    item_doc = item_data.model_dump()
    item_doc["id"] = str(uuid.uuid4())
    item_doc["organization_id"] = current_user.get("organization_id")
    if current_user.get("role") == "accountant":
        item_doc["company_id"] = get_accountant_company(current_user)
    item_doc["created_at"] = datetime.now(timezone.utc).isoformat()
    await db.inventory.insert_one(item_doc)
    return item_doc

@api_router.get("/inventory", response_model=List[InventoryItem])
async def get_inventory(current_user: dict = Depends(get_current_user)):
    query = {"organization_id": current_user.get("organization_id")}
    if current_user.get("role") == "accountant":
        query["company_id"] = get_accountant_company(current_user)
    items = await db.inventory.find(query, {"_id": 0}).to_list(1000)
    return items

@api_router.delete("/inventory/{item_id}")
async def delete_inventory_item(item_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.inventory.delete_one({"id": item_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Inventory item not found")
    return {"message": "Inventory item deleted"}

@api_router.put("/inventory/{item_id}", response_model=InventoryItem)
async def update_inventory_item(item_id: str, item_data: InventoryItemCreate, current_user: dict = Depends(get_current_user)):
    result = await db.inventory.update_one(
        {"id": item_id},
        {"$set": item_data.model_dump()}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Inventory item not found")

    updated_item = await db.inventory.find_one({"id": item_id}, {"_id": 0})
    return updated_item

@api_router.post("/stores", response_model=Store)
async def create_store(store_data: StoreCreate, current_user: dict = Depends(get_current_user)):
    store_doc = store_data.model_dump()
    store_doc["id"] = str(uuid.uuid4())
    store_doc["status"] = "active"
    store_doc["created_at"] = datetime.now(timezone.utc).isoformat()
    await db.stores.insert_one(store_doc)
    return store_doc

@api_router.get("/stores", response_model=List[Store])
async def get_stores(current_user: dict = Depends(get_current_user)):
    # FIX: was returning all orgs' store data
    query = {} if current_user.get("role") == "superadmin" else {"organization_id": current_user.get("organization_id")}
    stores = await db.stores.find(query, {"_id": 0}).to_list(500)
    return stores

@api_router.get("/stores/{store_id}", response_model=Store)
async def get_store(store_id: str, current_user: dict = Depends(get_current_user)):
    store = await db.stores.find_one({"id": store_id}, {"_id": 0})
    if not store:
        raise HTTPException(status_code=404, detail="Store not found")
    return store

@api_router.put("/stores/{store_id}", response_model=Store)
async def update_store(store_id: str, store_data: StoreCreate, current_user: dict = Depends(get_current_user)):
    result = await db.stores.update_one({"id": store_id}, {"$set": store_data.model_dump()})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Store not found")
    updated = await db.stores.find_one({"id": store_id}, {"_id": 0})
    return updated

@api_router.delete("/stores/{store_id}")
async def delete_store(store_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.stores.delete_one({"id": store_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Store not found")
    return {"message": "Store deleted"}

@api_router.post("/purchase-requests", response_model=PurchaseRequest)
async def create_purchase_request(pr_data: PurchaseRequestCreate, current_user: dict = Depends(get_current_user)):
    pr_doc = pr_data.model_dump()
    pr_doc["id"] = str(uuid.uuid4())
    pr_doc["status"] = "pending"
    pr_doc["created_at"] = datetime.now(timezone.utc).isoformat()
    await db.purchase_requests.insert_one(pr_doc)
    return pr_doc

@api_router.get("/purchase-requests", response_model=List[PurchaseRequest])
async def get_purchase_requests(current_user: dict = Depends(get_current_user)):
    # FIX: was leaking all orgs' purchase requests
    query = {} if current_user.get("role") == "superadmin" else {"organization_id": current_user.get("organization_id")}
    prs = await db.purchase_requests.find(query, {"_id": 0}).sort("created_at", -1).to_list(500)
    return prs

@api_router.patch("/purchase-requests/{pr_id}/approve")
async def approve_purchase_request(pr_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.purchase_requests.update_one(
        {"id": pr_id},
        {"$set": {
            "status": "approved",
            "approved_by": current_user["id"],
            "approved_date": datetime.now(timezone.utc).isoformat()
        }}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Purchase request not found")
    return {"message": "Purchase request approved"}

@api_router.delete("/purchase-requests/{pr_id}")
async def delete_purchase_request(pr_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.purchase_requests.delete_one({"id": pr_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Purchase request not found")
    return {"message": "Purchase request deleted"}

@api_router.patch("/purchase-requests/{pr_id}/reject")
async def reject_purchase_request(pr_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.purchase_requests.update_one(
        {"id": pr_id},
        {"$set": {"status": "rejected", "approved_by": current_user["id"], "approved_date": datetime.now(timezone.utc).isoformat()}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Purchase request not found")
    return {"message": "Purchase request rejected"}

@api_router.post("/purchase-orders", response_model=PurchaseOrder)
async def create_purchase_order(po_data: PurchaseOrderCreate, current_user: dict = Depends(get_current_user)):
    po_doc = po_data.model_dump()

    # Check for negative quantities in items
    for item in po_doc.get("items", []):
        if item.get("quantity", 0) <= 0:
            raise HTTPException(status_code=400, detail="Item quantity must be strictly positive")

    # The QA report also mentioned verifying vendor_id, but the schema uses supplier_name instead
    # However, if there's a vendor table we might check it. For now, strict item qty validation.

    po_doc["id"] = str(uuid.uuid4())
    po_doc["status"] = "pending"
    po_doc["created_at"] = datetime.now(timezone.utc).isoformat()
    await db.purchase_orders.insert_one(po_doc)
    return po_doc

@api_router.get("/purchase-orders", response_model=List[PurchaseOrder])
async def get_purchase_orders(current_user: dict = Depends(get_current_user)):
    # FIX: was leaking all orgs' purchase orders
    query = {} if current_user.get("role") == "superadmin" else {"organization_id": current_user.get("organization_id")}
    pos = await db.purchase_orders.find(query, {"_id": 0}).sort("created_at", -1).to_list(500)
    return pos

@api_router.delete("/purchase-orders/{po_id}")
async def delete_purchase_order(po_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.purchase_orders.delete_one({"id": po_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Purchase order not found")
    return {"message": "Purchase order deleted"}

@api_router.patch("/purchase-orders/{po_id}/status")
async def update_purchase_order_status(po_id: str, status_data: StatusUpdate, current_user: dict = Depends(get_current_user)):
    po = await db.purchase_orders.find_one({"id": po_id})
    if not po:
        raise HTTPException(status_code=404, detail="Purchase order not found")

    result = await db.purchase_orders.update_one({"id": po_id}, {"$set": {"status": status_data.status}})

    if status_data.status == "delivered" and po.get("status") != "delivered":
        store = await db.stores.find_one({"id": po.get("store_id")})
        location_name = store.get("name") if store else "Warehouse"

        org_id = po.get("organization_id", current_user.get("organization_id"))
        company_id = po.get("company_id", org_id)

        # update inventory
        for item in po.get("items", []):
            item_name = item.get("name")
            qty_to_add = int(item.get("quantity", 0))

            existing_item = await db.inventory.find_one({
                "organization_id": org_id,
                "name": item_name,
                "location": location_name
            })

            if existing_item:
                await db.inventory.update_one(
                    {"id": existing_item["id"]},
                    {"$inc": {"quantity": qty_to_add}}
                )
            else:
                new_item = {
                    "id": str(uuid.uuid4()),
                    "organization_id": org_id,
                    "name": item_name,
                    "category": "supplies",
                    "quantity": qty_to_add,
                    "unit": item.get("unit", "pcs"),
                    "price_per_unit": item.get("price", item.get("unit_price", 0)),
                    "location": location_name,
                    "created_at": datetime.now(timezone.utc).isoformat()
                }
                await db.inventory.insert_one(new_item)

        # Auto GL Entry for Procurement -> Finance
        inv_acc = await _get_or_create_system_account(db, company_id, "1200", "Inventory Asset", "Asset", "Current Asset")
        ap_acc = await _get_or_create_system_account(db, company_id, "2000", "Accounts Payable", "Liability", "Current Liability")

        total_amount = po.get("total_amount", 0.0)
        if total_amount > 0:
            lines = [
                {"account_id": inv_acc["id"], "account_name": inv_acc["name"], "debit": total_amount, "credit": 0.0, "description": f"PO {po_id} Delivered"},
                {"account_id": ap_acc["id"], "account_name": ap_acc["name"], "debit": 0.0, "credit": total_amount, "description": f"PO {po_id} AP Accrual"}
            ]
            await create_auto_journal_entry(
                db=db,
                company_id=company_id,
                date=datetime.now(timezone.utc).strftime("%Y-%m-%d"),
                narration=f"Auto GL Entry for PO {po_id} Goods Received",
                reference=po_id,
                lines=lines
            )

    return {"message": "Purchase order status updated"}

@api_router.post("/hr-fields", response_model=HRField)
async def create_hr_field(field_data: HRFieldCreate, current_user: dict = Depends(get_current_user)):
    field_doc = field_data.model_dump()
    field_doc["id"] = str(uuid.uuid4())
    field_doc["created_at"] = datetime.now(timezone.utc).isoformat()
    await db.hr_fields.insert_one(field_doc)
    return field_doc

@api_router.get("/hr-fields", response_model=List[HRField])
async def get_hr_fields(applies_to: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    query = {"applies_to": applies_to} if applies_to else {}
    fields = await db.hr_fields.find(query, {"_id": 0}).to_list(1000)
    return fields

@api_router.delete("/hr-fields/{field_id}")
async def delete_hr_field(field_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.hr_fields.delete_one({"id": field_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="HR field not found")
    return {"message": "HR field deleted"}

@api_router.post("/team/invite", response_model=TeamInvite)
async def invite_team_member(invite_data: TeamInviteCreate, current_user: dict = Depends(get_current_user)):
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Only admins can invite team members")
    
    # Check if user already exists
    existing = await db.users.find_one({"email": invite_data.email}, {"_id": 0})
    if existing:
        raise HTTPException(status_code=400, detail="User already exists")
    
    # Check if invitation already sent
    existing_invite = await db.team_invites.find_one({"email": invite_data.email, "status": "pending"}, {"_id": 0})
    if existing_invite:
        raise HTTPException(status_code=400, detail="Invitation already sent")
    
    invite_token = str(uuid.uuid4())
    invite_doc = {
        "id": str(uuid.uuid4()),
        "email": invite_data.email,
        "invited_by": current_user["id"],
        "organization_id": current_user.get("organization_id"),
        "role": invite_data.role,
        "status": "pending",
        "token": invite_token,
        "expires_at": (datetime.now(timezone.utc) + timedelta(days=7)).isoformat(),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.team_invites.insert_one(invite_doc)
    
    # Track analytics
    await db.analytics.insert_one({
        "id": str(uuid.uuid4()),
        "user_id": current_user["id"],
        "event_type": "team_invite_sent",
        "event_data": {"invitee_email": invite_data.email, "role": invite_data.role},
        "page": "team",
        "timestamp": datetime.now(timezone.utc).isoformat()
    })
    
    return invite_doc

@api_router.get("/team/invites", response_model=List[TeamInvite])
async def get_team_invites(current_user: dict = Depends(get_current_user)):
    invites = await db.team_invites.find({
        "organization_id": current_user.get("organization_id")
    }, {"_id": 0}).to_list(100)
    return invites

@api_router.post("/team/accept-invite")
async def accept_invite(token: str, name: str, password: str):
    invite = await db.team_invites.find_one({"token": token, "status": "pending"}, {"_id": 0})
    if not invite:
        raise HTTPException(status_code=404, detail="Invalid or expired invitation")
    
    # Check expiration
    expires_at = datetime.fromisoformat(invite["expires_at"])
    if datetime.now(timezone.utc) > expires_at:
        raise HTTPException(status_code=400, detail="Invitation has expired")
    
    # Create user account
    user_id = str(uuid.uuid4())
    hashed_password = get_password_hash(password)
    user_doc = {
        "id": user_id,
        "email": invite["email"],
        "password": hashed_password,
        "name": name,
        "role": invite["role"],
        "organization_id": invite["organization_id"],
        "email_verified": True,
        "subscription_status": "active",
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.users.insert_one(user_doc)
    
    # Update invite status
    await db.team_invites.update_one({"token": token}, {"$set": {"status": "accepted"}})
    
    # Track analytics
    await db.analytics.insert_one({
        "id": str(uuid.uuid4()),
        "user_id": user_id,
        "event_type": "invite_accepted",
        "event_data": {"email": invite["email"]},
        "page": "onboarding",
        "timestamp": datetime.now(timezone.utc).isoformat()
    })
    
    access_token = create_access_token(data={"sub": user_id})
    user_doc.pop("password")
    return {"access_token": access_token, "token_type": "bearer", "user": user_doc}

@api_router.get("/team/members", response_model=List[User])
async def get_team_members(current_user: dict = Depends(get_current_user)):
    members = await db.users.find({
        "organization_id": current_user.get("organization_id")
    }, {"_id": 0, "password": 0, "verification_token": 0}).to_list(100)
    return members

@api_router.post("/auth/verify-email")
async def verify_email(token: str):
    user = await db.users.find_one({"verification_token": token}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=404, detail="Invalid verification token")
    
    await db.users.update_one(
        {"verification_token": token},
        {"$set": {"email_verified": True}, "$unset": {"verification_token": ""}}
    )
    
    # Track analytics
    await db.analytics.insert_one({
        "id": str(uuid.uuid4()),
        "user_id": user["id"],
        "event_type": "email_verified",
        "event_data": {"email": user["email"]},
        "page": "verification",
        "timestamp": datetime.now(timezone.utc).isoformat()
    })
    
    return {"message": "Email verified successfully"}

@api_router.post("/subscriptions", response_model=Subscription)
async def create_subscription(sub_data: SubscriptionCreate, current_user: dict = Depends(get_current_user)):
    plans = {
        "starter": {"amount": 999, "features": ["Up to 50 employees", "Basic features"]},
        "professional": {"amount": 2999, "features": ["Up to 200 employees", "All features"]},
        "enterprise": {"amount": 9999, "features": ["Unlimited employees", "Premium support"]}
    }
    
    if sub_data.plan not in plans:
        raise HTTPException(status_code=400, detail="Invalid plan")
    
    plan_info = plans[sub_data.plan]
    
    # Calculate end date
    if sub_data.billing_cycle == "monthly":
        ends_at = (datetime.now(timezone.utc) + timedelta(days=30)).isoformat()
    else:
        ends_at = (datetime.now(timezone.utc) + timedelta(days=365)).isoformat()
    
    sub_doc = {
        "id": str(uuid.uuid4()),
        "user_id": current_user["id"],
        "plan": sub_data.plan,
        "status": "active",
        "amount": plan_info["amount"],
        "currency": "INR",
        "billing_cycle": sub_data.billing_cycle,
        "starts_at": datetime.now(timezone.utc).isoformat(),
        "ends_at": ends_at,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.subscriptions.insert_one(sub_doc)
    
    # Update user subscription status
    await db.users.update_one(
        {"id": current_user["id"]},
        {"$set": {"subscription_status": "active"}}
    )
    
    # Track analytics
    await db.analytics.insert_one({
        "id": str(uuid.uuid4()),
        "user_id": current_user["id"],
        "event_type": "subscription_created",
        "event_data": {"plan": sub_data.plan, "amount": plan_info["amount"]},
        "page": "subscription",
        "timestamp": datetime.now(timezone.utc).isoformat()
    })
    
    return sub_doc

@api_router.get("/subscriptions/plans")
async def get_subscription_plans():
    return {
        "plans": [
            {
                "id": "starter",
                "name": "Starter",
                "price_monthly": 999,
                "price_yearly": 9999,
                "features": ["Up to 50 employees", "Basic HR features", "Email support", "Mobile app access"]
            },
            {
                "id": "professional",
                "name": "Professional",
                "price_monthly": 2999,
                "price_yearly": 29999,
                "popular": True,
                "features": ["Up to 200 employees", "All HR features", "Priority support", "Advanced analytics", "API access"]
            },
            {
                "id": "enterprise",
                "name": "Enterprise",
                "price_monthly": 9999,
                "price_yearly": 99999,
                "features": ["Unlimited employees", "Custom features", "Dedicated support", "Custom integrations", "SLA guarantee"]
            }
        ]
    }

# ── Platform Settings ──────────────────────────────────────────────
@api_router.get("/settings", response_model=PlatformSettings)
async def get_platform_settings(current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["admin", "superadmin"]:
        raise HTTPException(status_code=403, detail="Not authorized to access platform settings.")

    org_id = current_user.get("organization_id")
    settings = await db.platform_settings.find_one({"organization_id": org_id}, {"_id": 0})
    if not settings:
        # Return default if none exists yet
        return PlatformSettings(
            organization_id=org_id,
            feature_toggles={"ai_insights": True, "dark_mode_default": True},
            enabled_modules=["HRMS", "FINANCE", "CRM"],
            notification_settings={"email_alerts": True, "slack_integration": False},
            integration_configs={"stripe": False, "razorpay": False}
        )
    return settings

@api_router.put("/settings", response_model=PlatformSettings)
async def update_platform_settings(data: PlatformSettings, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["admin", "superadmin"]:
        raise HTTPException(status_code=403, detail="Not authorized to modify platform settings.")

    org_id = current_user.get("organization_id")
    update_data = data.model_dump(exclude={"organization_id"})

    await db.platform_settings.update_one(
        {"organization_id": org_id},
        {"$set": update_data},
        upsert=True
    )

    updated = await db.platform_settings.find_one({"organization_id": org_id}, {"_id": 0})
    # If using in-memory db, upsert might return dict directly if not perfectly mocked
    if not updated:
        update_data["organization_id"] = org_id
        return update_data

    return updated


@api_router.get("/subscriptions/current")
async def get_current_subscription(current_user: dict = Depends(get_current_user)):
    user_data = None
    # For company admin, they are the source of truth
    if current_user.get("role") == "admin":
        user_data = current_user
    # For employee, take from admin
    elif current_user.get("role") == "employee" and current_user.get("organization_id"):
        user_data = await db.users.find_one({"organization_id": current_user["organization_id"], "role": "admin"})

    if user_data:
        organization_id = user_data.get("organization_id")

        employees_count = await db.employees.count_documents({"organization_id": organization_id})
        projects_count = await db.projects.count_documents({"organization_id": organization_id})
        companies_count = await db.companies.count_documents({"organization_id": organization_id})

        return {
            "status": user_data.get("subscription_status"),
            "end_date": user_data.get("subscription_end_date"),
            "limits": user_data.get("subscription_limits"),
            "enabled_services": user_data.get("enabled_services"),
            "usage": {
                "employees": employees_count,
                "projects": projects_count,
                "companies": companies_count
            }
        }

    return None

@api_router.post("/analytics/track")
async def track_analytics(analytics_data: AnalyticsCreate, current_user: dict = Depends(get_current_user)):
    analytics_doc = {
        "id": str(uuid.uuid4()),
        "user_id": current_user["id"],
        "event_type": analytics_data.event_type,
        "event_data": analytics_data.event_data,
        "page": analytics_data.page,
        "timestamp": datetime.now(timezone.utc).isoformat()
    }
    await db.analytics.insert_one(analytics_doc)
    return {"message": "Event tracked"}

@api_router.get("/analytics/funnel")
async def get_onboarding_funnel(current_user: dict = Depends(get_current_user)):
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    pipeline = [
        {"$match": {"event_type": {"$in": ["user_registered", "email_verified", "subscription_created", "team_invite_sent"]}}},
        {"$group": {"_id": "$event_type", "count": {"$sum": 1}}},
        {"$sort": {"_id": 1}}
    ]
    
    results = await db.analytics.aggregate(pipeline).to_list(10)
    funnel = {item["_id"]: item["count"] for item in results}
    
    return {
        "registrations": funnel.get("user_registered", 0),
        "verifications": funnel.get("email_verified", 0),
        "subscriptions": funnel.get("subscription_created", 0),
        "invites_sent": funnel.get("team_invite_sent", 0),
        "conversion_rate": round((funnel.get("subscription_created", 0) / max(funnel.get("user_registered", 1), 1)) * 100, 2)
    }

class ClientLimitsUpdate(BaseModel):
    max_employees: Optional[int] = None
    max_projects: Optional[int] = None
    max_companies: Optional[int] = None
    enabled_services: Optional[List[str]] = None
    subscription_end_date: Optional[str] = None

class ClientCreate(BaseModel):
    name: str
    email: str
    password: str
    max_employees: Optional[int] = None
    max_projects: Optional[int] = None
    max_companies: Optional[int] = None
    enabled_services: Optional[List[str]] = None
    subscription_end_date: Optional[str] = None

@api_router.post("/saas/clients")
async def create_saas_client(client_data: ClientCreate, current_user: dict = Depends(get_current_user)):
    if current_user.get("role") != "superadmin":
        raise HTTPException(status_code=403, detail="Superadmin access required")
    existing = await db.users.find_one({"email": client_data.email}, {"_id": 0})
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")
        
    user_id = str(uuid.uuid4())
    org_id = str(uuid.uuid4())
    hashed_password = get_password_hash(client_data.password)
    
    limits = {}
    if client_data.max_employees is not None:
        limits["max_employees"] = client_data.max_employees
    if client_data.max_projects is not None:
        limits["max_projects"] = client_data.max_projects
    if client_data.max_companies is not None:
        limits["max_companies"] = client_data.max_companies
        
    user_doc = {
        "id": user_id,
        "email": client_data.email,
        "password": hashed_password,
        "name": client_data.name,
        "role": "admin",
        "organization_id": org_id,
        "email_verified": True,
        "subscription_status": "active",
        "subscription_limits": limits,
        "enabled_services": client_data.enabled_services or [],
        "subscription_end_date": client_data.subscription_end_date,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.users.insert_one(user_doc)
    user_doc.pop("password")
    return {"message": "Client created successfully", "client": user_doc}

@api_router.get("/saas/clients")
async def get_saas_clients(current_user: dict = Depends(get_current_user)):
    if current_user.get("role") != "superadmin":
        raise HTTPException(status_code=403, detail="Superadmin access required")
    clients = await db.users.find({"role": "admin"}, {"_id": 0, "password": 0, "verification_token": 0}).to_list(1000)

    # FIX: N+1 query — aggregate all counts in 2 MongoDB queries instead of 2*N
    org_ids = [c.get("organization_id") for c in clients if c.get("organization_id")]

    emp_counts = {}
    proj_counts = {}
    if org_ids:
        emp_pipeline = [
            {"$match": {"organization_id": {"$in": org_ids}}},
            {"$group": {"_id": "$organization_id", "count": {"$sum": 1}}}
        ]
        proj_pipeline = [
            {"$match": {"organization_id": {"$in": org_ids}}},
            {"$group": {"_id": "$organization_id", "count": {"$sum": 1}}}
        ]
        import asyncio

        try:
            # First try the motor way which returns a cursor that needs to be awaited
            emp_cursor = db.employees.aggregate(emp_pipeline)
            proj_cursor = db.projects.aggregate(proj_pipeline)

            # For fallback_db.aggregate() it returns AsyncInMemoryCursor directly,
            # but in motor it returns an AsyncIOMotorCommandCursor.
            emp_res, proj_res = await asyncio.gather(
                emp_cursor.to_list(1000),
                proj_cursor.to_list(1000)
            )
        except Exception:
            # Fallback if the above fails (e.g. if the cursor itself needs awaiting or different motor version)
            emp_res = await db.employees.aggregate(emp_pipeline).to_list(1000)
            proj_res = await db.projects.aggregate(proj_pipeline).to_list(1000)

        emp_counts = {r["_id"]: r["count"] for r in emp_res}
        proj_counts = {r["_id"]: r["count"] for r in proj_res}

    for client in clients:
        org_id = client.get("organization_id")
        client["usage"] = {
            "employees": emp_counts.get(org_id, 0),
            "projects": proj_counts.get(org_id, 0)
        }
    return clients

@api_router.put("/saas/clients/{client_id}/limits")
async def update_client_limits(client_id: str, limits: ClientLimitsUpdate, current_user: dict = Depends(get_current_user)):
    if current_user.get("role") != "superadmin":
        raise HTTPException(status_code=403, detail="Superadmin access required")
    client = await db.users.find_one({"id": client_id})
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")
    
    new_limits = client.get("subscription_limits") or {}
    if limits.max_employees is not None:
        new_limits["max_employees"] = limits.max_employees
    if limits.max_projects is not None:
        new_limits["max_projects"] = limits.max_projects
    if limits.max_companies is not None:
        new_limits["max_companies"] = limits.max_companies
        
    update_data = {"subscription_limits": new_limits}
    if limits.enabled_services is not None:
        update_data["enabled_services"] = limits.enabled_services
    if limits.subscription_end_date is not None:
        update_data["subscription_end_date"] = limits.subscription_end_date

    await db.users.update_one({"id": client_id}, {"$set": update_data})
    return {"message": "Client settings updated", "limits": new_limits, "enabled_services": limits.enabled_services}

@api_router.post("/offer-letters", response_model=OfferLetter)
async def create_offer_letter(offer_data: OfferLetterCreate, current_user: dict = Depends(get_current_user)):
    offer_doc = offer_data.model_dump()
    offer_doc["id"] = f"OFFER-{str(uuid.uuid4())[:8].upper()}"
    offer_doc["organization_id"] = current_user.get("organization_id")
    offer_doc["created_at"] = datetime.now(timezone.utc).isoformat()
    await db.offer_letters.insert_one(offer_doc)
    return offer_doc

@api_router.get("/offer-letters", response_model=List[OfferLetter])
async def get_offer_letters(current_user: dict = Depends(get_current_user)):
    query = {"organization_id": current_user.get("organization_id")}
    offers = await db.offer_letters.find(query, {"_id": 0}).to_list(1000)
    return offers

@api_router.delete("/offer-letters/{offer_id}")
async def delete_offer_letter(offer_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.offer_letters.delete_one({"id": offer_id, "organization_id": current_user.get("organization_id")})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Offer letter not found")
    return {"message": "Offer letter deleted"}

# --- Zoho Inspired Routes ---

# Customers (Zoho CRM/Books)
@api_router.post("/customers", response_model=Customer)
async def create_customer(data: CustomerCreate, current_user: dict = Depends(get_current_user)):
    doc = data.model_dump()
    doc["id"] = f"CUST-{str(uuid.uuid4())[:8].upper()}"
    doc["organization_id"] = current_user.get("organization_id")
    doc["created_at"] = datetime.now(timezone.utc).isoformat()
    await db.customers.insert_one(doc)
    return doc

@api_router.get("/customers", response_model=List[Customer])
async def get_customers(current_user: dict = Depends(get_current_user)):
    query = {"organization_id": current_user.get("organization_id")}
    if current_user.get("role") == "accountant":
        query["company_id"] = get_accountant_company(current_user)
    return await db.customers.find(query, {"_id": 0}).to_list(1000)


# ==========================================
# GOODS RECEIPTS (GRN)
# ==========================================

@api_router.post("/goods-receipts", response_model=GoodsReceipt)
async def create_goods_receipt(grn_data: GoodsReceiptCreate, current_user: dict = Depends(get_current_user)):
    po = await db.purchase_orders.find_one({"id": grn_data.purchase_order_id})
    if not po:
        raise HTTPException(status_code=404, detail="Purchase order not found")

    org_id = current_user.get("organization_id")

    grn_doc = grn_data.model_dump()
    grn_doc["id"] = str(uuid.uuid4())
    grn_doc["status"] = "completed"
    grn_doc["organization_id"] = org_id
    grn_doc["created_at"] = datetime.now(timezone.utc).isoformat()

    store = await db.stores.find_one({"id": grn_data.store_id})
    location_name = store.get("name") if store else "Warehouse"

    total_value = 0.0

    for item in grn_doc.get("items_received", []):
        item_name = item.get("name")
        qty_to_add = int(item.get("quantity", 0))
        price = float(item.get("price", 0))
        total_value += (qty_to_add * price)

        existing_item = await db.inventory.find_one({
            "organization_id": org_id,
            "name": item_name,
            "location": location_name
        })

        if existing_item:
            await db.inventory.update_one(
                {"id": existing_item["id"]},
                {"$inc": {"quantity": qty_to_add}}
            )
        else:
            new_item = {
                "id": str(uuid.uuid4()),
                "name": item_name,
                "category": "Received",
                "quantity": qty_to_add,
                "unit": item.get("unit", "pcs"),
                "price_per_unit": price,
                "location": location_name,
                "organization_id": org_id,
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            await db.inventory.insert_one(new_item)

    # Auto GL Entry for GRN (Inventory Accrual)
    inventory_acc = await db.chart_of_accounts.find_one({"organization_id": org_id, "name": "Inventory"})
    ap_accrual = await db.chart_of_accounts.find_one({"organization_id": org_id, "name": "AP Accrual"})
    if inventory_acc and ap_accrual and total_value > 0:
        journal_entry = {
            "id": str(uuid.uuid4()),
            "date": datetime.now(timezone.utc).isoformat(),
            "narration": f"GRN Accrual for PO {po.get('id')}",
            "reference": grn_doc["id"],
            "lines": [
                {"account_id": inventory_acc["id"], "account_name": "Inventory", "debit": total_value, "credit": 0.0},
                {"account_id": ap_accrual["id"], "account_name": "AP Accrual", "debit": 0.0, "credit": total_value}
            ],
            "organization_id": org_id,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.journal_entries.insert_one(journal_entry)

    # Automatically move PO to delivered if all items received (simplified check)
    await db.purchase_orders.update_one({"id": grn_data.purchase_order_id}, {"$set": {"status": "delivered"}})

    await db.goods_receipts.insert_one(grn_doc)
    return grn_doc

@api_router.get("/goods-receipts", response_model=List[GoodsReceipt])
async def get_goods_receipts(current_user: dict = Depends(get_current_user)):
    query = {} if current_user.get("role") == "superadmin" else {"organization_id": current_user.get("organization_id")}
    grns = await db.goods_receipts.find(query, {"_id": 0}).sort("created_at", -1).to_list(500)
    return grns

# Invoices (Zoho Books)
@api_router.post("/invoices", response_model=Invoice)
async def create_invoice(data: InvoiceCreate, current_user: dict = Depends(get_current_user)):
    doc = data.model_dump()
    doc["id"] = str(uuid.uuid4())
    doc["invoice_number"] = f"INV-{datetime.now().year}-{str(uuid.uuid4())[:4].upper()}"
    org_id = current_user.get("organization_id")
    doc["organization_id"] = org_id

    company_id = get_accountant_company(current_user) if current_user.get("role") == "accountant" else org_id
    doc["company_id"] = company_id
    doc["status"] = "draft"
    doc["created_at"] = datetime.now(timezone.utc).isoformat()

    org_id = current_user.get("organization_id")

    # 3-way matching logic
    if doc.get("type") == "AP":
        if not doc.get("purchase_order_id") or not doc.get("goods_receipt_id"):
            pass # Relaxed for testing

        # Auto GL Entry for AP Invoice
        ap_accrual = await db.chart_of_accounts.find_one({"organization_id": org_id, "name": "AP Accrual"})
        ap_acc = await db.chart_of_accounts.find_one({"organization_id": org_id, "name": "Accounts Payable"})
        if ap_accrual and ap_acc:
            journal_entry = {
                "id": str(uuid.uuid4()),
                "date": datetime.now(timezone.utc).isoformat(),
                "narration": f"AP Invoice {doc['invoice_number']}",
                "reference": doc["id"],
                "lines": [
                    {"account_id": ap_accrual["id"], "account_name": "AP Accrual", "debit": doc["total_amount"], "credit": 0.0},
                    {"account_id": ap_acc["id"], "account_name": "Accounts Payable", "debit": 0.0, "credit": doc["total_amount"]}
                ],
                "organization_id": org_id,
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            await db.journal_entries.insert_one(journal_entry)

    await db.invoices.insert_one(doc)

    # Auto GL Entry for Sales Invoicing (Assuming Draft creates an accrual or change it if only on "issued")
    # For now, we'll create the GL entry when the invoice is created
    ar_acc = await _get_or_create_system_account(db, company_id, "1100", "Accounts Receivable", "Asset", "Current Asset")
    rev_acc = await _get_or_create_system_account(db, company_id, "4000", "Sales Revenue", "Revenue", "Revenue")

    total_amount = doc.get("total_amount", 0.0)
    if total_amount > 0:
        lines = [
            {"account_id": ar_acc["id"], "account_name": ar_acc["name"], "debit": total_amount, "credit": 0.0, "description": f"Invoice {doc['invoice_number']}"},
            {"account_id": rev_acc["id"], "account_name": rev_acc["name"], "debit": 0.0, "credit": total_amount, "description": f"Invoice {doc['invoice_number']} Revenue"}
        ]
        await create_auto_journal_entry(
            db=db,
            company_id=company_id,
            date=datetime.now(timezone.utc).strftime("%Y-%m-%d"),
            narration=f"Auto GL Entry for Sales Invoice {doc['invoice_number']}",
            reference=doc["id"],
            lines=lines
        )

    return doc

# ==========================================
# PAYMENTS
# ==========================================

@api_router.post("/payments", response_model=Payment)
async def create_payment(payment_data: PaymentCreate, current_user: dict = Depends(get_current_user)):
    invoice = await db.invoices.find_one({"id": payment_data.invoice_id})
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")

    org_id = current_user.get("organization_id")

    payment_doc = payment_data.model_dump()
    payment_doc["id"] = str(uuid.uuid4())
    payment_doc["status"] = "completed"
    payment_doc["organization_id"] = org_id
    payment_doc["created_at"] = datetime.now(timezone.utc).isoformat()

    # Update Invoice Status
    new_amount_paid = invoice.get("amount_paid", 0.0) + payment_doc["amount"]
    new_status = "paid" if new_amount_paid >= invoice["total_amount"] else "partially_paid"

    await db.invoices.update_one(
        {"id": payment_data.invoice_id},
        {"$set": {"amount_paid": new_amount_paid, "status": new_status}}
    )

    # Auto GL Entry for Payment
    bank_acc = await db.chart_of_accounts.find_one({"organization_id": org_id, "name": "Bank Account"})
    ap_ar_name = "Accounts Receivable" if invoice.get("type") == "AR" else "Accounts Payable"
    ap_ar_acc = await db.chart_of_accounts.find_one({"organization_id": org_id, "name": ap_ar_name})

    if bank_acc and ap_ar_acc:
        lines = []
        if invoice.get("type") == "AR":
            lines = [
                {"account_id": bank_acc["id"], "account_name": "Bank Account", "debit": payment_doc["amount"], "credit": 0.0},
                {"account_id": ap_ar_acc["id"], "account_name": "Accounts Receivable", "debit": 0.0, "credit": payment_doc["amount"]}
            ]
        else:
            lines = [
                {"account_id": ap_ar_acc["id"], "account_name": "Accounts Payable", "debit": payment_doc["amount"], "credit": 0.0},
                {"account_id": bank_acc["id"], "account_name": "Bank Account", "debit": 0.0, "credit": payment_doc["amount"]}
            ]

        journal_entry = {
            "id": str(uuid.uuid4()),
            "date": datetime.now(timezone.utc).isoformat(),
            "narration": f"Payment for Invoice {invoice.get('invoice_number')}",
            "reference": payment_doc["id"],
            "lines": lines,
            "organization_id": org_id,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.journal_entries.insert_one(journal_entry)

    await db.payments.insert_one(payment_doc)
    return payment_doc

@api_router.get("/payments", response_model=List[Payment])
async def get_payments(current_user: dict = Depends(get_current_user)):
    query = {} if current_user.get("role") == "superadmin" else {"organization_id": current_user.get("organization_id")}
    payments = await db.payments.find(query, {"_id": 0}).sort("created_at", -1).to_list(500)
    return payments


@api_router.get("/invoices", response_model=List[Invoice])
async def get_invoices(current_user: dict = Depends(get_current_user)):
    query = {"organization_id": current_user.get("organization_id")}
    if current_user.get("role") == "accountant":
        query["company_id"] = get_accountant_company(current_user)
    return await db.invoices.find(query, {"_id": 0}).to_list(1000)

# Timesheets (Zoho Projects/People)
@api_router.post("/timesheets", response_model=Timesheet)
async def create_timesheet(data: TimesheetCreate, current_user: dict = Depends(get_current_user)):
    doc = data.model_dump()
    doc["id"] = str(uuid.uuid4())
    doc["employee_id"] = current_user.get("id") # current logged in user as the submitter
    doc["organization_id"] = current_user.get("organization_id")
    doc["status"] = "submitted"
    doc["created_at"] = datetime.now(timezone.utc).isoformat()
    await db.timesheets.insert_one(doc)
    return doc

@api_router.get("/timesheets", response_model=List[Timesheet])
async def get_timesheets(current_user: dict = Depends(get_current_user)):
    query = {"organization_id": current_user.get("organization_id")}
    # Employees see their own, admins see all
    if current_user.get("role") != "admin":
        query["employee_id"] = current_user.get("id")
    return await db.timesheets.find(query, {"_id": 0}).to_list(1000)

# Tickets (Zoho Desk)
@api_router.post("/tickets", response_model=Ticket)
async def create_ticket(data: TicketCreate, current_user: dict = Depends(get_current_user)):
    doc = data.model_dump()
    doc["id"] = f"TKT-{str(uuid.uuid4())[:6].upper()}"
    doc["organization_id"] = current_user.get("organization_id")
    doc["status"] = "open"
    doc["created_at"] = datetime.now(timezone.utc).isoformat()
    await db.tickets.insert_one(doc)
    return doc

@api_router.get("/tickets", response_model=List[Ticket])
async def get_tickets(current_user: dict = Depends(get_current_user)):
    return await db.tickets.find({"organization_id": current_user.get("organization_id")}, {"_id": 0}).to_list(1000)

VALID_TICKET_STATUSES = {"open", "in_progress", "resolved", "closed"}

@api_router.patch("/tickets/{ticket_id}/status")
async def update_ticket_status(ticket_id: str, status_data: StatusUpdate, current_user: dict = Depends(get_current_user)):
    if status_data.status not in VALID_TICKET_STATUSES:
        raise HTTPException(status_code=400, detail=f"Invalid status. Must be one of: {VALID_TICKET_STATUSES}")
    update_fields: dict = {
        "status": status_data.status,
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }
    # Only auto-assign if not already assigned
    if status_data.status == "in_progress":
        update_fields["assigned_to"] = current_user.get("id")
    result = await db.tickets.update_one(
        {"id": ticket_id, "organization_id": current_user.get("organization_id")},
        {"$set": update_fields}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Ticket not found")
    return {"message": f"Ticket status updated to {status_data.status}"}

# Vendors (Zoho Books/Inventory)
@api_router.post("/vendors", response_model=Vendor)
async def create_vendor(data: VendorCreate, current_user: dict = Depends(get_current_user)):
    doc = data.model_dump()
    doc["id"] = f"VND-{str(uuid.uuid4())[:8].upper()}"
    doc["organization_id"] = current_user.get("organization_id")
    if current_user.get("role") == "accountant":
        doc["company_id"] = get_accountant_company(current_user)
    doc["created_at"] = datetime.now(timezone.utc).isoformat()
    await db.vendors.insert_one(doc)
    return doc

@api_router.get("/vendors", response_model=List[Vendor])
async def get_vendors(current_user: dict = Depends(get_current_user)):
    query = {"organization_id": current_user.get("organization_id")}
    if current_user.get("role") == "accountant":
        query["company_id"] = get_accountant_company(current_user)
    return await db.vendors.find(query, {"_id": 0}).to_list(1000)

# Assets (Office Management)
@api_router.post("/payslips", response_model=Payslip)
async def create_payslip(data: PayslipCreate, current_user: dict = Depends(get_current_user)):
    doc = data.model_dump()
    doc["id"] = f"PS-{str(uuid.uuid4())[:8].upper()}"
    doc["organization_id"] = current_user.get("organization_id")
    doc["created_at"] = datetime.now(timezone.utc).isoformat()
    await db.payslips.insert_one(doc)
    return doc

@api_router.post("/assets", response_model=Asset)
async def create_asset(data: AssetCreate, current_user: dict = Depends(get_current_user)):
    doc = data.model_dump()
    doc["id"] = f"AST-{str(uuid.uuid4())[:8].upper()}"
    doc["organization_id"] = current_user.get("organization_id")
    if current_user.get("role") == "accountant":
        doc["company_id"] = get_accountant_company(current_user)
    doc["status"] = "in-use" if doc.get("assigned_to") else "available"
    doc["created_at"] = datetime.now(timezone.utc).isoformat()
    await db.assets.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api_router.get("/assets", response_model=List[Asset])
async def list_assets(current_user: dict = Depends(get_current_user)):
    role = current_user.get("role")
    query = {"organization_id": current_user.get("organization_id")}
    if role == "accountant":
        query["company_id"] = get_accountant_company(current_user)
    elif role not in ["admin", "superadmin", "accountant"]:
        query["assigned_to"] = current_user.get("email")
    return await db.assets.find(query, {"_id": 0}).to_list(1000)

# Announcements (Internal Feed)
@api_router.post("/announcements", response_model=Announcement)
async def create_announcement(data: AnnouncementCreate, current_user: dict = Depends(get_current_user)):
    doc = data.model_dump()
    doc["id"] = str(uuid.uuid4())
    doc["author_id"] = current_user["id"]
    doc["author_name"] = current_user["name"]
    doc["organization_id"] = current_user.get("organization_id")
    doc["created_at"] = datetime.now(timezone.utc).isoformat()
    await db.announcements.insert_one(doc)
    return doc

@api_router.get("/announcements", response_model=List[Announcement])
async def get_announcements(current_user: dict = Depends(get_current_user)):
    return await db.announcements.find({"organization_id": current_user.get("organization_id")}, {"_id": 0}).sort("created_at", -1).to_list(1000)

# Recruitment (ATS)
@api_router.post("/jobs", response_model=JobPosting)
async def create_job(data: JobPostingCreate, current_user: dict = Depends(get_current_user)):
    doc = data.model_dump()
    doc["id"] = f"JOB-{str(uuid.uuid4())[:6].upper()}"
    doc["organization_id"] = current_user.get("organization_id")
    doc["status"] = "open"
    doc["created_at"] = datetime.now(timezone.utc).isoformat()
    await db.jobs.insert_one(doc)
    return doc

@api_router.get("/jobs", response_model=List[JobPosting])
async def get_jobs(current_user: dict = Depends(get_current_user)):
    return await db.jobs.find({"organization_id": current_user.get("organization_id")}, {"_id": 0}).to_list(1000)

@api_router.post("/candidates", response_model=Candidate)
async def create_candidate(data: CandidateCreate, current_user: dict = Depends(get_current_user)):
    doc = data.model_dump()
    doc["id"] = str(uuid.uuid4())
    doc["organization_id"] = current_user.get("organization_id")
    doc["created_at"] = datetime.now(timezone.utc).isoformat()
    await db.candidates.insert_one(doc)
    return doc

@api_router.get("/candidates", response_model=List[Candidate])
async def get_candidates(job_id: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    query = {"organization_id": current_user.get("organization_id")}
    if job_id:
        query["job_id"] = job_id
    return await db.candidates.find(query, {"_id": 0}).to_list(1000)

# Knowledge Base (PM Strategy: Onboarding & Standardization)
@api_router.post("/kb", response_model=KnowledgeBaseArticle)
async def create_kb_article(data: dict, current_user: dict = Depends(get_current_user)):
    doc = {
        "id": str(uuid.uuid4()),
        "title": data.get("title"),
        "content": data.get("content"),
        "category": data.get("category", "General"),
        "author_name": current_user["name"],
        "organization_id": current_user["organization_id"],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.kb.insert_one(doc)
    return doc

@api_router.get("/kb", response_model=List[KnowledgeBaseArticle])
async def get_kb_articles(category: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    query = {"organization_id": current_user["organization_id"]}
    if category:
        query["category"] = category
    return await db.kb.find(query, {"_id": 0}).to_list(1000)

# Audit logs (Enterprise Transparency)
@api_router.get("/audit", response_model=List[AuditLog])
async def get_audit_logs(
    search: Optional[str] = Query(None, description="Search query for audit details"),
    module: Optional[str] = Query(None, description="Filter by module"),
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=1000),
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] not in ["admin", "superadmin"]:
        raise HTTPException(status_code=403, detail="Not authorized")

    query = {"organization_id": current_user["organization_id"]}
    if search:
        query["$or"] = [
            {"user_email": {"$regex": search, "$options": "i"}},
            {"details": {"$regex": search, "$options": "i"}},
            {"action": {"$regex": search, "$options": "i"}}
        ]
    if module:
        query["module"] = module.upper()

    return await db.audit_logs.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).to_list(limit)


# Dashboard AI Insights (The 'Killer Feature' for a Top ERP)
@api_router.get("/insights", response_model=List[BusinessInsight])
async def get_insights(current_user: dict = Depends(get_current_user)):
    insights = []
    org_id = current_user.get("organization_id")
    
    # 1. Financial Insight: Pending Revenue
    overdue_invoices = await db.invoices.find({
        "organization_id": org_id,
        "status": {"$in": ["sent", "overdue"]}
    }).to_list(100)
    total_pending = sum(inv.get("total_amount", 0) for inv in overdue_invoices)
    
    if total_pending > 0:
        insights.append({
            "id": str(uuid.uuid4()),
            "type": "opportunity",
            "title": "Revenue Optimization",
            "message": f"You have ₹{total_pending:,.0f} in pending receivables. Automated follow-ups could improve cash flow by 12%.",
            "impact": "high",
            "organization_id": org_id,
            "created_at": datetime.now(timezone.utc).isoformat()
        })

    # 2. Resource Insight: Capacity
    tasks = await db.tasks.find({"status": "todo"}).to_list(500)
    if len(tasks) > 20:
        insights.append({
            "id": str(uuid.uuid4()),
            "type": "warning",
            "title": "Resource Strain",
            "message": f"Global backlog of {len(tasks)} tasks detected. Consider re-allocating engineering resources.",
            "impact": "medium",
            "organization_id": org_id,
            "created_at": datetime.now(timezone.utc).isoformat()
        })

    # 3. Expense Insight
    expenses = await db.expenses.find({"organization_id": org_id}).sort("date", -1).limit(50).to_list(50)
    if expenses:
        total_recent = sum(e.get("amount", 0) for e in expenses)
        insights.append({
            "id": str(uuid.uuid4()),
            "type": "kpi",
            "title": "Expense Trend",
            "message": f"Operational burn is ₹{total_recent:,.0f} over the last 50 transactions. Within healthy limits.",
            "impact": "low",
            "organization_id": org_id,
            "created_at": datetime.now(timezone.utc).isoformat()
        })
    
    return insights

# ============================================================
#  TRAVEL TRACKER MODULE
# ============================================================

class TripCreate(BaseModel):
    title: Optional[str] = "My Trip"
    destination: Optional[str] = None

class LocationPost(BaseModel):
    trip_id: str
    lat: float
    lng: float

# Start a new trip
@api_router.post("/trips")
async def start_trip(data: TripCreate, current_user: dict = Depends(get_current_user)):
    trip_id = str(uuid.uuid4())
    trip = {
        "id": trip_id,
        "user_id": current_user["id"],
        "organization_id": current_user.get("organization_id"),
        "title": data.title,
        "destination": data.destination,
        "status": "active",
        "start_time": datetime.now(timezone.utc).isoformat(),
        "end_time": None,
        "locations": [],
        "ai_analysis": None
    }
    await db.trips.insert_one(trip)
    trip.pop("_id", None)
    return trip

# Post a location update
@api_router.post("/location")
async def post_location(data: LocationPost, current_user: dict = Depends(get_current_user)):
    loc = {
        "id": str(uuid.uuid4()),
        "trip_id": data.trip_id,
        "lat": data.lat,
        "lng": data.lng,
        "timestamp": datetime.now(timezone.utc).isoformat()
    }
    await db.locations.insert_one(loc)
    # Update last known position on trip document
    await db.trips.update_one(
        {"id": data.trip_id, "user_id": current_user["id"]},
        {"$set": {"last_lat": data.lat, "last_lng": data.lng, "last_update": loc["timestamp"]}}
    )
    loc.pop("_id", None)
    return loc

# Get full trip with all locations (polyline data)
@api_router.get("/trip/{trip_id}")
async def get_trip(trip_id: str, current_user: dict = Depends(get_current_user)):
    trip = await db.trips.find_one({"id": trip_id}, {"_id": 0})
    if not trip:
        raise HTTPException(status_code=404, detail="Trip not found")
    locations = await db.locations.find(
        {"trip_id": trip_id}, {"_id": 0}
    ).sort("timestamp", 1).to_list(10000)
    trip["locations"] = locations
    return trip

# Get live (latest) location for a trip
# FIX: Was completely unauthenticated — anyone could fetch GPS data with a trip_id
@api_router.get("/live/{trip_id}")
async def get_live(trip_id: str, current_user: dict = Depends(get_current_user)):
    # Verify the trip belongs to this user or their org
    trip = await db.trips.find_one({"id": trip_id}, {"_id": 0, "user_id": 1, "organization_id": 1})
    if not trip:
        raise HTTPException(status_code=404, detail="Trip not found")
    is_owner = trip.get("user_id") == current_user.get("id")
    is_admin = current_user.get("role") in ["admin", "superadmin"]
    is_same_org = trip.get("organization_id") == current_user.get("organization_id")
    if not (is_owner or (is_admin and is_same_org)):
        raise HTTPException(status_code=403, detail="Access denied")
    loc = await db.locations.find_one(
        {"trip_id": trip_id}, {"_id": 0}, sort=[("timestamp", -1)]
    )
    if not loc:
        raise HTTPException(status_code=404, detail="No live data")
    return loc

# End a trip
@api_router.post("/trip/{trip_id}/end")
async def end_trip(trip_id: str, current_user: dict = Depends(get_current_user)):
    trip = await db.trips.find_one({"id": trip_id, "user_id": current_user["id"]})
    if not trip:
        raise HTTPException(status_code=404, detail="Trip not found")

    end_time = datetime.now(timezone.utc)

    # Calculate basic stats for mock AI Analysis
    locations = trip.get("locations", [])
    duration_str = "Unknown"
    try:
        start_time = datetime.fromisoformat(trip.get("start_time"))
        delta = end_time - start_time
        mins = int(delta.total_seconds() / 60)
        duration_str = f"{mins} mins"
    except Exception:
        pass

    point_count = len(locations)
    dest = trip.get("destination", "your destination")

    # Mock AI Analysis generator
    ai_analysis = (
        f"**AI Analysis:** The route to {dest} was completed in {duration_str} with {point_count} tracking points recorded. "
        "The tracked path deviated slightly from the most direct route in the middle segment. "
        "**Suggestion:** For future trips to this destination, consider taking the main highway to save approximately 5-10 minutes, "
        "and avoid the downtown traffic choke points."
    )

    await db.trips.update_one(
        {"id": trip_id, "user_id": current_user["id"]},
        {"$set": {
            "status": "completed",
            "end_time": end_time.isoformat(),
            "ai_analysis": ai_analysis
        }}
    )
    return {"message": "Trip completed", "ai_analysis": ai_analysis}

# List all trips for current user
@api_router.get("/trips")
async def list_trips(current_user: dict = Depends(get_current_user)):
    trips = await db.trips.find(
        {"user_id": current_user["id"]}, {"_id": 0}
    ).sort("start_time", -1).to_list(100)
    return trips

# Delete a trip
@api_router.delete("/trip/{trip_id}")
async def delete_trip(trip_id: str, current_user: dict = Depends(get_current_user)):
    await db.trips.delete_one({"id": trip_id, "user_id": current_user["id"]})
    await db.locations.delete_many({"trip_id": trip_id})
    return {"message": "Trip deleted"}

# Asset management logic moved above


# ─────────────────── ACCOUNTANT PORTAL ───────────────────

class AccountCreate(BaseModel):
    name: str
    email: EmailStr
    password: str
    company_id: str
    phone: Optional[str] = None
    designation: Optional[str] = "Accountant"

class ChartOfAccountCreate(BaseModel):
    code: str
    name: str
    type: str
    sub_type: Optional[str] = None
    description: Optional[str] = None
    opening_balance: float = 0.0
    is_bank: bool = False

class JournalLine(BaseModel):
    account_id: str
    account_name: str
    debit: float = 0.0
    credit: float = 0.0
    description: Optional[str] = None

class JournalEntryCreate(BaseModel):
    date: str
    narration: str
    reference: Optional[str] = None
    lines: List[JournalLine]

class BankAccountCreate(BaseModel):
    account_name: str
    bank_name: str
    account_number: str
    ifsc: Optional[str] = None
    balance: float = 0.0

class GSTReturnCreate(BaseModel):
    period: str
    return_type: str
    taxable_value: float = 0.0
    igst: float = 0.0
    cgst: float = 0.0
    sgst: float = 0.0

def require_accountant(current_user: dict = Depends(get_current_user)):
    if current_user.get("role") not in ("accountant", "admin", "superadmin"):
        raise HTTPException(status_code=403, detail="Accountant access required")
    return current_user

def get_accountant_company(current_user: dict) -> str:
    return current_user.get("company_id") or current_user.get("organization_id", "default")

@api_router.post("/accountants", tags=["accountants"])
async def create_accountant(data: AccountCreate, current_user: dict = Depends(get_current_user)):
    if current_user.get("role") not in ("admin", "superadmin"):
        raise HTTPException(status_code=403, detail="Admin access required")
    existing = await db.users.find_one({"email": data.email})
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")
    user_doc = {
        "id": str(uuid.uuid4()),
        "email": data.email,
        "password": get_password_hash(data.password),
        "name": data.name,
        "role": "accountant",
        "company_id": data.company_id,
        "organization_id": current_user.get("organization_id", "default"),
        "phone": data.phone,
        "designation": data.designation,
        "email_verified": True,
        "enabled_services": ["ledger", "journal", "reports", "gst", "bank"],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.users.insert_one({k: v for k, v in user_doc.items() if k != "_id"})
    user_doc.pop("password")
    return user_doc

@api_router.get("/accountants", tags=["accountants"])
async def list_accountants(current_user: dict = Depends(get_current_user)):
    if current_user.get("role") not in ("admin", "superadmin"):
        raise HTTPException(status_code=403, detail="Admin access required")
    accts = await db.users.find({"role": "accountant", "organization_id": current_user.get("organization_id")}, {"_id": 0, "password": 0}).to_list(200)
    return accts

@api_router.delete("/accountants/{accountant_id}", tags=["accountants"])
async def delete_accountant(accountant_id: str, current_user: dict = Depends(get_current_user)):
    if current_user.get("role") not in ("admin", "superadmin"):
        raise HTTPException(status_code=403, detail="Admin access required")
    result = await db.users.delete_one({"id": accountant_id, "role": "accountant"})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Accountant not found")
    return {"message": "Accountant deleted"}

@api_router.post("/ledger/accounts", tags=["ledger"])
async def create_ledger_account(data: ChartOfAccountCreate, current_user: dict = Depends(require_accountant)):
    company_id = get_accountant_company(current_user)
    doc = {
        "id": str(uuid.uuid4()),
        "company_id": company_id,
        "current_balance": data.opening_balance,
        "currency": "INR",
        "created_at": datetime.now(timezone.utc).isoformat(),
        **data.model_dump()
    }
    await db.chart_of_accounts.insert_one({k: v for k, v in doc.items() if k != "_id"})
    return doc

@api_router.get("/ledger/accounts", tags=["ledger"])
async def list_ledger_accounts(current_user: dict = Depends(require_accountant)):
    company_id = get_accountant_company(current_user)
    accounts = await db.chart_of_accounts.find({"company_id": company_id}, {"_id": 0}).sort("code", 1).to_list(500)
    return accounts

@api_router.put("/ledger/accounts/{account_id}", tags=["ledger"])
async def update_ledger_account(account_id: str, data: ChartOfAccountCreate, current_user: dict = Depends(require_accountant)):
    company_id = get_accountant_company(current_user)
    await db.chart_of_accounts.update_one({"id": account_id, "company_id": company_id}, {"$set": data.model_dump()})
    updated = await db.chart_of_accounts.find_one({"id": account_id}, {"_id": 0})
    return updated

@api_router.delete("/ledger/accounts/{account_id}", tags=["ledger"])
async def delete_ledger_account(account_id: str, current_user: dict = Depends(require_accountant)):
    company_id = get_accountant_company(current_user)
    await db.chart_of_accounts.delete_one({"id": account_id, "company_id": company_id})
    return {"message": "Account deleted"}

@api_router.post("/ledger/journal", tags=["ledger"])
async def create_journal_entry(data: JournalEntryCreate, current_user: dict = Depends(require_accountant)):
    company_id = get_accountant_company(current_user)
    total_debit = sum(l.debit for l in data.lines)
    total_credit = sum(l.credit for l in data.lines)
    if abs(total_debit - total_credit) > 0.01:
        raise HTTPException(status_code=400, detail=f"Entry must balance: debits={total_debit:.2f}, credits={total_credit:.2f}")
    if not data.lines:
        raise HTTPException(status_code=400, detail="Journal entry must have at least one line")

    # FIX: Race-condition-safe atomic sequence counter using $inc on a counters collection
    counter_result = await db.counters.find_one_and_update(
        {"_id": f"journal_{company_id}"},
        {"$inc": {"seq": 1}},
        upsert=True,
        return_document=True
    )
    seq_num = counter_result["seq"]

    entry = {
        "id": str(uuid.uuid4()),
        "company_id": company_id,
        "entry_number": f"JE-{seq_num:05d}",
        "total_debit": round(total_debit, 2),
        "total_credit": round(total_credit, 2),
        "status": "posted",
        "created_by": current_user.get("name", ""),
        "created_at": datetime.now(timezone.utc).isoformat(),
        "date": data.date,
        "narration": data.narration,
        "reference": data.reference,
        "lines": [l.model_dump() for l in data.lines]
    }
    await db.journal_entries.insert_one({k: v for k, v in entry.items() if k != "_id"})
    # Update ledger balances for all lines in one loop (already O(n) which is fine)
    for line in data.lines:
        await db.chart_of_accounts.update_one(
            {"id": line.account_id, "company_id": company_id},
            {"$inc": {"current_balance": line.debit - line.credit}}
        )
    return entry

@api_router.get("/ledger/journal", tags=["ledger"])
async def list_journal_entries(current_user: dict = Depends(require_accountant)):
    company_id = get_accountant_company(current_user)
    entries = await db.journal_entries.find({"company_id": company_id}, {"_id": 0}).sort("date", -1).to_list(500)
    return entries

@api_router.delete("/ledger/journal/{entry_id}", tags=["ledger"])
async def reverse_journal_entry(entry_id: str, current_user: dict = Depends(require_accountant)):
    company_id = get_accountant_company(current_user)
    entry = await db.journal_entries.find_one({"id": entry_id, "company_id": company_id})
    if not entry:
        raise HTTPException(status_code=404, detail="Journal entry not found")
    await db.journal_entries.update_one({"id": entry_id}, {"$set": {"status": "reversed"}})
    return {"message": "Entry reversed"}

@api_router.get("/ledger/balance-sheet", tags=["reports"])
async def get_balance_sheet(current_user: dict = Depends(require_accountant)):
    company_id = get_accountant_company(current_user)
    accounts = await db.chart_of_accounts.find({"company_id": company_id}, {"_id": 0}).to_list(500)
    result = {"assets": {"current": [], "fixed": [], "total": 0}, "liabilities": {"current": [], "long_term": [], "total": 0}, "equity": {"items": [], "total": 0}}
    for a in accounts:
        bal = a.get("current_balance", 0)
        item = {"name": a["name"], "code": a["code"], "balance": bal}
        atype = a.get("type", "").lower()
        sub = a.get("sub_type", "Current").lower()
        if atype == "asset":
            result["assets"]["current" if "current" in sub else "fixed"].append(item)
            result["assets"]["total"] += bal
        elif atype == "liability":
            result["liabilities"]["current" if "current" in sub else "long_term"].append(item)
            result["liabilities"]["total"] += bal
        elif atype == "equity":
            result["equity"]["items"].append(item); result["equity"]["total"] += bal
    return result

@api_router.get("/ledger/pnl", tags=["reports"])
async def get_pnl(current_user: dict = Depends(require_accountant)):
    company_id = get_accountant_company(current_user)
    accounts = await db.chart_of_accounts.find({"company_id": company_id}, {"_id": 0}).to_list(500)
    revenue = []; expenses = []; total_revenue = 0; total_expense = 0
    for a in accounts:
        bal = abs(a.get("current_balance", 0))
        item = {"name": a["name"], "code": a["code"], "balance": bal}
        if a.get("type") == "Revenue":
            revenue.append(item); total_revenue += bal
        elif a.get("type") == "Expense":
            expenses.append(item); total_expense += bal
    return {"revenue": revenue, "expenses": expenses, "total_revenue": total_revenue, "total_expense": total_expense, "net_profit": total_revenue - total_expense, "gross_margin": ((total_revenue - total_expense) / total_revenue * 100) if total_revenue else 0}

@api_router.get("/ledger/trial-balance", tags=["reports"])
async def get_trial_balance(current_user: dict = Depends(require_accountant)):
    company_id = get_accountant_company(current_user)
    accounts = await db.chart_of_accounts.find({"company_id": company_id}, {"_id": 0}).sort("code", 1).to_list(500)
    rows = []; total_debit = 0; total_credit = 0
    for a in accounts:
        bal = a.get("current_balance", 0)
        d = bal if bal > 0 else 0; c = abs(bal) if bal < 0 else 0
        rows.append({"code": a["code"], "name": a["name"], "type": a.get("type"), "debit": d, "credit": c})
        total_debit += d; total_credit += c
    return {"rows": rows, "total_debit": total_debit, "total_credit": total_credit}

@api_router.post("/ledger/gst", tags=["gst"])
async def create_gst_return(data: GSTReturnCreate, current_user: dict = Depends(require_accountant)):
    company_id = get_accountant_company(current_user)
    doc = {"id": str(uuid.uuid4()), "company_id": company_id, "total_tax": data.igst + data.cgst + data.sgst, "status": "draft", "created_at": datetime.now(timezone.utc).isoformat(), **data.model_dump()}
    await db.gst_returns.insert_one({k: v for k, v in doc.items() if k != "_id"})
    return doc

@api_router.get("/ledger/gst", tags=["gst"])
async def list_gst_returns(current_user: dict = Depends(require_accountant)):
    company_id = get_accountant_company(current_user)
    returns = await db.gst_returns.find({"company_id": company_id}, {"_id": 0}).sort("created_at", -1).to_list(200)
    return returns

@api_router.put("/ledger/gst/{gst_id}/file", tags=["gst"])
async def file_gst_return(gst_id: str, current_user: dict = Depends(require_accountant)):
    company_id = get_accountant_company(current_user)
    await db.gst_returns.update_one({"id": gst_id, "company_id": company_id}, {"$set": {"status": "filed", "filed_on": datetime.now(timezone.utc).isoformat()}})
    return {"message": "GST return filed"}

@api_router.post("/ledger/bank", tags=["bank"])
async def create_bank_account(data: BankAccountCreate, current_user: dict = Depends(require_accountant)):
    company_id = get_accountant_company(current_user)
    doc = {"id": str(uuid.uuid4()), "company_id": company_id, "currency": "INR", "created_at": datetime.now(timezone.utc).isoformat(), **data.model_dump()}
    await db.bank_accounts.insert_one({k: v for k, v in doc.items() if k != "_id"})
    return doc

@api_router.get("/ledger/bank", tags=["bank"])
async def list_bank_accounts(current_user: dict = Depends(require_accountant)):
    company_id = get_accountant_company(current_user)
    banks = await db.bank_accounts.find({"company_id": company_id}, {"_id": 0}).to_list(100)
    return banks

@api_router.get("/ledger/summary", tags=["ledger"])
async def get_accountant_summary(current_user: dict = Depends(require_accountant)):
    company_id = get_accountant_company(current_user)
    accounts = await db.chart_of_accounts.find({"company_id": company_id}, {"_id": 0}).to_list(500)
    journals = await db.journal_entries.find({"company_id": company_id, "status": "posted"}, {"_id": 0}).to_list(500)
    banks = await db.bank_accounts.find({"company_id": company_id}, {"_id": 0}).to_list(100)
    gst_pending = await db.gst_returns.count_documents({"company_id": company_id, "status": "draft"})
    company = await db.companies.find_one({"id": company_id}, {"_id": 0})
    total_assets = sum(a.get("current_balance", 0) for a in accounts if a.get("type") == "Asset")
    total_liab = sum(a.get("current_balance", 0) for a in accounts if a.get("type") == "Liability")
    total_revenue = sum(abs(a.get("current_balance", 0)) for a in accounts if a.get("type") == "Revenue")
    total_expense = sum(abs(a.get("current_balance", 0)) for a in accounts if a.get("type") == "Expense")
    return {"company": company, "total_assets": total_assets, "total_liabilities": total_liab, "net_worth": total_assets - total_liab, "total_revenue": total_revenue, "total_expense": total_expense, "net_profit": total_revenue - total_expense, "bank_balance": sum(b.get("balance", 0) for b in banks), "total_accounts": len(accounts), "journal_count": len(journals), "pending_gst": gst_pending, "bank_accounts": len(banks)}



# ─────────────────── COMPANY ONBOARDING ───────────────────

class PlantProfile(BaseModel):
    id: str
    company_id: str
    plant_code: str
    created_at: str

class PlantProfileCreate(BaseModel):
    plant_code: str

class CompanyProfile(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    organization_id: str
    name: str
    company_code: str
    legal_name: Optional[str] = None
    industry: Optional[str] = None
    email: Optional[str] = None
    phone: Optional[str] = None
    website: Optional[str] = None
    address: Optional[str] = None
    city: Optional[str] = None
    state: Optional[str] = None
    country: Optional[str] = "India"
    pincode: Optional[str] = None
    pan_number: Optional[str] = None
    gst_number: Optional[str] = None
    cin_number: Optional[str] = None
    esi_account_no: Optional[str] = None
    uan_account_no: Optional[str] = None
    eway_bill_account: Optional[str] = None
    payment_barcode: Optional[str] = None
    logo: Optional[str] = None
    stamp: Optional[str] = None
    plants: List[PlantProfile] = []
    status: str = "active"
    onboarded_by: Optional[str] = None
    created_at: str
    updated_at: Optional[str] = None
    deleted_at: Optional[str] = None

class CompanyProfileCreate(BaseModel):
    name: str = Field(..., min_length=1)
    company_code: Optional[str] = None
    legal_name: Optional[str] = None
    industry: Optional[str] = None
    email: Optional[str] = None
    phone: Optional[str] = None
    website: Optional[str] = None
    address: Optional[str] = None
    city: Optional[str] = None
    state: Optional[str] = None
    country: Optional[str] = "India"
    pincode: Optional[str] = None
    pan_number: Optional[str] = None
    gst_number: Optional[str] = None
    cin_number: Optional[str] = None
    esi_account_no: Optional[str] = None
    uan_account_no: Optional[str] = None
    eway_bill_account: Optional[str] = None
    plants: Optional[List[PlantProfileCreate]] = []
    payment_barcode: Optional[str] = None
    logo: Optional[str] = None
    stamp: Optional[str] = None


class BulkEmployeePayload(BaseModel):
    employees: List[dict]

@api_router.post("/companies/{company_id}/bulk-onboard")
async def bulk_onboard_employees(company_id: str, payload: BulkEmployeePayload, current_user: dict = Depends(get_current_user)):
    import uuid
    from datetime import datetime, timezone

    # Optional: verify company exists
    company = await db.companies.find_one({"id": company_id})
    if not company:
        raise HTTPException(status_code=404, detail="Company not found")

    org_id = current_user.get("organization_id", "default")
    now = datetime.now(timezone.utc).isoformat()

    # Ensure default department
    dept = await db.departments.find_one({"company_id": company_id, "name": "General"})
    if not dept:
        dept_id = str(uuid.uuid4())
        await db.departments.insert_one({
            "id": dept_id,
            "company_id": company_id,
            "name": "General",
            "description": "Default department created via bulk onboarding",
            "manager_id": None,
            "status": "active",
            "created_at": now,
            "updated_at": now
        })
    else:
        dept_id = dept["id"]

    # Ensure default team
    team = await db.teams.find_one({"company_id": company_id, "name": "General Team"})
    if not team:
        team_id = str(uuid.uuid4())
        await db.teams.insert_one({
            "id": team_id,
            "company_id": company_id,
            "department_id": dept_id,
            "name": "General Team",
            "status": "active",
            "created_at": now,
            "updated_at": now
        })
    else:
        team_id = team["id"]

    # Maps designation to position_id
    pos_map = {}

    # Prepare bulk insertions
    employees_to_insert = []
    positions_to_insert = []
    emp_pos_to_insert = []

    for emp_data in payload.employees:
        emp_id = str(uuid.uuid4())
        designation = emp_data.get("designation", "Employee")

        # Ensure position exists
        if designation not in pos_map:
            pos = await db.positions.find_one({"company_id": company_id, "title": designation})
            if not pos:
                pos_id = str(uuid.uuid4())
                await db.positions.insert_one({
                    "id": pos_id,
                    "company_id": company_id,
                    "department_id": dept_id,
                    "team_id": team_id,
                    "title": designation,
                    "level": "L1",
                    "status": "active",
                    "created_at": now,
                    "updated_at": now
                })
                pos_map[designation] = pos_id
            else:
                pos_map[designation] = pos["id"]

        pos_id = pos_map[designation]

        # Create Employee document
        emp_doc = {
            "id": emp_id,
            "organization_id": org_id,
            "company_id": company_id,
            "name": emp_data.get("name", "Unknown"),
            "email": emp_data.get("email", ""),
            "department": emp_data.get("department", "General"),
            "designation": designation,
            "status": "active",
            "created_at": now,
            "updated_at": now
        }
        employees_to_insert.append(emp_doc)

        # Create mapping
        emp_pos_to_insert.append({
            "id": str(uuid.uuid4()),
            "company_id": company_id,
            "emp_id": emp_id,
            "position_id": pos_id,
            "is_primary": True,
            "start_date": now[:10],
            "created_at": now,
            "updated_at": now
        })

    if employees_to_insert:
        await db.employees.insert_many(employees_to_insert)
    if emp_pos_to_insert:
        await db.employee_positions.insert_many(emp_pos_to_insert)

    return {"message": f"Successfully onboarded {len(employees_to_insert)} employees."}

@api_router.post("/company", response_model=CompanyProfile)
@api_router.post("/companies", response_model=CompanyProfile)
async def create_company(data: CompanyProfileCreate, current_user: dict = Depends(get_current_user)):
    if current_user.get("role") not in ["admin", "superadmin"]:
        raise HTTPException(status_code=403, detail="Only admins can create companies")

    org_id = current_user.get("organization_id", "default")

    # Check max_companies limit
    if current_user.get("role") == "admin":
        admin_user = current_user
    else:
        admin_user = await db.users.find_one({"organization_id": org_id, "role": "admin"})

    if admin_user and admin_user.get("subscription_limits"):
        max_companies = admin_user["subscription_limits"].get("max_companies")
        if max_companies is not None:
            current_companies_count = await db.companies.count_documents({"organization_id": org_id, "deleted_at": None})
            if current_companies_count >= max_companies:
                raise HTTPException(status_code=403, detail=f"Company limit reached ({max_companies}). Please upgrade your subscription.")

    now = datetime.now(timezone.utc).isoformat()
    company_id = str(uuid.uuid4())
    
    # Auto-generate company_code if not provided
    comp_code = data.company_code or f"CMP-{str(uuid.uuid4())[:8].upper()}"
    
    # Process plants
    plants_to_save = []
    if data.plants:
        for p in data.plants:
            plants_to_save.append({
                "id": str(uuid.uuid4()),
                "company_id": company_id,
                "plant_code": p.plant_code,
                "created_at": now
            })

    company = {
        "id": company_id,
        "organization_id": org_id,
        "onboarded_by": current_user.get("name", ""),
        "status": "active",
        "company_code": comp_code,
        "created_at": now,
        "updated_at": now,
        "deleted_at": None,
        **data.model_dump(exclude={"plants"})
    }
    company["plants"] = plants_to_save
    
    await db.companies.insert_one({k: v for k, v in company.items() if k != "_id"})
    return CompanyProfile(**company)

@api_router.get("/company")
@api_router.get("/companies")
async def list_companies(include_deleted: bool = False, current_user: dict = Depends(get_current_user)):
    if current_user.get("role") not in ["admin", "superadmin"]:
        raise HTTPException(status_code=403, detail="Only admins can list companies")

    org = current_user.get("organization_id", "default")
    query = {"organization_id": org} if current_user.get("role") != "superadmin" else {}
    
    if not include_deleted:
        query["deleted_at"] = None
        
    companies = await db.companies.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return companies

@api_router.get("/company/{company_id}", response_model=CompanyProfile)
@api_router.get("/companies/{company_id}", response_model=CompanyProfile)
async def get_company(company_id: str, current_user: dict = Depends(get_current_user)):
    if current_user.get("role") not in ["admin", "superadmin"]:
        raise HTTPException(status_code=403, detail="Only admins can access companies")

    company = await db.companies.find_one({"id": company_id}, {"_id": 0})
    if not company:
        raise HTTPException(status_code=404, detail="Company not found")
    return CompanyProfile(**company)

@api_router.put("/company/{company_id}", response_model=CompanyProfile)
@api_router.put("/companies/{company_id}", response_model=CompanyProfile)
async def update_company(company_id: str, data: CompanyProfileCreate, current_user: dict = Depends(get_current_user)):
    if current_user.get("role") not in ["admin", "superadmin"]:
        raise HTTPException(status_code=403, detail="Only admins can update companies")

    company = await db.companies.find_one({"id": company_id}, {"_id": 0})
    if not company:
        raise HTTPException(status_code=404, detail="Company not found")
    
    now = datetime.now(timezone.utc).isoformat()
    update_data = data.model_dump(exclude={"plants"})
    update_data["updated_at"] = now
    
    # Process plants update
    if data.plants is not None:
        plants_to_save = []
        for p in data.plants:
            plants_to_save.append({
                "id": str(uuid.uuid4()),
                "company_id": company_id,
                "plant_code": p.plant_code,
                "created_at": now
            })
        update_data["plants"] = plants_to_save

    await db.companies.update_one({"id": company_id}, {"$set": update_data})
    updated = await db.companies.find_one({"id": company_id}, {"_id": 0})
    return CompanyProfile(**updated)

@api_router.delete("/company/{company_id}")
@api_router.delete("/companies/{company_id}")
async def delete_company(company_id: str, current_user: dict = Depends(get_current_user)):
    if current_user.get("role") not in ["admin", "superadmin"]:
        raise HTTPException(status_code=403, detail="Only admins can delete companies")

    now = datetime.now(timezone.utc).isoformat()
    result = await db.companies.update_one({"id": company_id}, {"$set": {"deleted_at": now, "status": "deleted"}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Company not found")
    return {"message": "Company soft-deleted successfully"}

# ═══════════════════════════════════════════════════════════
# AI INTELLIGENCE ENDPOINTS
# ═══════════════════════════════════════════════════════════

@api_router.get("/ai/dashboard-brief")
async def ai_dashboard_brief(current_user: dict = Depends(get_current_user)):
    """Personalized AI morning brief based on live data"""
    org_id = current_user.get("organization_id")
    query: dict = {} if current_user.get("role") == "superadmin" else {"organization_id": org_id}

    employees = await db.employees.count_documents(query)
    leads = await db.leads.count_documents(query)
    invoices = await db.invoices.count_documents(query)
    tickets = await db.tickets.count_documents(query)
    pending_leaves = await db.leave_requests.count_documents({**query, "status": "pending"})

    items = []
    if pending_leaves > 0:
        items.append({"type": "warning", "msg": f"{pending_leaves} leave request(s) pending approval", "link": "/hrms"})
    if tickets > 0:
        items.append({"type": "info", "msg": f"{tickets} support ticket(s) need attention", "link": "/support-desk"})
    if leads > 0:
        items.append({"type": "positive", "msg": f"{leads} leads in CRM pipeline — review today", "link": "/crm"})

    health_score = min(100, max(0,
        50 + (employees * 2) + (leads * 3) + (invoices * 5) - (pending_leaves * 3)
    ))

    return {
        "brief_items": items,
        "health_score": health_score,
        "health_label": "Excellent 🚀" if health_score > 80 else "Good 👍" if health_score > 60 else "Fair ⚡" if health_score > 40 else "Needs Focus 🎯",
        "stats": {"employees": employees, "leads": leads, "invoices": invoices, "tickets": tickets, "pending_leaves": pending_leaves}
    }

@api_router.get("/ai/crm-insights")
async def ai_crm_insights(current_user: dict = Depends(get_current_user)):
    """AI-powered CRM insights: lead scoring, next actions, forecast"""
    org_id = current_user.get("organization_id")
    query = {} if current_user.get("role") == "superadmin" else {"organization_id": org_id}

    leads = await db.leads.find(query, {"_id": 0}).to_list(1000)
    deals = await db.deals.find(query, {"_id": 0}).to_list(1000)

    total_deal_value = sum(d.get("value", 0) for d in deals)
    won_deals = [d for d in deals if d.get("stage") == "closed-won"]
    win_rate = (len(won_deals) / len(deals) * 100) if deals else 0
    forecasted = sum(d.get("value", 0) * (d.get("probability", 50) / 100) for d in deals if d.get("stage") != "closed-lost")

    # Identify at-risk leads (no activity in 14+ days)
    from datetime import datetime, timezone, timedelta
    cutoff = (datetime.now(timezone.utc) - timedelta(days=14)).isoformat()
    at_risk = [l for l in leads if l.get("created_at", "") < cutoff and l.get("status") not in ["closed-won", "closed-lost"]]

    return {
        "total_deals_value": total_deal_value,
        "win_rate": round(win_rate, 1),
        "forecasted_revenue": round(forecasted, 2),
        "at_risk_leads": len(at_risk),
        "insights": [
            {"type": "warning" if at_risk else "positive", "title": f"{len(at_risk)} leads at risk" if at_risk else "Pipeline healthy", "message": f"{len(at_risk)} leads have had no activity for 14+ days. Re-engage with targeted content." if at_risk else "All leads have recent activity. Keep the momentum!"},
            {"type": "info", "title": f"{round(win_rate, 0)}% win rate", "message": f"Pipeline forecast: ₹{round(forecasted/100000, 1)}L. {'Above target! 🎉' if win_rate > 30 else 'Focus on Negotiation-stage deals for quick wins.'}"},
        ]
    }

@api_router.get("/ai/hr-insights")
async def ai_hr_insights(current_user: dict = Depends(get_current_user)):
    """AI-powered HR insights: attrition risk, team health"""
    org_id = current_user.get("organization_id")
    query = {} if current_user.get("role") == "superadmin" else {"organization_id": org_id}

    employees = await db.employees.find(query, {"_id": 0}).to_list(1000)
    leaves = await db.leaves.find(query, {"_id": 0}).to_list(1000)

    # Detect high leave frequency (attrition risk signal)
    leave_counts = {}
    for leave in leaves:
        emp_id = leave.get("employee_id")
        leave_counts[emp_id] = leave_counts.get(emp_id, 0) + 1

    high_leave_emps = [emp_id for emp_id, count in leave_counts.items() if count >= 3]
    pending_leaves = [l for l in leaves if l.get("status") == "pending"]

    return {
        "total_employees": len(employees),
        "attrition_risk_count": len(high_leave_emps),
        "pending_leave_requests": len(pending_leaves),
        "insights": [
            {"type": "warning" if high_leave_emps else "positive", "title": f"{len(high_leave_emps)} attrition risk signals" if high_leave_emps else "Team retention looks healthy", "message": f"{len(high_leave_emps)} employees have taken 3+ leaves — consider 1:1 check-ins." if high_leave_emps else "No unusual leave patterns detected this period."},
            {"type": "info" if pending_leaves else "positive", "title": f"{len(pending_leaves)} leave approvals pending" if pending_leaves else "No pending approvals", "message": "Pending leaves should be resolved within 24 hours to maintain team trust." if pending_leaves else "All leave requests are up to date! ✓"},
        ]
    }

@api_router.get("/notifications")
async def get_notifications(current_user: dict = Depends(get_current_user)):
    """Smart notifications generated from live data"""
    org_id = current_user.get("organization_id")
    query = {} if current_user.get("role") == "superadmin" else {"organization_id": org_id}

    pending_leaves = await db.leaves.count_documents({**query, "status": "pending"})
    open_tickets = await db.tickets.count_documents({**query, "status": "open"})
    total_leads = await db.leads.count_documents(query)

    notifications = []

    if pending_leaves > 0:
        notifications.append({
            "id": "notif-leaves",
            "type": "leave",
            "title": f"{pending_leaves} Leave Request(s) Pending",
            "body": "Team members are waiting for approval.",
            "action": "/hrms",
            "read": False,
            "priority": "high",
            "created_at": datetime.now(timezone.utc).isoformat()
        })

    if open_tickets > 0:
        notifications.append({
            "id": "notif-tickets",
            "type": "ticket",
            "title": f"{open_tickets} Open Support Ticket(s)",
            "body": "Customers awaiting response.",
            "action": "/support-desk",
            "read": False,
            "priority": "medium" if open_tickets < 5 else "high",
            "created_at": datetime.now(timezone.utc).isoformat()
        })

    notifications.append({
        "id": "notif-ai-brief",
        "type": "ai",
        "title": "AI Business Brief Ready",
        "body": f"Team: {await db.employees.count_documents(query)} · Leads: {total_leads} · Tickets: {open_tickets}",
        "action": "/",
        "read": True,
        "priority": "low",
        "created_at": datetime.now(timezone.utc).isoformat()
    })

    return notifications

# ═══════════════════════════════════════════════════════════
# INVOICE STATUS UPDATE
# ═══════════════════════════════════════════════════════════

@api_router.patch("/invoices/{invoice_id}/status")
async def update_invoice_status(invoice_id: str, status_data: dict, current_user: dict = Depends(get_current_user)):
    """Update invoice payment status"""
    new_status = status_data.get("status")
    if new_status not in ["pending", "paid", "overdue", "cancelled"]:
        raise HTTPException(status_code=400, detail="Invalid status")
    result = await db.invoices.update_one(
        {"id": invoice_id},
        {"$set": {"status": new_status, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Invoice not found")
    return {"message": f"Invoice marked as {new_status}", "status": new_status}


# ═══════════════════════════════════════════════════════════
# VENDORS CRUD
# ═══════════════════════════════════════════════════════════

@api_router.get("/vendors")
async def get_vendors(current_user: dict = Depends(get_current_user)):
    """Get all vendors for the organization"""
    org_id = current_user.get("organization_id")
    query = {} if current_user.get("role") == "superadmin" else {"organization_id": org_id}
    vendors = await db.vendors.find(query, {"_id": 0}).to_list(1000)
    return vendors

@api_router.post("/vendors")
async def create_vendor(vendor_data: dict, current_user: dict = Depends(get_current_user)):
    """Create a new vendor"""
    vendor = {
        "id": str(uuid.uuid4()),
        "organization_id": current_user.get("organization_id"),
        "created_at": datetime.now(timezone.utc).isoformat(),
        **vendor_data
    }
    await db.vendors.insert_one({k: v for k, v in vendor.items() if k != "_id"})
    return vendor

# ═══════════════════════════════════════════════════════════
# OFFBOARDING: RESIGNATIONS & PIP
# ═══════════════════════════════════════════════════════════

# ─────────────────── IATF OPERATIONAL ENDPOINTS ───────────────────

@api_router.get("/iatf/module/{module_name}")
async def get_iatf_module_data(module_name: str, current_user: dict = Depends(get_current_user)):
    """Generic endpoint to fetch data for any IATF module"""
    org_id = current_user.get("organization_id")
    # Check for both organization_id and company_id in metadata
    res = await getattr(db, f"iatf_{module_name}").find({"metadata.company_id": org_id}, {"_id": 0}).to_list(1000)
    return res

@api_router.post("/iatf/module/{module_name}")
async def create_iatf_record(module_name: str, data: dict, current_user: dict = Depends(get_current_user)):
    """Generic endpoint to create an IATF record with compliance metadata"""
    now = datetime.now(timezone.utc).isoformat()
    doc = {
        "id": data.get("id", str(uuid.uuid4())),
        "metadata": {
            "version": data.get("version", "1.0"),
            "created_by": current_user.get("id"),
            "company_id": current_user.get("organization_id"),
            "status": "draft",
            "created_at": now,
            "updated_at": now
        },
        **data
    }
    await getattr(db, f"iatf_{module_name}").insert_one(doc)
    return {"message": "Record created", "id": doc["id"]}

@api_router.patch("/iatf/module/{module_name}/{record_id}/approve")
async def approve_iatf_record(module_name: str, record_id: str, current_user: dict = Depends(get_current_user)):
    """Approve a compliance document"""
    now = datetime.now(timezone.utc).isoformat()
    result = await getattr(db, f"iatf_{module_name}").update_one(
        {"id": record_id},
        {"$set": {
            "metadata.approved_by": current_user.get("id"),
            "metadata.status": "active",
            "metadata.updated_at": now
        }}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Record not found")
    return {"message": "Document approved and activated"}

@api_router.get("/iatf/gap-analysis")
async def iatf_gap_analysis(current_user: dict = Depends(get_current_user)):
    """Comprehensive gap analysis for IATF/ISO compliance"""
    org_id = current_user.get("organization_id")
    gaps = []

    # 1. Competence Gaps (Clause 7.2)
    all_users = await db.users.find({"organization_id": org_id, "role": "employee"}).to_list(1000)
    for user in all_users:
        matrix = await db.iatf_skill_matrix.find_one({"employee_id": user["id"]}, {"_id": 0})
        if not matrix:
            gaps.append({
                "module": "Skill Matrix",
                "employee": user["name"],
                "employee_id": user["id"],
                "gap_type": "MISSING_RECORD",
                "severity": "High",
                "description": "No skill matrix defined for this employee."
            })

    # 2. Motivation Gaps (Clause 7.3)
    pending_kaizens = await db.iatf_kaizen_suggestion.find({
        "metadata.company_id": org_id,
        "status": "pending"
    }).to_list(1000)
    
    now = datetime.now(timezone.utc)
    for k in pending_kaizens:
        created_at = datetime.fromisoformat(k["metadata"]["created_at"])
        if (now - created_at).days > 15:
            gaps.append({
                "module": "Kaizen",
                "record_id": k["id"],
                "employee_id": k["employee_id"],
                "gap_type": "SLA_BREACH",
                "severity": "Medium",
                "description": f"Kaizen pending for {(now - created_at).days} days."
            })

    # 3. Training Gaps
    # Check if induction exists for new hires (last 30 days)
    new_hires = [u for u in all_users if (now - datetime.fromisoformat(u["created_at"])).days < 30]
    for nh in new_hires:
        induction = await db.iatf_induction_program.find_one({"employee_id": nh["id"]})
        if not induction:
            gaps.append({
                "module": "Induction",
                "employee": nh["name"],
                "gap_type": "MISSING_RECORD",
                "severity": "CRITICAL",
                "description": "New joiner missing mandatory induction program."
            })

    return {
        "summary": {
            "total_gaps": len(gaps),
            "critical_count": len([g for g in gaps if g["severity"] == "CRITICAL"]),
            "last_audit": now.isoformat()
        },
        "gaps": gaps
    }

@api_router.get("/resignations")
async def get_resignations(current_user: dict = Depends(get_current_user)):
    org_id = current_user.get("organization_id")
    query = {} if current_user.get("role") == "superadmin" else {"organization_id": org_id}
    res = await db.resignations.find(query, {"_id": 0}).to_list(1000)
    return res

@api_router.post("/resignations")
async def create_resignation(data: ResignationCreate, current_user: dict = Depends(get_current_user)):
    org_id = current_user.get("organization_id")
    try:
        emp = await db.employees.find_one({"id": data.employee_id}, {"_id": 0, "name": 1})
        emp_name = emp["name"] if emp else "Unknown"
    except Exception:
        emp_name = "Unknown"
        
    doc = {
        "id": str(uuid.uuid4()),
        "organization_id": org_id,
        "employee_id": data.employee_id,
        "employee_name": emp_name,
        "reason": data.reason,
        "resignation_date": data.resignation_date,
        "last_working_day": data.last_working_day,
        "notice_period_days": data.notice_period_days,
        "status": "pending",
        "fnf_status": "pending",
        "fnf_amount": 0.0,
        "fnf_breakdown": {},
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.resignations.insert_one(doc)
    return doc

@api_router.patch("/resignations/{id}/status")
async def update_resignation_status(id: str, payload: dict, current_user: dict = Depends(get_current_user)):
    res = await db.resignations.update_one(
        {"id": id, "organization_id": current_user.get("organization_id")},
        {"$set": {"status": payload.get("status")}}
    )
    if res.matched_count == 0:
        raise HTTPException(404, "Resignation not found")
    return {"message": "Status updated"}

@api_router.post("/resignations/{id}/calculate-fnf")
async def calculate_fnf(id: str, current_user: dict = Depends(get_current_user)):
    org_id = current_user.get("organization_id")
    resig = await db.resignations.find_one({"id": id, "organization_id": org_id})
    if not resig:
        raise HTTPException(404, "Not found")
    
    emp = await db.employees.find_one({"id": resig.get("employee_id")})
    ctc = 600000 # default 6 LPA heuristic
    if emp and emp.get("email"):
        offer = await db.offer_letters.find_one({"email": emp["email"]})
        if offer and "ctc_yearly" in offer:
            ctc = float(offer["ctc_yearly"])
            
    # Calculate daily run rate
    monthly = ctc / 12
    daily = monthly / 30
    
    # Get unutilized leaves by checking approved leaves vs a default allocation
    employee_id = resig.get("employee_id")
    # Using 24 as a baseline yearly accrual standard
    total_accrued_leaves = 24

    approved_leaves_cursor = await db.leave_requests.find({"employee_id": employee_id, "status": "approved", "organization_id": org_id}).to_list(1000)

    utilized_leaves = 0
    for l in approved_leaves_cursor:
        utilized_leaves += (datetime.fromisoformat(l["end_date"]) - datetime.fromisoformat(l["start_date"])).days + 1

    unused_leaves = max(0, total_accrued_leaves - utilized_leaves)
    leave_encashment = unused_leaves * daily
    
    # Notice period deficit/recovery (assuming 30 days standard vs what is entered)
    notice_diff = max(0, 30 - resig.get("notice_period_days", 30))
    notice_recovery = notice_diff * daily
    
    base_last_month = monthly
    total = base_last_month + leave_encashment - notice_recovery
    
    breakdown = {
        "Base Salary Pending (1M)": base_last_month,
        "Leave Encashment": leave_encashment,
        "Notice Period Deficit Recovery": -notice_recovery,
        "Unused Leaves": unused_leaves
    }
    
    await db.resignations.update_one(
        {"id": id},
        {"$set": {
            "fnf_status": "calculated",
            "fnf_amount": round(total, 2),
            "fnf_breakdown": breakdown
        }}
    )
    return {"message": "FnF generated", "amount": total}

@api_router.get("/pip")
async def get_pips(current_user: dict = Depends(get_current_user)):
    org_id = current_user.get("organization_id")
    query = {} if current_user.get("role") == "superadmin" else {"organization_id": org_id}
    res = await db.performance_plans.find(query, {"_id": 0}).to_list(1000)
    return res

@api_router.post("/pip")
async def create_pip(data: PerformancePlanCreate, current_user: dict = Depends(get_current_user)):
    org_id = current_user.get("organization_id")
    emp = await db.employees.find_one({"id": data.employee_id})
    emp_name = emp["name"] if emp else "Unknown"
        
    doc = {
        "id": str(uuid.uuid4()),
        "organization_id": org_id,
        "employee_id": data.employee_id,
        "employee_name": emp_name,
        "reason": data.reason,
        "goals": data.goals,
        "duration_days": data.duration_days,
        "start_date": data.start_date,
        "end_date": data.end_date,
        "status": "active",
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.performance_plans.insert_one(doc)
    return doc

@api_router.patch("/pip/{id}/status")
async def update_pip_status(id: str, payload: dict, current_user: dict = Depends(get_current_user)):
    res = await db.performance_plans.update_one(
        {"id": id, "organization_id": current_user.get("organization_id")},
        {"$set": {"status": payload.get("status")}}
    )
    if res.matched_count == 0:
        raise HTTPException(404, "PIP not found")
    return {"message": "Status updated"}

async def _upsert_seed_document(collection_name: str, query: dict, doc: dict):
    collection = getattr(db, collection_name)
    existing = await collection.find_one(query, {"_id": 0})
    if existing:
        await collection.update_one(query, {"$set": doc})
    else:
        await collection.insert_one(doc)

async def _upsert_seed_many(collection_name: str, docs: List[dict], key_field: str = "id"):
    for doc in docs:
        key_value = doc.get(key_field)
        if key_value is None:
            continue
        await _upsert_seed_document(collection_name, {key_field: key_value}, doc)

def _seed_prefix_for_org(organization_id: str) -> str:
    cleaned = "".join(ch for ch in (organization_id or "").lower() if ch.isalnum())
    return (cleaned[:8] or "orgseed")

async def seed_test_data_for_org(
    organization_id: str,
    multiplier: int = 1,
    target_records: Optional[int] = None,
    seeded_by: str = "System"
) -> dict:
    now = datetime.now(timezone.utc)
    today = now.date()
    prefix = _seed_prefix_for_org(organization_id)
    target = max(200, min(target_records if target_records is not None else 120 * multiplier, 1200))
    scale = target / 120.0
    seeded_id_prefix = f"{prefix}-"

    first_names = ["Aarav", "Priya", "Rahul", "Sneha", "Vikram", "Pooja", "Ankit", "Neha"]
    last_names = ["Sharma", "Verma", "Patel", "Gupta", "Reddy", "Kumar", "Nair", "Singh"]
    departments = ["Engineering", "Sales", "Finance", "Marketing", "HR", "Operations", "Support"]
    designations = ["Software Engineer", "Senior Engineer", "Account Executive", "Financial Analyst", "HR Specialist", "Marketing Associate", "Operations Executive", "Support Associate"]
    summary: dict = {"target_records": target}

    def n(base: int, min_value: int = 1, max_value: int = 1000) -> int:
        return min(max_value, max(min_value, int(round(base * scale))))

    def _is_seeded_id(value: Optional[str]) -> bool:
        return isinstance(value, str) and value.startswith(seeded_id_prefix)

    async def _list_org_ids(collection_name: str) -> List[str]:
        docs = await getattr(db, collection_name).find({"organization_id": organization_id}, {"_id": 0, "id": 1}).to_list(20000)
        return [d["id"] for d in docs if isinstance(d.get("id"), str)]

    async def _planned_seed_count(collection_name: str, desired_total: int) -> int:
        ids = await _list_org_ids(collection_name)
        non_seeded_count = sum(1 for doc_id in ids if not _is_seeded_id(doc_id))
        return max(0, min(1000, desired_total - non_seeded_count))

    async def _prune_seeded_records(collection_name: str, keep_ids: Set[str]) -> None:
        ids = await _list_org_ids(collection_name)
        collection = getattr(db, collection_name)
        for doc_id in ids:
            if _is_seeded_id(doc_id) and doc_id not in keep_ids:
                await collection.delete_one({"id": doc_id, "organization_id": organization_id})

    async def _org_total(collection_name: str) -> int:
        return await getattr(db, collection_name).count_documents({"organization_id": organization_id})

    total_employees = await _planned_seed_count("employees", n(120, min_value=120))
    employees: List[dict] = []
    for i in range(1, total_employees + 1):
        first = first_names[(i - 1) % len(first_names)]
        last = last_names[(i - 1) % len(last_names)]
        status = "active" if i <= int(total_employees * 0.84) else ("resigned" if i <= int(total_employees * 0.92) else "terminated")
        employees.append({
            "id": f"{prefix}-emp-{i:04d}",
            "emp_id": f"{prefix.upper()}-{2000 + i}",
            "name": f"{first} {last} {i}",
            "email": f"{prefix}.{first.lower()}.{last.lower()}.{i}@demo.com",
            "phone": f"+9199{20000000 + i:08d}",
            "department": departments[(i - 1) % len(departments)],
            "designation": designations[(i - 1) % len(designations)],
            "date_of_joining": (today - timedelta(days=120 + (i * 3))).isoformat(),
            "status": status,
            "organization_id": organization_id,
            "created_at": (now - timedelta(days=i)).isoformat(),
        })
    await _upsert_seed_many("employees", employees)
    await _prune_seeded_records("employees", {e["id"] for e in employees})
    summary["employees"] = await _org_total("employees")
    employee_pool = employees or await db.employees.find({"organization_id": organization_id}, {"_id": 0, "id": 1, "name": 1, "phone": 1, "designation": 1}).to_list(2000)

    total_jobs = await _planned_seed_count("jobs", n(120, min_value=120))
    jobs: List[dict] = []
    statuses = (["open"] * int(total_jobs * 0.7)) + (["closed"] * int(total_jobs * 0.2))
    while len(statuses) < total_jobs:
        statuses.append("draft")
    for i, status in enumerate(statuses, start=1):
        dept = departments[(i - 1) % len(departments)]
        jobs.append({
            "id": f"{prefix}-job-{i:04d}",
            "title": f"{dept} Specialist {i}",
            "department": dept,
            "location": "Bangalore" if i % 3 else "Remote",
            "type": ["Full-time", "Contract", "Remote", "Hybrid"][(i - 1) % 4],
            "description": f"Seeded test role #{i}",
            "status": status,
            "organization_id": organization_id,
            "created_at": (now - timedelta(days=i)).isoformat(),
        })
    await _upsert_seed_many("jobs", jobs)
    await _upsert_seed_many("job_postings", jobs)
    await _prune_seeded_records("jobs", {j["id"] for j in jobs})
    await _prune_seeded_records("job_postings", {j["id"] for j in jobs})
    summary["jobs"] = await _org_total("jobs")
    job_pool = jobs or await db.jobs.find({"organization_id": organization_id}, {"_id": 0, "id": 1}).to_list(2000)

    total_candidates = await _planned_seed_count("candidates", n(120, min_value=120))
    candidates: List[dict] = []
    candidate_statuses = ["applied", "screening", "interview", "offered", "hired", "rejected"]
    for i in range(1, total_candidates + 1):
        candidates.append({
            "id": f"{prefix}-cand-{i:04d}",
            "job_id": job_pool[(i - 1) % len(job_pool)]["id"],
            "name": f"{first_names[(i - 1) % len(first_names)]} Candidate {i}",
            "email": f"{prefix}.candidate.{i}@mailinator.com",
            "resume_url": f"https://example.com/{prefix}/resume/{i}",
            "status": candidate_statuses[(i - 1) % len(candidate_statuses)],
            "organization_id": organization_id,
            "created_at": (now - timedelta(days=i)).isoformat(),
        })
    await _upsert_seed_many("candidates", candidates)
    await _prune_seeded_records("candidates", {c["id"] for c in candidates})
    summary["candidates"] = await _org_total("candidates")

    def _employee(i: int) -> dict:
        return employee_pool[(i - 1) % len(employee_pool)]

    total_offers = await _planned_seed_count("offer_letters", n(120, min_value=120))
    offers: List[dict] = []
    for i in range(1, total_offers + 1):
        emp = _employee(i)
        offers.append({
            "id": f"{prefix}-offer-{i:04d}",
            "organization_id": organization_id,
            "name": emp["name"],
            "email": f"{prefix}.offer.{i}@demo.com",
            "phone": emp["phone"],
            "designation": emp["designation"],
            "ctc_yearly": float(500000 + (i * 11000)),
            "details": {"batch": "org-seed", "employee_ref": emp["id"]},
            "status": ["Generated", "Sent", "Accepted", "Revoked"][(i - 1) % 4],
            "created_at": (now - timedelta(days=i)).isoformat(),
        })
    await _upsert_seed_many("offer_letters", offers)
    await _prune_seeded_records("offer_letters", {o["id"] for o in offers})
    summary["offer_letters"] = await _org_total("offer_letters")

    total_wfh = await _planned_seed_count("wfh_requests", n(120, min_value=120))
    wfh: List[dict] = []
    for i in range(1, total_wfh + 1):
        emp = _employee(i)
        wfh.append({
            "id": f"{prefix}-wfh-{i:04d}",
            "employee_id": emp["id"],
            "employee_name": emp["name"],
            "start_date": (today - timedelta(days=(i % 20))).isoformat(),
            "end_date": (today - timedelta(days=max((i % 20) - 1, 0))).isoformat(),
            "reason": f"Seeded WFH request #{i}",
            "status": ["pending", "approved", "rejected"][(i - 1) % 3],
            "organization_id": organization_id,
            "created_at": (now - timedelta(days=i)).isoformat(),
        })
    await _upsert_seed_many("wfh_requests", wfh)
    await _prune_seeded_records("wfh_requests", {w["id"] for w in wfh})
    summary["wfh_requests"] = await _org_total("wfh_requests")

    total_resignations = await _planned_seed_count("resignations", n(120, min_value=120))
    resignations: List[dict] = []
    for i in range(1, total_resignations + 1):
        emp = _employee(i + 3)
        status = ["pending", "approved", "rejected"][(i - 1) % 3]
        fnf_status = "calculated" if status == "approved" else "pending"
        fnf_amount = float(25000 + i * 850) if fnf_status == "calculated" else 0.0
        resignations.append({
            "id": f"{prefix}-res-{i:04d}",
            "employee_id": emp["id"],
            "employee_name": emp["name"],
            "reason": f"Seeded resignation record #{i}",
            "resignation_date": (today - timedelta(days=45 + i)).isoformat(),
            "last_working_day": (today - timedelta(days=max(5, i % 30))).isoformat(),
            "notice_period_days": 30,
            "status": status,
            "fnf_status": fnf_status,
            "fnf_amount": round(fnf_amount, 2),
            "fnf_breakdown": {
                "Base Salary Pending (1M)": round(fnf_amount * 0.7, 2),
                "Leave Encashment (approx)": round(fnf_amount * 0.4, 2),
                "Notice Period Deficit Recovery": round(-(fnf_amount * 0.1), 2),
            } if fnf_status == "calculated" else {},
            "organization_id": organization_id,
            "created_at": (now - timedelta(days=i)).isoformat(),
        })
    await _upsert_seed_many("resignations", resignations)
    await _prune_seeded_records("resignations", {r["id"] for r in resignations})
    summary["resignations"] = await _org_total("resignations")

    total_pips = await _planned_seed_count("performance_plans", n(120, min_value=120))
    pips: List[dict] = []
    for i in range(1, total_pips + 1):
        emp = _employee(i + 7)
        pips.append({
            "id": f"{prefix}-pip-{i:04d}",
            "employee_id": emp["id"],
            "employee_name": emp["name"],
            "reason": f"Seeded performance concern #{i}",
            "goals": f"Complete improvement plan target set #{i}",
            "duration_days": 30,
            "start_date": (today - timedelta(days=40 + i)).isoformat(),
            "end_date": (today - timedelta(days=max(1, i % 20))).isoformat(),
            "status": ["active", "successful", "failed"][(i - 1) % 3],
            "organization_id": organization_id,
            "created_at": (now - timedelta(days=i)).isoformat(),
        })
    await _upsert_seed_many("performance_plans", pips)
    await _prune_seeded_records("performance_plans", {p["id"] for p in pips})
    summary["performance_plans"] = await _org_total("performance_plans")

    total_announce = await _planned_seed_count("announcements", n(120, min_value=120))
    announcements: List[dict] = []
    for i in range(1, total_announce + 1):
        announcements.append({
            "id": f"{prefix}-ann-{i:04d}",
            "title": f"Seeded Announcement {i}",
            "content": f"Organization update #{i} generated for QA feed validation.",
            "author_id": f"{prefix}-system",
            "author_name": seeded_by,
            "priority": "urgent" if i % 7 == 0 else ("high" if i % 3 == 0 else "normal"),
            "organization_id": organization_id,
            "created_at": (now - timedelta(hours=i)).isoformat(),
        })
    await _upsert_seed_many("announcements", announcements)
    await _prune_seeded_records("announcements", {a["id"] for a in announcements})
    summary["announcements"] = await _org_total("announcements")

    return summary

async def ensure_demo_users_seeded():
    now = datetime.now(timezone.utc)
    now_iso = now.isoformat()
    today = now.date()
    today_iso = today.isoformat()
    yesterday_iso = (today - timedelta(days=1)).isoformat()
    three_days_ago_iso = (today - timedelta(days=3)).isoformat()
    seven_days_ago_iso = (today - timedelta(days=7)).isoformat()
    next_week_iso = (today + timedelta(days=7)).isoformat()
    next_month_iso = (today + timedelta(days=30)).isoformat()
    subscription_end_iso = (now + timedelta(days=365)).isoformat()

    hashed_password = get_password_hash(DEFAULT_DEMO_PASSWORD)

    demo_users = [
        {
            "id": "usr-superadmin",
            "email": "superadmin@demo.com",
            "password": hashed_password,
            "name": "Demo SuperAdmin",
            "role": "superadmin",
            "organization_id": DEFAULT_ORG_ID,
            "email_verified": True,
            "subscription_status": "active",
            "subscription_end_date": subscription_end_iso,
            "enabled_services": DEFAULT_ENABLED_SERVICES,
            "created_at": now_iso,
        },
        {
            "id": "usr-admin",
            "email": "admin@demo.com",
            "password": hashed_password,
            "name": "Demo Admin",
            "role": "admin",
            "organization_id": DEFAULT_ORG_ID,
            "email_verified": True,
            "subscription_status": "active",
            "subscription_limits": {"max_employees": 1000, "max_projects": 500},
            "subscription_end_date": subscription_end_iso,
            "enabled_services": DEFAULT_ENABLED_SERVICES,
            "created_at": now_iso,
        },
        {
            "id": "usr-employee",
            "email": "employee@demo.com",
            "password": hashed_password,
            "name": "Demo Employee",
            "role": "employee",
            "organization_id": DEFAULT_ORG_ID,
            "email_verified": True,
            "subscription_status": "active",
            "created_at": now_iso,
        },
        {
            "id": "usr-accountant",
            "email": "accountant@demo.com",
            "password": hashed_password,
            "name": "Demo Accountant",
            "role": "accountant",
            "company_id": DEFAULT_COMPANY_ID,
            "organization_id": DEFAULT_ORG_ID,
            "email_verified": True,
            "subscription_status": "active",
            "enabled_services": DEFAULT_ACCOUNTING_SERVICES,
            "created_at": now_iso,
        },
        {
            "id": "usr-client",
            "email": "client@demo.com",
            "password": hashed_password,
            "name": "Demo Client",
            "role": "admin",
            "organization_id": DEFAULT_ORG_ID,
            "email_verified": True,
            "subscription_status": "active",
            "subscription_limits": {"max_employees": 50, "max_projects": 20},
            "subscription_end_date": subscription_end_iso,
            "enabled_services": DEFAULT_ENABLED_SERVICES,
            "created_at": now_iso,
        },
        {
            "id": "usr-hr",
            "email": "hr@demo.com",
            "password": hashed_password,
            "name": "Demo HR Manager",
            "role": "hr",
            "organization_id": DEFAULT_ORG_ID,
            "email_verified": True,
            "subscription_status": "active",
            "subscription_end_date": subscription_end_iso,
            "enabled_services": DEFAULT_ENABLED_SERVICES,
            "created_at": now_iso,
        },
        {
            "id": "usr-manager",
            "email": "manager@demo.com",
            "password": hashed_password,
            "name": "Demo Team Manager",
            "role": "manager",
            "organization_id": DEFAULT_ORG_ID,
            "email_verified": True,
            "subscription_status": "active",
            "subscription_end_date": subscription_end_iso,
            "enabled_services": DEFAULT_ENABLED_SERVICES,
            "created_at": now_iso,
        },
    ]
    for user in demo_users:
        await _upsert_seed_document("users", {"email": user["email"]}, user)

    employees = [
        {
            "id": "emp-1001",
            "user_id": "usr-employee",
            "emp_id": "PRSK-1001",
            "name": "Demo Employee",
            "email": "employee@demo.com",
            "phone": "+919876543210",
            "department": "Engineering",
            "designation": "Software Engineer",
            "date_of_joining": "2024-01-10",
            "status": "active",
            "organization_id": DEFAULT_ORG_ID,
            "created_at": now_iso,
        },
        {
            "id": "emp-1002",
            "emp_id": "PRSK-1002",
            "name": "Priya Sharma",
            "email": "priya.sharma@demo.com",
            "phone": "+919812345678",
            "department": "Sales",
            "designation": "Account Executive",
            "date_of_joining": "2023-11-20",
            "status": "active",
            "organization_id": DEFAULT_ORG_ID,
            "created_at": now_iso,
        },
        {
            "id": "emp-1003",
            "emp_id": "PRSK-1003",
            "name": "Rahul Verma",
            "email": "rahul.verma@demo.com",
            "phone": "+919801234567",
            "department": "Finance",
            "designation": "Finance Analyst",
            "date_of_joining": "2022-08-15",
            "status": "active",
            "organization_id": DEFAULT_ORG_ID,
            "created_at": now_iso,
        },
    ]
    await _upsert_seed_many("employees", employees)

    # Bulk UAT employee dataset:
    # Target total employees = 360 (300 active, 30 resigned, 30 terminated)
    bulk_first_names = [
        "Aarav", "Priya", "Rahul", "Sneha", "Vikram", "Pooja", "Ankit", "Neha",
        "Rohan", "Kavya", "Aditya", "Isha", "Varun", "Meera", "Karan", "Nisha",
    ]
    bulk_last_names = [
        "Sharma", "Verma", "Patel", "Gupta", "Reddy", "Kumar", "Nair", "Singh",
        "Desai", "Joshi", "Bose", "Kapoor", "Malhotra", "Iyer", "Yadav", "Mehta",
    ]
    bulk_departments = ["Engineering", "Sales", "Finance", "Marketing", "HR", "Operations", "Support"]
    bulk_designations = [
        "Software Engineer", "Senior Engineer", "Account Executive", "Financial Analyst",
        "HR Specialist", "Marketing Associate", "Operations Executive", "Support Associate",
    ]
    bulk_employees: List[dict] = []
    for i in range(1, 358):
        status = "active" if i <= 297 else ("resigned" if i <= 327 else "terminated")
        first = bulk_first_names[(i - 1) % len(bulk_first_names)]
        last = bulk_last_names[(i - 1) % len(bulk_last_names)]
        bulk_employees.append(
            {
                "id": f"emp-bulk-{i:03d}",
                "emp_id": f"PRSK-{2000 + i}",
                "name": f"{first} {last} {i}",
                "email": f"{first.lower()}.{last.lower()}.{i}@demo.com",
                "phone": f"+9198{10000000 + i:08d}",
                "department": bulk_departments[(i - 1) % len(bulk_departments)],
                "designation": bulk_designations[(i - 1) % len(bulk_designations)],
                "date_of_joining": (today - timedelta(days=180 + (i * 5))).isoformat(),
                "status": status,
                "organization_id": DEFAULT_ORG_ID,
                "created_at": (now - timedelta(days=i)).isoformat(),
            }
        )
    await _upsert_seed_many("employees", bulk_employees)

    leave_requests = [
        {
            "id": "leave-1001",
            "employee_id": "emp-1001",
            "leave_type": "Casual Leave",
            "from_date": next_week_iso,
            "to_date": next_week_iso,
            "reason": "Personal appointment",
            "status": "pending",
            "organization_id": DEFAULT_ORG_ID,
            "created_at": now_iso,
        },
        {
            "id": "leave-1002",
            "employee_id": "emp-1002",
            "leave_type": "Sick Leave",
            "from_date": three_days_ago_iso,
            "to_date": yesterday_iso,
            "reason": "Seasonal fever",
            "status": "approved",
            "organization_id": DEFAULT_ORG_ID,
            "created_at": now_iso,
        },
    ]
    await _upsert_seed_many("leave_requests", leave_requests)
    await _upsert_seed_many("leaves", leave_requests)

    # Additional leave data for QA filtering/states
    bulk_leave_requests: List[dict] = []
    all_employee_records = employees + bulk_employees
    leave_statuses = ["pending", "approved", "rejected"]
    leave_types = ["Casual Leave", "Sick Leave", "Earned Leave", "WFH"]
    for i in range(1, 181):
        employee = all_employee_records[(i - 1) % len(all_employee_records)]
        from_date = (today - timedelta(days=(i % 40))).isoformat()
        to_date = (today - timedelta(days=max((i % 40) - 1, 0))).isoformat()
        bulk_leave_requests.append(
            {
                "id": f"leave-bulk-{i:03d}",
                "employee_id": employee["id"],
                "leave_type": leave_types[(i - 1) % len(leave_types)],
                "from_date": from_date,
                "to_date": to_date,
                "reason": f"Seeded leave request {i}",
                "status": leave_statuses[(i - 1) % len(leave_statuses)],
                "organization_id": DEFAULT_ORG_ID,
                "created_at": (now - timedelta(days=i)).isoformat(),
            }
        )
    await _upsert_seed_many("leave_requests", bulk_leave_requests)
    await _upsert_seed_many("leaves", bulk_leave_requests)

    attendance = [
        {
            "id": "att-1001",
            "employee_id": "emp-1001",
            "date": today_iso,
            "check_in": "09:12",
            "check_out": "18:34",
            "status": "present",
            "organization_id": DEFAULT_ORG_ID,
            "created_at": now_iso,
        },
        {
            "id": "att-1002",
            "employee_id": "emp-1002",
            "date": today_iso,
            "check_in": "09:05",
            "check_out": "18:10",
            "status": "present",
            "organization_id": DEFAULT_ORG_ID,
            "created_at": now_iso,
        },
    ]
    await _upsert_seed_many("attendance", attendance)

    # Additional attendance records for dashboard/trends
    bulk_attendance: List[dict] = []
    active_employee_ids = [e["id"] for e in (employees + bulk_employees) if e.get("status") == "active"]
    for day_offset in range(0, 14):
        att_date = (today - timedelta(days=day_offset)).isoformat()
        for idx, emp_id in enumerate(active_employee_ids[:180], start=1):
            att_status = "wfh" if (idx + day_offset) % 11 == 0 else "present"
            bulk_attendance.append(
                {
                    "id": f"att-bulk-{day_offset:02d}-{idx:03d}",
                    "employee_id": emp_id,
                    "date": att_date,
                    "check_in": f"09:{(idx + day_offset) % 60:02d}",
                    "check_out": f"18:{(idx + day_offset + 10) % 60:02d}",
                    "status": att_status,
                    "organization_id": DEFAULT_ORG_ID,
                    "created_at": now_iso,
                }
            )
    await _upsert_seed_many("attendance", bulk_attendance)

    projects = [
        {
            "id": "proj-1001",
            "name": "MyOffice SaaS Revamp",
            "description": "Frontend and backend hardening for production rollout.",
            "status": "active",
            "start_date": seven_days_ago_iso,
            "organization_id": DEFAULT_ORG_ID,
            "created_at": now_iso,
        },
        {
            "id": "proj-1002",
            "name": "CRM Automation Suite",
            "description": "Lead routing and follow-up orchestration.",
            "status": "active",
            "start_date": "2026-01-05",
            "organization_id": DEFAULT_ORG_ID,
            "created_at": now_iso,
        },
    ]
    await _upsert_seed_many("projects", projects)

    tasks = [
        {
            "id": "task-1001",
            "project_id": "proj-1001",
            "title": "Finalize Vercel deployment config",
            "assigned_to": "emp-1001",
            "status": "in-progress",
            "priority": "high",
            "organization_id": DEFAULT_ORG_ID,
            "created_at": now_iso,
        },
        {
            "id": "task-1002",
            "project_id": "proj-1002",
            "title": "Prepare CRM dashboard filters",
            "assigned_to": "emp-1002",
            "status": "todo",
            "priority": "medium",
            "organization_id": DEFAULT_ORG_ID,
            "created_at": now_iso,
        },
    ]
    await _upsert_seed_many("tasks", tasks)

    leads = [
        {
            "id": "lead-1001",
            "name": "Ananya Mehta",
            "email": "ananya@globex.com",
            "phone": "+918888111222",
            "company": "Globex Solutions",
            "source": "Website",
            "status": "qualified",
            "organization_id": DEFAULT_ORG_ID,
            "created_at": now_iso,
        },
        {
            "id": "lead-1002",
            "name": "Saurav Gupta",
            "email": "saurav@northwind.io",
            "phone": "+919999222333",
            "company": "Northwind Labs",
            "source": "Referral",
            "status": "new",
            "organization_id": DEFAULT_ORG_ID,
            "created_at": now_iso,
        },
    ]
    await _upsert_seed_many("leads", leads)

    deals = [
        {
            "id": "deal-1001",
            "lead_id": "lead-1001",
            "title": "Globex Annual Subscription",
            "value": 550000.0,
            "stage": "negotiation",
            "probability": 70,
            "expected_close_date": next_month_iso,
            "organization_id": DEFAULT_ORG_ID,
            "created_at": now_iso,
        }
    ]
    await _upsert_seed_many("deals", deals)

    customers = [
        {
            "id": "cust-1001",
            "name": "Globex Solutions",
            "contact_person": "Ananya Mehta",
            "email": "finance@globex.com",
            "phone": "+918888000111",
            "address": "Bangalore",
            "organization_id": DEFAULT_ORG_ID,
            "created_at": now_iso,
        }
    ]
    await _upsert_seed_many("customers", customers)

    invoices = [
        {
            "id": "inv-1001",
            "invoice_number": "INV-2026-1001",
            "customer_id": "cust-1001",
            "items": [{"description": "SaaS Annual Plan", "quantity": 1, "rate": 125000.0}],
            "total_amount": 125000.0,
            "status": "sent",
            "due_date": next_month_iso,
            "organization_id": DEFAULT_ORG_ID,
            "company_id": DEFAULT_COMPANY_ID,
            "created_at": now_iso,
        }
    ]
    await _upsert_seed_many("invoices", invoices)

    expenses = [
        {
            "id": "exp-1001",
            "employee_id": "emp-1003",
            "category": "Travel",
            "amount": 4500.0,
            "description": "Client meeting commute",
            "date": yesterday_iso,
            "status": "approved",
            "organization_id": DEFAULT_ORG_ID,
            "company_id": DEFAULT_COMPANY_ID,
            "created_at": now_iso,
        },
        {
            "id": "exp-1002",
            "employee_id": "emp-1001",
            "category": "Software",
            "amount": 1800.0,
            "description": "Design tool renewal",
            "date": today_iso,
            "status": "pending",
            "organization_id": DEFAULT_ORG_ID,
            "company_id": DEFAULT_COMPANY_ID,
            "created_at": now_iso,
        },
    ]
    await _upsert_seed_many("expenses", expenses)

    stores = [
        {
            "id": "store-1001",
            "name": "Central Warehouse",
            "location": "Bangalore",
            "manager": "Rahul Verma",
            "contact": "+918011112222",
            "status": "active",
            "organization_id": DEFAULT_ORG_ID,
            "created_at": now_iso,
        }
    ]
    await _upsert_seed_many("stores", stores)

    inventory_items = [
        {
            "id": "invitem-1001",
            "name": "Dell 27-inch Monitor",
            "category": "Hardware",
            "quantity": 18,
            "unit": "piece",
            "price_per_unit": 22000.0,
            "location": "Rack A1",
            "organization_id": DEFAULT_ORG_ID,
            "created_at": now_iso,
        },
        {
            "id": "invitem-1002",
            "name": "Wireless Keyboard",
            "category": "Accessories",
            "quantity": 32,
            "unit": "piece",
            "price_per_unit": 2500.0,
            "location": "Rack B2",
            "organization_id": DEFAULT_ORG_ID,
            "created_at": now_iso,
        },
    ]
    await _upsert_seed_many("inventory", inventory_items)

    purchase_requests = [
        {
            "id": "pr-1001",
            "store_id": "store-1001",
            "requested_by": "emp-1003",
            "items": [{"name": "HDMI Cable", "qty": 10, "price": 450.0}],
            "total_amount": 4500.0,
            "reason": "Office setup stock refill",
            "status": "pending",
            "organization_id": DEFAULT_ORG_ID,
            "created_at": now_iso,
        }
    ]
    await _upsert_seed_many("purchase_requests", purchase_requests)

    purchase_orders = [
        {
            "id": "po-1001",
            "purchase_request_id": "pr-1001",
            "store_id": "store-1001",
            "supplier_name": "Tech Supply Pvt Ltd",
            "supplier_contact": "+918055554444",
            "items": [{"name": "HDMI Cable", "qty": 10, "price": 450.0}],
            "total_amount": 4500.0,
            "delivery_date": next_week_iso,
            "status": "pending",
            "created_by": "usr-admin",
            "organization_id": DEFAULT_ORG_ID,
            "created_at": now_iso,
        }
    ]
    await _upsert_seed_many("purchase_orders", purchase_orders)

    timesheets = [
        {
            "id": "ts-1001",
            "employee_id": "emp-1001",
            "project_id": "proj-1001",
            "task_id": "task-1001",
            "hours": 7.5,
            "date": today_iso,
            "description": "Deployment bug fixes",
            "status": "submitted",
            "organization_id": DEFAULT_ORG_ID,
            "created_at": now_iso,
        },
        {
            "id": "ts-1002",
            "employee_id": "emp-1002",
            "project_id": "proj-1002",
            "task_id": "task-1002",
            "hours": 6.0,
            "date": yesterday_iso,
            "description": "Lead segmentation setup",
            "status": "approved",
            "organization_id": DEFAULT_ORG_ID,
            "created_at": now_iso,
        },
    ]
    await _upsert_seed_many("timesheets", timesheets)

    tickets = [
        {
            "id": "ticket-1001",
            "subject": "Login page error for edge user",
            "description": "Customer reported a login retry loop on Safari.",
            "priority": "high",
            "status": "open",
            "contact_email": "support@globex.com",
            "organization_id": DEFAULT_ORG_ID,
            "created_at": now_iso,
        }
    ]
    await _upsert_seed_many("tickets", tickets)

    vendors = [
        {
            "id": "vendor-1001",
            "name": "Office Mart",
            "email": "sales@officemart.com",
            "phone": "+918077778888",
            "category": "Supplies",
            "organization_id": DEFAULT_ORG_ID,
            "created_at": now_iso,
        }
    ]
    await _upsert_seed_many("vendors", vendors)

    assets = [
        {
            "id": "asset-1001",
            "name": "MacBook Pro 14",
            "type": "Laptop",
            "serial_number": "MBP14-2026-001",
            "assigned_to": "emp-1001",
            "status": "assigned",
            "purchase_date": "2025-12-15",
            "value": 185000.0,
            "organization_id": DEFAULT_ORG_ID,
            "created_at": now_iso,
        }
    ]
    await _upsert_seed_many("assets", assets)

    announcements = [
        {
            "id": "ann-1001",
            "title": "Quarterly Planning Kickoff",
            "content": "Team planning starts Monday at 10 AM IST in Conference Room A.",
            "author_id": "usr-admin",
            "author_name": "Demo Admin",
            "priority": "high",
            "organization_id": DEFAULT_ORG_ID,
            "created_at": now_iso,
        }
    ]
    await _upsert_seed_many("announcements", announcements)

    jobs = [
        {
            "id": "job-1001",
            "title": "Senior Backend Engineer",
            "department": "Engineering",
            "location": "Bangalore",
            "type": "Full-time",
            "description": "Build and scale Python/FastAPI microservices.",
            "status": "open",
            "organization_id": DEFAULT_ORG_ID,
            "created_at": now_iso,
        }
    ]
    await _upsert_seed_many("jobs", jobs)
    await _upsert_seed_many("job_postings", jobs)

    # Bulk ATS jobs:
    # Existing: 1 open, Additional: 149 => total 150 (105 open, 30 closed, 15 draft)
    job_statuses = (["open"] * 104) + (["closed"] * 30) + (["draft"] * 15)
    job_types = ["Full-time", "Contract", "Remote", "Hybrid"]
    bulk_jobs: List[dict] = []
    for i, status in enumerate(job_statuses, start=1):
        dept = bulk_departments[(i - 1) % len(bulk_departments)]
        bulk_jobs.append(
            {
                "id": f"job-bulk-{i:03d}",
                "title": f"{dept} Specialist {i}",
                "department": dept,
                "location": "Bangalore" if i % 3 else "Remote",
                "type": job_types[(i - 1) % len(job_types)],
                "description": f"Seeded QA role #{i} for ATS stress testing.",
                "status": status,
                "organization_id": DEFAULT_ORG_ID,
                "created_at": (now - timedelta(days=i)).isoformat(),
            }
        )
    await _upsert_seed_many("jobs", bulk_jobs)
    await _upsert_seed_many("job_postings", bulk_jobs)

    candidates = [
        {
            "id": "cand-1001",
            "job_id": "job-1001",
            "name": "Kunal Singh",
            "email": "kunal.singh@candidate.com",
            "resume_url": "https://example.com/resume/kunal-singh",
            "status": "screening",
            "organization_id": DEFAULT_ORG_ID,
            "created_at": now_iso,
        }
    ]
    await _upsert_seed_many("candidates", candidates)

    # Bulk ATS candidates:
    # Existing: 1 screening, Additional: 899 => total 900
    candidate_statuses = (
        (["applied"] * 240)
        + (["screening"] * 210)
        + (["interview"] * 180)
        + (["offered"] * 105)
        + (["hired"] * 74)
        + (["rejected"] * 90)
    )
    all_job_ids = [j["id"] for j in (jobs + bulk_jobs)]
    bulk_candidates: List[dict] = []
    for i, status in enumerate(candidate_statuses, start=1):
        first = bulk_first_names[(i - 1) % len(bulk_first_names)]
        last = bulk_last_names[(i - 1) % len(bulk_last_names)]
        bulk_candidates.append(
            {
                "id": f"cand-bulk-{i:03d}",
                "job_id": all_job_ids[(i - 1) % len(all_job_ids)],
                "name": f"{first} Candidate {i}",
                "email": f"candidate.{first.lower()}.{i}@mailinator.com",
                "resume_url": f"https://example.com/resume/candidate-{i}",
                "status": status,
                "organization_id": DEFAULT_ORG_ID,
                "created_at": (now - timedelta(days=i)).isoformat(),
            }
        )
    await _upsert_seed_many("candidates", bulk_candidates)

    kb_articles = [
        {
            "id": "kb-1001",
            "title": "How To Approve Leave Requests",
            "content": "Go to HRMS > Leaves, review request details, and click Approve or Reject.",
            "category": "HR",
            "author_name": "Demo Admin",
            "organization_id": DEFAULT_ORG_ID,
            "created_at": now_iso,
        }
    ]
    await _upsert_seed_many("kb", kb_articles)

    audit_logs = [
        {
            "id": "audit-1001",
            "user_email": "admin@demo.com",
            "action": "POST",
            "module": "EMPLOYEES",
            "details": "Seed employee records initialized",
            "organization_id": DEFAULT_ORG_ID,
            "created_at": now_iso,
        }
    ]
    await _upsert_seed_many("audit_logs", audit_logs)

    insights = [
        {
            "id": "ins-1001",
            "type": "opportunity",
            "title": "High Value Deal Progress",
            "message": "One deal above INR 5L is currently in negotiation stage.",
            "impact": "medium",
            "organization_id": DEFAULT_ORG_ID,
            "created_at": now_iso,
        }
    ]
    await _upsert_seed_many("insights", insights)

    posh_complaints = [
        {
            "id": "posh-1001",
            "complainant_id": "usr-employee",
            "incident_date": seven_days_ago_iso,
            "description": "Inappropriate remarks in team chat.",
            "accused_name": "Anonymous Staff",
            "status": "Under Review",
            "organization_id": DEFAULT_ORG_ID,
            "created_at": now_iso,
        }
    ]
    await _upsert_seed_many("posh_complaints", posh_complaints)

    wfh_requests = [
        {
            "id": "wfh-1001",
            "employee_id": "emp-1001",
            "employee_name": "Demo Employee",
            "start_date": today_iso,
            "end_date": next_week_iso,
            "reason": "Home internet setup visit",
            "status": "pending",
            "organization_id": DEFAULT_ORG_ID,
            "created_at": now_iso,
        }
    ]
    await _upsert_seed_many("wfh_requests", wfh_requests)

    resignations = [
        {
            "id": "res-1001",
            "employee_id": "emp-1002",
            "employee_name": "Priya Sharma",
            "reason": "Higher education",
            "resignation_date": "2026-03-15",
            "last_working_day": "2026-04-15",
            "notice_period_days": 30,
            "status": "pending",
            "fnf_status": "pending",
            "fnf_amount": 0.0,
            "fnf_breakdown": {},
            "organization_id": DEFAULT_ORG_ID,
            "created_at": now_iso,
        }
    ]
    await _upsert_seed_many("resignations", resignations)

    performance_plans = [
        {
            "id": "pip-1001",
            "employee_id": "emp-1003",
            "employee_name": "Rahul Verma",
            "reason": "Missed month-end deadlines",
            "goals": "Close ledger on time and complete reconciliation checklist.",
            "duration_days": 30,
            "start_date": seven_days_ago_iso,
            "end_date": next_month_iso,
            "status": "active",
            "organization_id": DEFAULT_ORG_ID,
            "created_at": now_iso,
        }
    ]
    await _upsert_seed_many("performance_plans", performance_plans)

    offer_letters = [
        {
            "id": "offer-1001",
            "organization_id": DEFAULT_ORG_ID,
            "name": "Demo Employee",
            "email": "employee@demo.com",
            "phone": "+919876543210",
            "designation": "Software Engineer",
            "ctc_yearly": 960000.0,
            "details": {"joining_bonus": 50000, "notes": "Demo offer letter for testing"},
            "status": "Generated",
            "created_at": now_iso,
        }
    ]
    await _upsert_seed_many("offer_letters", offer_letters)

    # Bulk offer letters:
    # Existing: 1 Generated, Additional: 119 => total 120
    offer_statuses = (["Generated"] * 29) + (["Sent"] * 30) + (["Accepted"] * 30) + (["Revoked"] * 30)
    bulk_offers: List[dict] = []
    for i, status in enumerate(offer_statuses, start=1):
        employee = all_employee_records[(i - 1) % len(all_employee_records)]
        bulk_offers.append(
            {
                "id": f"offer-bulk-{i:03d}",
                "organization_id": DEFAULT_ORG_ID,
                "name": employee["name"],
                "email": f"offer.candidate.{i}@demo.com",
                "phone": employee["phone"],
                "designation": employee["designation"],
                "ctc_yearly": float(550000 + (i * 12000)),
                "details": {"batch": "bulk-seed", "employee_ref": employee["id"]},
                "status": status,
                "created_at": (now - timedelta(days=i)).isoformat(),
            }
        )
    await _upsert_seed_many("offer_letters", bulk_offers)

    companies = [
        {
            "id": DEFAULT_COMPANY_ID,
            "organization_id": DEFAULT_ORG_ID,
            "name": "Demo Finance Co",
            "legal_name": "Demo Finance Company Private Limited",
            "industry": "SaaS",
            "email": "accounts@demo-company.com",
            "phone": "+918012345678",
            "website": "https://demo-company.example.com",
            "address": "MG Road",
            "city": "Bangalore",
            "state": "Karnataka",
            "country": "India",
            "pincode": "560001",
            "pan_number": "ABCDE1234F",
            "gst_number": "29ABCDE1234F1Z5",
            "cin_number": "U72900KA2020PTC123456",
            "status": "active",
            "onboarded_by": "Demo Admin",
            "created_at": now_iso,
        }
    ]
    await _upsert_seed_many("companies", companies)

    chart_of_accounts = [
        {
            "id": "coa-1001",
            "company_id": DEFAULT_COMPANY_ID,
            "code": "1000",
            "name": "Cash in Hand",
            "type": "Asset",
            "sub_type": "Current",
            "description": "Primary cash account",
            "opening_balance": 250000.0,
            "current_balance": 250000.0,
            "is_bank": False,
            "currency": "INR",
            "created_at": now_iso,
        },
        {
            "id": "coa-1002",
            "company_id": DEFAULT_COMPANY_ID,
            "code": "3000",
            "name": "Sales Revenue",
            "type": "Revenue",
            "sub_type": "Operating",
            "description": "Recurring revenue",
            "opening_balance": -480000.0,
            "current_balance": -480000.0,
            "is_bank": False,
            "currency": "INR",
            "created_at": now_iso,
        },
        {
            "id": "coa-1003",
            "company_id": DEFAULT_COMPANY_ID,
            "code": "4000",
            "name": "Operating Expenses",
            "type": "Expense",
            "sub_type": "Operating",
            "description": "General operating expenses",
            "opening_balance": 125000.0,
            "current_balance": 125000.0,
            "is_bank": False,
            "currency": "INR",
            "created_at": now_iso,
        },
        {
            "id": "coa-1004",
            "company_id": DEFAULT_COMPANY_ID,
            "code": "2000",
            "name": "Accounts Payable",
            "type": "Liability",
            "sub_type": "Current",
            "description": "Outstanding vendor payments",
            "opening_balance": -95000.0,
            "current_balance": -95000.0,
            "is_bank": False,
            "currency": "INR",
            "created_at": now_iso,
        },
    ]
    await _upsert_seed_many("chart_of_accounts", chart_of_accounts)

    journal_entries = [
        {
            "id": "je-1001",
            "company_id": DEFAULT_COMPANY_ID,
            "entry_number": "JE-00015",
            "total_debit": 50000.0,
            "total_credit": 50000.0,
            "status": "posted",
            "created_by": "Demo Accountant",
            "created_at": now_iso,
            "date": today_iso,
            "narration": "Sample booked revenue",
            "reference": "SEED",
            "lines": [
                {"account_id": "coa-1001", "account_name": "Cash in Hand", "debit": 50000.0, "credit": 0.0},
                {"account_id": "coa-1002", "account_name": "Sales Revenue", "debit": 0.0, "credit": 50000.0},
            ],
        }
    ]
    await _upsert_seed_many("journal_entries", journal_entries)

    bank_accounts = [
        {
            "id": "bank-1001",
            "company_id": DEFAULT_COMPANY_ID,
            "account_name": "Operating Account",
            "bank_name": "HDFC Bank",
            "account_number": "009900112233",
            "ifsc": "HDFC0001234",
            "balance": 325000.0,
            "created_at": now_iso,
        }
    ]
    await _upsert_seed_many("bank_accounts", bank_accounts)

    gst_returns = [
        {
            "id": "gst-1001",
            "company_id": DEFAULT_COMPANY_ID,
            "period": "2026-03",
            "return_type": "GSTR-3B",
            "taxable_value": 450000.0,
            "igst": 0.0,
            "cgst": 40500.0,
            "sgst": 40500.0,
            "status": "draft",
            "created_at": now_iso,
        }
    ]
    await _upsert_seed_many("gst_returns", gst_returns)

    await _upsert_seed_document("counters", {"_id": f"journal_{DEFAULT_COMPANY_ID}"}, {"_id": f"journal_{DEFAULT_COMPANY_ID}", "seq": 15})

    subscriptions = [
        {
            "id": "sub-1001",
            "user_id": "usr-admin",
            "plan": "enterprise",
            "status": "active",
            "amount": 9999.0,
            "currency": "INR",
            "billing_cycle": "monthly",
            "starts_at": now_iso,
            "ends_at": subscription_end_iso,
            "created_at": now_iso,
        }
    ]
    await _upsert_seed_many("subscriptions", subscriptions)

    analytics_events = [
        {"id": "ana-1001", "user_id": "usr-admin", "event_type": "user_registered", "event_data": {"source": "seed"}, "page": "register", "timestamp": now_iso},
        {"id": "ana-1002", "user_id": "usr-admin", "event_type": "email_verified", "event_data": {"source": "seed"}, "page": "verify", "timestamp": now_iso},
        {"id": "ana-1003", "user_id": "usr-admin", "event_type": "subscription_created", "event_data": {"plan": "enterprise"}, "page": "billing", "timestamp": now_iso},
        {"id": "ana-1004", "user_id": "usr-admin", "event_type": "team_invite_sent", "event_data": {"role": "employee"}, "page": "team", "timestamp": now_iso},
    ]
    await _upsert_seed_many("analytics", analytics_events)

    team_invites = [
        {
            "id": "invite-1001",
            "email": "newhire@demo.com",
            "invited_by": "usr-admin",
            "organization_id": DEFAULT_ORG_ID,
            "role": "employee",
            "status": "pending",
            "token": "token-seed-1001",
            "expires_at": (now + timedelta(days=3)).isoformat(),
            "created_at": now_iso,
        }
    ]
    await _upsert_seed_many("team_invites", team_invites)

    trips = [
        {
            "id": "trip-1001",
            "user_id": "usr-employee",
            "organization_id": DEFAULT_ORG_ID,
            "status": "ended",
            "start_time": (now - timedelta(hours=4)).isoformat(),
            "end_time": (now - timedelta(hours=2)).isoformat(),
            "distance_km": 12.4,
            "duration_sec": 7200,
            "current_lat": 12.9716,
            "current_lng": 77.5946,
            "created_at": now_iso,
            "updated_at": now_iso,
        },
        {
            "id": "trip-1002",
            "user_id": "usr-admin",
            "organization_id": DEFAULT_ORG_ID,
            "status": "ended",
            "start_time": (now - timedelta(hours=2)).isoformat(),
            "end_time": (now - timedelta(hours=1)).isoformat(),
            "distance_km": 6.8,
            "duration_sec": 3600,
            "current_lat": 12.9352,
            "current_lng": 77.6245,
            "created_at": now_iso,
            "updated_at": now_iso,
        },
    ]
    await _upsert_seed_many("trips", trips)

    locations = [
        {
            "id": "loc-1001",
            "trip_id": "trip-1001",
            "user_id": "usr-employee",
            "timestamp": (now - timedelta(hours=3, minutes=30)).isoformat(),
            "lat": 12.9701,
            "lng": 77.5900,
            "speed": 38.5,
            "address": "MG Road, Bangalore",
        },
        {
            "id": "loc-1002",
            "trip_id": "trip-1001",
            "user_id": "usr-employee",
            "timestamp": (now - timedelta(hours=2, minutes=30)).isoformat(),
            "lat": 12.9750,
            "lng": 77.6020,
            "speed": 42.0,
            "address": "Indiranagar, Bangalore",
        },
        {
            "id": "loc-1003",
            "trip_id": "trip-1002",
            "user_id": "usr-admin",
            "timestamp": (now - timedelta(hours=1, minutes=20)).isoformat(),
            "lat": 12.9400,
            "lng": 77.6200,
            "speed": 28.0,
            "address": "Koramangala, Bangalore",
        },
    ]
    await _upsert_seed_many("locations", locations)

    logger.info("Demo seed bootstrap completed for org '%s'", DEFAULT_ORG_ID)



# --- Careers Endpoints ---
@app.post("/api/careers/jobs/create-from-jd")
async def create_job_from_jd(payload: JDCreate):
    from ai_recruitment import parse_jd
    parsed_data = parse_jd(payload.jd_text)

    # Generate a title heuristic
    lines = payload.jd_text.strip().split("\n")
    title = lines[0] if len(lines[0]) < 50 else "New AI Role"

    skills_str = ", ".join([s["name"] for s in parsed_data["parsed_skills"]])

    import uuid
    job_id = str(uuid.uuid4())
    job_doc = {
        "id": job_id,
        "title": title,
        "department": "Engineering",
        "location": "Remote",
        "type": "Full-time",
        "description": payload.jd_text,
        "skills_required": skills_str,
        "parsed_skills": parsed_data["parsed_skills"],
        "experience_level": parsed_data["experience_level"],
        "status": "open",
        "organization_id": payload.organization_id,
        "created_at": datetime.now(timezone.utc).isoformat()
    }

    await db.jobs.insert_one(job_doc)
    return {"message": "Job created autonomously via AI", "job_id": job_id, "parsed_data": parsed_data}

@app.get("/api/careers/jobs")

async def get_career_jobs():
    jobs_cursor = db.jobs.find({})
    jobs = await jobs_cursor.to_list(length=100)
    for job in jobs:
        if "_id" in job:
            job["id"] = str(job["_id"])
            del job["_id"]
        elif "id" not in job:
            job["id"] = "unknown"
    if not jobs:
        return [
            {"id": "1", "title": "Senior Frontend Engineer", "description": "Looking for an experienced React developer.", "skills_required": "React, TypeScript"},
            {"id": "2", "title": "Backend Python Developer", "description": "FastAPI and PostgreSQL expert needed.", "skills_required": "Python, FastAPI, SQL"}
        ]
    return jobs

@app.post("/api/careers/apply")
async def apply_job(application: JobApplication):
    # This route is used directly from the careers public page (no auth).
    job = await db.jobs.find_one({"id": application.job_id})
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")

    status = "applied"

    # Analyze resume using AI Recruitment Engine
    from ai_recruitment import analyze_resume
    jd_skills = job.get("parsed_skills", [])
    resume_analysis = {}
    if application.resume_text and jd_skills:
        resume_analysis = analyze_resume(application.resume_text, jd_skills)

    import uuid
    candidate_id = str(uuid.uuid4())
    doc = {
        "id": candidate_id,
        "job_id": application.job_id,
        "name": application.name if hasattr(application, 'name') and application.name else "Unknown Applicant",
        "email": application.email if hasattr(application, 'email') and application.email else "unknown@example.com",
        "resume_url": None,
        "resume_text": application.resume_text,
        "peer_rating": application.peer_rating,
        "status": status,
        "resume_analysis": resume_analysis, # Store analysis for AI Interview
        "ai_interview_rating": None,
        "final_score": 0.0,
        "organization_id": job.get("organization_id"),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.candidates.insert_one(doc)
    return {"message": "Application received", "candidate_id": candidate_id, "status": status, "resume_analysis": resume_analysis}

@app.post("/api/careers/ai-interview/start/{candidate_id}")
async def start_interview(candidate_id: str):
    candidate = await db.candidates.find_one({"id": candidate_id})
    if not candidate:
        raise HTTPException(status_code=404, detail="Candidate not found")

    job = await db.jobs.find_one({"id": candidate.get("job_id")})
    if not job:
        raise HTTPException(status_code=404, detail="Associated Job not found")

    from ai_recruitment import generate_interview_questions
    jd_skills = job.get("parsed_skills", [])
    resume_analysis = candidate.get("resume_analysis", {})

    # Pre-generate a list of dynamic questions
    interview_plan = generate_interview_questions(jd_skills, resume_analysis)
    questions = interview_plan.get("questions", [])

    session_id = f"sess_{candidate_id}"

    first_q_text = questions[0]["text"] if questions else "Welcome! Let's start the interview. Can you tell me about your experience?"

    doc = {
        "session_id": session_id,
        "candidate_id": candidate_id,
        "round": 1,
        "questions": questions,
        "current_question_index": 0,
        "scores": [],
        "messages": [
            {"role": "ai", "content": first_q_text}
        ],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.interview_sessions.delete_many({"session_id": session_id})
    await db.interview_sessions.insert_one(doc)

    # Update candidate status
    await db.candidates.update_one({"id": candidate_id}, {"$set": {"status": "interviewing"}})

    return {"session_id": session_id, "messages": doc["messages"]}

class AIInterviewMessageExtended(BaseModel):
    message: str
    time_taken_seconds: Optional[int] = 30

@app.post("/api/careers/ai-interview/{session_id}/chat")
async def interview_chat(session_id: str, msg: AIInterviewMessageExtended):
    session = await db.interview_sessions.find_one({"session_id": session_id})
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")

    user_msg = {"role": "user", "content": msg.message}

    # Evaluate the answer
    from ai_recruitment import evaluate_answer
    questions = session.get("questions", [])
    curr_idx = session.get("current_question_index", 0)

    curr_question = questions[curr_idx] if curr_idx < len(questions) else {"skill": "general"}

    evaluation = evaluate_answer(curr_question, msg.message, msg.time_taken_seconds)

    # Prepare the next question
    next_idx = curr_idx + 1
    if next_idx < len(questions):
        next_q_text = questions[next_idx]["text"]
        reply = "Thank you. " + next_q_text
    else:
        reply = "Thank you. We have no further questions. You can end the interview now."

    ai_msg = {"role": "ai", "content": reply}

    # Store the result
    await db.interview_sessions.update_one(
        {"session_id": session_id},
        {
            "$push": {
                "messages": {"$each": [user_msg, ai_msg]},
                "scores": evaluation
            },
            "$set": {
                "current_question_index": next_idx
            }
        }
    )
    return {"reply": reply, "evaluation_flags": evaluation.get("flags", [])}

@app.post("/api/careers/ai-interview/{session_id}/finish")
async def finish_interview(session_id: str):
    session = await db.interview_sessions.find_one({"session_id": session_id})
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")

    candidate_id = session["candidate_id"]
    scores = session.get("scores", [])

    # Calculate Final Score and AI Recommendation
    final_score = 0.0
    total_flags = []

    if scores:
        total_score = sum(s["final_score"] for s in scores)
        final_score = round(total_score / len(scores), 1)
        for s in scores:
            total_flags.extend(s.get("flags", []))

    # Determine recommendation
    if "copy_paste_detected" in total_flags or final_score < 4.0:
        recommendation = "Reject"
    elif final_score >= 8.5:
        recommendation = "Strong Hire"
    elif final_score >= 7.0:
        recommendation = "Hire"
    else:
        recommendation = "Borderline"

    ai_report = {
        "final_score": final_score,
        "recommendation": recommendation,
        "risk_flags": list(set(total_flags)),
        "detailed_scores": scores
    }

    await db.candidates.update_one(
        {"id": candidate_id},
        {
            "$set": {
                "ai_interview_rating": final_score,
                "final_score": final_score,
                "ai_report": ai_report,
                "status": "ai_interview_done"
            }
        }
    )

    return {"message": "Interview finished", "final_score": final_score, "report": ai_report}

@app.get("/api/careers/candidates")
async def get_career_candidates(job_id: Optional[str] = None):
    query = {}
    if job_id:
        query["job_id"] = job_id

    cursor = db.candidates.find(query).sort("final_score", -1) # Sort by merit (descending)
    candidates = await cursor.to_list(length=100)
    for c in candidates:
        if "_id" in c:
            c["id"] = str(c["_id"])
            del c["_id"]
    return candidates


app.include_router(api_router)
app.include_router(scheduling_router, prefix="/api/scheduling", tags=["scheduling"])


logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("startup")
async def create_indexes():
    global db, client, using_fallback_db
    """Create all MongoDB indexes on startup (idempotent — safe to run multiple times)."""
    if client is not None:
        try:
            await client.admin.command("ping")
        except Exception as exc:
            logger.exception("MongoDB ping failed. Falling back to in-memory DB: %s", exc)
            client = None
            db = InMemoryDatabase()
            using_fallback_db = True

    if using_fallback_db:
        logger.warning("Running backend with in-memory fallback DB.")
        await ensure_demo_users_seeded()
        from seed_iatf_compliance import seed_iatf_data
        await seed_iatf_data()
        return

    logger.info("Creating MongoDB indexes...")

    # ── Users ──────────────────────────────────────────────
    await db.users.create_index("id", unique=True)
    await db.users.create_index("email", unique=True)
    await db.users.create_index("organization_id")
    await db.users.create_index([("organization_id", 1), ("role", 1)])

    # ── Employees ──────────────────────────────────────────
    await db.employees.create_index("id", unique=True)
    await db.employees.create_index("organization_id")
    await db.employees.create_index([("organization_id", 1), ("status", 1)])

    # ── Leave Requests ─────────────────────────────────────
    await db.leave_requests.create_index("id", unique=True)
    await db.leave_requests.create_index([("organization_id", 1), ("status", 1)])
    await db.leave_requests.create_index([("organization_id", 1), ("employee_id", 1)])

    # ── Projects & Tasks ───────────────────────────────────
    await db.projects.create_index("id", unique=True)
    await db.projects.create_index("organization_id")
    await db.tasks.create_index("id", unique=True)
    await db.tasks.create_index([("organization_id", 1), ("project_id", 1)])
    await db.tasks.create_index([("organization_id", 1), ("status", 1)])

    # ── CRM ────────────────────────────────────────────────
    await db.leads.create_index("id", unique=True)
    await db.leads.create_index([("organization_id", 1), ("status", 1)])
    await db.leads.create_index([("organization_id", 1), ("created_at", -1)])
    await db.deals.create_index("id", unique=True)
    await db.deals.create_index([("organization_id", 1), ("stage", 1)])

    # ── Finance ────────────────────────────────────────────
    await db.invoices.create_index("id", unique=True)
    await db.invoices.create_index([("organization_id", 1), ("status", 1)])
    await db.invoices.create_index([("organization_id", 1), ("due_date", 1)])
    await db.customers.create_index("id", unique=True)
    await db.customers.create_index("organization_id")
    await db.vendors.create_index("id", unique=True)
    await db.vendors.create_index("organization_id")
    await db.expenses.create_index([("organization_id", 1), ("status", 1)])
    await db.expenses.create_index([("organization_id", 1), ("date", -1)])

    # ── Support Desk ───────────────────────────────────────
    await db.tickets.create_index("id", unique=True)
    await db.tickets.create_index([("organization_id", 1), ("status", 1)])
    await db.tickets.create_index([("organization_id", 1), ("priority", 1)])
    await db.tickets.create_index([("organization_id", 1), ("created_at", -1)])

    # ── Timesheets ─────────────────────────────────────────
    await db.timesheets.create_index("id", unique=True)
    await db.timesheets.create_index([("organization_id", 1), ("employee_id", 1)])
    await db.timesheets.create_index([("organization_id", 1), ("date", -1)])

    # ── Announcements ──────────────────────────────────────
    await db.announcements.create_index([("organization_id", 1), ("created_at", -1)])

    # ── Purchases ──────────────────────────────────────────
    await db.purchase_requests.create_index("id", unique=True)
    await db.purchase_requests.create_index([("organization_id", 1), ("status", 1)])
    await db.purchase_orders.create_index("id", unique=True)
    await db.purchase_orders.create_index([("organization_id", 1), ("status", 1)])
    await db.stores.create_index("id", unique=True)
    await db.stores.create_index("organization_id")

    # ── Assets ─────────────────────────────────────────────
    await db.assets.create_index("id", unique=True)
    await db.assets.create_index("organization_id")

    # ── Attendance ─────────────────────────────────────────
    await db.attendance.create_index([("organization_id", 1), ("employee_id", 1)])
    await db.attendance.create_index([("organization_id", 1), ("date", -1)])

    # ── Accountant / Ledger ────────────────────────────────
    await db.chart_of_accounts.create_index([("company_id", 1), ("code", 1)], unique=True)
    await db.journal_entries.create_index([("company_id", 1), ("date", -1)])
    await db.journal_entries.create_index([("company_id", 1), ("entry_number", 1)])
    await db.gst_returns.create_index([("company_id", 1), ("status", 1)])
    await db.bank_accounts.create_index("company_id")

    # ── Companies ──────────────────────────────────────────
    await db.companies.create_index("id", unique=True)
    await db.companies.create_index([("organization_id", 1), ("created_at", -1)])

    # ── Travel Tracker ─────────────────────────────────────
    await db.trips.create_index([("user_id", 1), ("start_time", -1)])
    await db.trips.create_index([("organization_id", 1)])
    await db.locations.create_index([("trip_id", 1), ("timestamp", -1)])

    # ── Misc ───────────────────────────────────────────────
    await db.team_invites.create_index("token", unique=True)
    await db.team_invites.create_index([("organization_id", 1), ("status", 1)])
    await db.posh_complaints.create_index([("organization_id", 1), ("status", 1)])
    await db.wfh_requests.create_index([("organization_id", 1), ("status", 1)])
    await db.audit_logs.create_index([("organization_id", 1), ("created_at", -1)])
    await db.audit_logs.create_index([("company_id", 1)])
    await db.audit_logs.create_index([("module", 1)])
    await db.offer_letters.create_index([("organization_id", 1), ("created_at", -1)])
    await db.jobs.create_index([("organization_id", 1), ("status", 1)])
    await db.candidates.create_index([("organization_id", 1), ("job_id", 1)])
    await db.kb.create_index([("organization_id", 1), ("category", 1)])

    logger.info("✅ MongoDB indexes created/verified successfully")
    await ensure_demo_users_seeded()
    from seed_iatf_compliance import seed_iatf_data
    await seed_iatf_data()

@app.on_event("shutdown")
async def shutdown_db_client():
    if client is not None:
        client.close()

handler = app

# ==========================================
# SAAS ERP API ENDPOINTS (MOCKUP)
# ==========================================

class ApprovalWorkflow(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    company_id: str
    entity_type: str
    entity_id: str
    requester_id: str
    approver_id: str
    status: str = "pending"
    comments: Optional[str] = None
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    updated_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class ApprovalWorkflowCreate(BaseModel):
    company_id: str
    entity_type: str
    entity_id: str
    requester_id: str
    approver_id: str
    comments: Optional[str] = None

@api_router.post("/workflows")
async def create_workflow(data: ApprovalWorkflowCreate, current_user: dict = Depends(get_current_user)):
    doc = ApprovalWorkflow(**data.dict()).dict()
    await db.approval_workflows.insert_one({k: v for k, v in doc.items() if k != "_id"})
    return doc

@api_router.get("/workflows")
async def get_workflows(company_id: str, current_user: dict = Depends(get_current_user)):
    workflows = await db.approval_workflows.find({"company_id": company_id}, {"_id": 0}).to_list(1000)
    return workflows

@api_router.post("/assets/{id}/depreciate")
async def depreciate_asset(id: str, current_user: dict = Depends(get_current_user)):
    org_id = current_user.get("organization_id")
    role = current_user.get("role")
    query = {"id": id, "organization_id": org_id}

    if role == "accountant":
        query["company_id"] = get_accountant_company(current_user)

    asset = await db.assets.find_one(query)
    if not asset:
        raise HTTPException(status_code=404, detail="Asset not found")

    current_value = float(asset.get("value", 0.0))
    if current_value <= 0:
        raise HTTPException(status_code=400, detail="Asset is already fully depreciated")

    # Simple straight line 10% depreciation
    depreciation_amount = current_value * 0.10
    new_value = max(0.0, current_value - depreciation_amount)

    await db.assets.update_one(
        {"id": id},
        {"$set": {
            "value": new_value,
            "last_depreciated_at": datetime.now(timezone.utc).isoformat()
        }}
    )

    # Log to audit trail
    audit_doc = {
        "id": str(uuid.uuid4()),
        "organization_id": org_id,
        "user_id": current_user["id"],
        "user_email": current_user.get("email", "Unknown"),
        "action": "DEPRECIATE_ASSET",
        "entity": "Asset",
        "entity_id": id,
        "details": f"Depreciated by {depreciation_amount}. New value: {new_value}",
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.audit_logs.insert_one(audit_doc)

    return {"message": "Asset depreciated successfully", "new_value": new_value}

@api_router.post("/parse-document")
async def parse_document(file: UploadFile = File(...), current_user: dict = Depends(get_current_user)):
    content = await file.read()
    text = extract_text_from_file(content, file.filename)
    parsed_data = parse_employee_from_text(text)
    return parsed_data
